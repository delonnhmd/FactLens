# PHASE 4 STEP 1
# PHASE 4 STEP 2
# PHASE 4 STEP 3
# PHASE 4 STEP 4
# PHASE 4 STEP 5
# PHASE 4 STEP 5B
# PHASE 4 STEP 5C
# PHASE 4 STEP 5D
# PHASE 4 STEP 5F-2
# PHASE 4 STEP 7
# PHASE 4 STEP 8
# PHASE 4 STEP 9
# PHASE 4 STEP 10
# PHASE 4 STEP 17
# PHASE 4 STEP 18
# PHASE 4 STEP 27
import json
import hashlib
import hmac
import os
import re
import sys
from io import BytesIO
from datetime import datetime, timedelta, timezone
from html import escape
from pathlib import Path
from time import monotonic
from typing import Any, Literal
from urllib.parse import quote, urlparse
from urllib.request import Request as UrlRequest, urlopen
from uuid import UUID

from dotenv import load_dotenv
from fastapi import BackgroundTasks, FastAPI, File, Form, HTTPException, Request, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, HTMLResponse, JSONResponse
from pydantic import BaseModel

try:
    from services.openai_factcheck import (
        analyze_claim_with_openai,
        get_openai_model,
    )
    from services.ai_library_loader import load_verifact_ai_library
    from services.source_page_fetcher import fetch_source_page
    from services.source_credibility import score_source_url
    from services.embedding_service import (
        classify_claim_stance,
        generate_claim_embedding,
    )
    # CONTENT SAFETY (NEW, additive) — objectionable-content gate at submission.
    from services.content_safety import check_content_safety
    from services.content_safety import classify_for_gate
    from services.content_safety import get_content_safety_openai_status
    from services.citation_service import (
        score_citation_source,
        validate_citation,
        verify_citation_exists,
    )
    # PHASE 6 STEP 3 — AI SEO tagging (NEW)
    from services.seo_service import (
        fetch_latest_claim_seo,
        generate_claim_seo,
    )
    # PHASE 6 STEP 4 — Topic clustering (NEW)
    from services.topic_cluster_service import (
        CLUSTER_MATCH_THRESHOLD,
        fetch_topic_row,
        find_or_create_topic_cluster,
        update_cluster_stats,
    )
    # VERDICT FORMULA v1 (NEW)
    from services.verdict_engine import (
        compute_verdict,
        map_verdict_to_claim_status,
    )
except ModuleNotFoundError:  # Allows repo-root command: uvicorn backend.main:app
    from backend.services.openai_factcheck import (
        analyze_claim_with_openai,
        get_openai_model,
    )
    from backend.services.ai_library_loader import load_verifact_ai_library
    from backend.services.source_page_fetcher import fetch_source_page
    from backend.services.source_credibility import score_source_url
    from backend.services.embedding_service import (
        classify_claim_stance,
        generate_claim_embedding,
    )
    # CONTENT SAFETY (NEW, additive) — objectionable-content gate at submission.
    from backend.services.content_safety import check_content_safety
    from backend.services.content_safety import classify_for_gate
    from backend.services.content_safety import get_content_safety_openai_status
    from backend.services.citation_service import (
        score_citation_source,
        validate_citation,
        verify_citation_exists,
    )
    # PHASE 6 STEP 3 — AI SEO tagging (NEW)
    from backend.services.seo_service import (
        fetch_latest_claim_seo,
        generate_claim_seo,
    )
    # PHASE 6 STEP 4 — Topic clustering (NEW)
    from backend.services.topic_cluster_service import (
        CLUSTER_MATCH_THRESHOLD,
        fetch_topic_row,
        find_or_create_topic_cluster,
        update_cluster_stats,
    )
    # VERDICT FORMULA v1 (NEW)
    from backend.services.verdict_engine import (
        compute_verdict,
        map_verdict_to_claim_status,
    )


load_dotenv()
# PHASE 4 STEP 27
app = FastAPI(title="Verifact backend", docs_url=None, redoc_url=None, openapi_url=None)

DEFAULT_CORS_ALLOWED_ORIGINS = (
    "https://factfight.com",
    "https://www.factfight.com",
    # Preserved while the legacy Verifact web entry point remains available.
    "https://verifact.pennyfloat.com",
    "http://localhost:3000",
    "http://127.0.0.1:3000",
)


def get_cors_allowed_origins() -> list[str]:
    configured_origins = os.environ.get("CORS_ALLOWED_ORIGINS", "")
    if not configured_origins.strip():
        return list(DEFAULT_CORS_ALLOWED_ORIGINS)

    origins: list[str] = []
    for value in configured_origins.split(","):
        origin = value.strip().rstrip("/")
        parsed = urlparse(origin)
        if (
            parsed.scheme in {"http", "https"}
            and parsed.netloc
            and not any((parsed.path, parsed.params, parsed.query, parsed.fragment))
            and parsed.username is None
            and parsed.password is None
        ):
            origins.append(origin)

    return origins or list(DEFAULT_CORS_ALLOWED_ORIGINS)

PUBLIC_SITE_URL = "https://factfight.com"
FALLBACK_SUPABASE_URL = "https://islcxqkevxxopatqvlqz.supabase.co"
FALLBACK_SUPABASE_ANON_KEY = (
    "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9."
    "eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6ImlzbGN4cWtldnh4b3BhdHF2bHF6Iiwicm9sZSI6ImFub24iLCJpYXQiOjE3Nzk3NDE5ODksImV4cCI6MjA5NTMxNzk4OX0."
    "96zZEPyRz2_RLkKJTx1GJIzQ-E1EcGA1X82FLPohTlg"
)
SUPABASE_PUBLIC_URL = (
    os.environ.get("EXPO_PUBLIC_SUPABASE_URL")
    or os.environ.get("SUPABASE_URL")
    or FALLBACK_SUPABASE_URL
).rstrip("/")
SUPABASE_PUBLIC_ANON_KEY = os.environ.get("EXPO_PUBLIC_SUPABASE_ANON_KEY") or FALLBACK_SUPABASE_ANON_KEY

AUTH_CALLBACK_HTML = """
<!DOCTYPE html>
<html lang="en">
  <head>
    <meta charset="utf-8" />
    <meta name="viewport" content="width=device-width, initial-scale=1" />
    <meta name="description" content="Verifact email verification callback." />
    <title>Verifact Email Verification</title>
    <style>
      :root {
        --ink: #172033;
        --muted: #667085;
        --border: #e4e7ec;
        --surface: #ffffff;
        --soft: #f5f7fa;
        --navy: #0d1b3e;
        --danger: #e24b4a;
        --danger-soft: #fcebeb;
        --success-soft: #e1f5ee;
      }

      * {
        box-sizing: border-box;
      }

      body {
        align-items: center;
        background: var(--soft);
        color: var(--ink);
        display: flex;
        font-family: Inter, ui-sans-serif, system-ui, -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
        justify-content: center;
        line-height: 1.5;
        margin: 0;
        min-height: 100vh;
        padding: 20px;
      }

      main {
        background: var(--surface);
        border: 1px solid var(--border);
        border-radius: 18px;
        max-width: 480px;
        padding: 28px;
        text-align: center;
        width: 100%;
      }

      .mark {
        align-items: center;
        background: var(--success-soft);
        border-radius: 999px;
        color: var(--navy);
        display: inline-flex;
        font-size: 28px;
        font-weight: 800;
        height: 68px;
        justify-content: center;
        margin-bottom: 18px;
        width: 68px;
      }

      main[data-state="error"] .mark {
        background: var(--danger-soft);
        color: var(--danger);
      }

      h1 {
        color: var(--navy);
        font-size: 24px;
        letter-spacing: 0;
        line-height: 1.2;
        margin: 0 0 10px;
      }

      p {
        color: var(--muted);
        font-size: 15px;
        margin: 0 0 18px;
      }

      .actions {
        display: grid;
        gap: 10px;
      }

      a {
        align-items: center;
        border-radius: 8px;
        display: inline-flex;
        font-size: 14px;
        font-weight: 800;
        justify-content: center;
        min-height: 46px;
        padding: 11px 14px;
        text-decoration: none;
      }

      .primary {
        background: var(--navy);
        color: #ffffff;
      }

      .secondary {
        background: #ffffff;
        border: 1px solid var(--border);
        color: var(--navy);
      }

      .small {
        color: #98a2b3;
        font-size: 12px;
        margin: 16px 0 0;
      }
    </style>
  </head>
  <body>
    <main id="card" data-state="success">
      <div id="mark" class="mark">OK</div>
      <h1 id="title">Email verified successfully</h1>
      <p id="copy">Your account has been verified. You may now close this page and return to the Verifact app to sign in.</p>
    </main>
    <script>
      const search = window.location.search || "";
      const hash = window.location.hash || "";
      const hashParams = new URLSearchParams(hash.replace(/^#/, ""));
      const searchParams = new URLSearchParams(search.replace(/^\\?/, ""));
      const error =
        hashParams.get("error_description") ||
        searchParams.get("error_description") ||
        hashParams.get("error") ||
        searchParams.get("error") ||
        "";
      const card = document.getElementById("card");
      const mark = document.getElementById("mark");
      const title = document.getElementById("title");
      const copy = document.getElementById("copy");

      if (window.history && window.history.replaceState) {
        window.history.replaceState(null, "", error ? "/auth/callback" : "/auth/confirmed");
      }

      if (error) {
        card.dataset.state = "error";
        mark.textContent = "!";
        title.textContent = "Verification link problem";
        copy.textContent = "This verification link could not be used. Request a new verification email from the Verifact app and try again.";
      }
    </script>
  </body>
</html>
"""

RESET_PASSWORD_HTML = """
<!DOCTYPE html>
<html lang="en">
  <head>
    <meta charset="utf-8" />
    <meta name="viewport" content="width=device-width, initial-scale=1" />
    <meta name="robots" content="noindex, nofollow" />
    <meta name="description" content="Reset your Verifact password." />
    <title>Reset your password | Verifact</title>
    <style>
      :root {
        --ink: #172033;
        --muted: #667085;
        --border: #e4e7ec;
        --surface: #ffffff;
        --soft: #f5f7fa;
        --navy: #0d1b3e;
        --danger: #e24b4a;
        --success: #1d9e75;
        --success-soft: #e1f5ee;
      }

      * {
        box-sizing: border-box;
      }

      body {
        align-items: center;
        background:
          radial-gradient(circle at top left, rgba(29, 158, 117, 0.12), transparent 28%),
          linear-gradient(180deg, #ffffff 0%, var(--soft) 100%);
        color: var(--ink);
        display: flex;
        font-family: Inter, ui-sans-serif, system-ui, -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
        justify-content: center;
        line-height: 1.5;
        margin: 0;
        min-height: 100vh;
        padding: 20px;
      }

      main {
        background: var(--surface);
        border: 1px solid var(--border);
        border-radius: 18px;
        max-width: 480px;
        padding: 28px;
        width: 100%;
      }

      .brand {
        align-items: center;
        display: flex;
        gap: 12px;
        margin-bottom: 24px;
      }

      .mark {
        align-items: center;
        background: var(--navy);
        border-radius: 12px;
        color: #ffffff;
        display: flex;
        font-size: 15px;
        font-weight: 800;
        height: 42px;
        justify-content: center;
        width: 42px;
      }

      .brand-name {
        color: var(--navy);
        font-size: 21px;
        font-weight: 800;
        line-height: 1.1;
      }

      .brand-line {
        color: var(--muted);
        font-size: 13px;
        margin-top: 2px;
      }

      h1 {
        color: var(--navy);
        font-size: 26px;
        letter-spacing: 0;
        line-height: 1.2;
        margin: 0 0 10px;
      }

      p {
        color: var(--muted);
        font-size: 15px;
        margin: 0 0 18px;
      }

      form {
        display: grid;
        gap: 14px;
      }

      label {
        color: var(--ink);
        display: grid;
        font-size: 13px;
        font-weight: 800;
        gap: 7px;
      }

      input {
        border: 1px solid var(--border);
        border-radius: 8px;
        color: var(--ink);
        font: inherit;
        min-height: 48px;
        padding: 11px 12px;
        width: 100%;
      }

      input:focus {
        border-color: var(--navy);
        outline: 3px solid rgba(13, 27, 62, 0.12);
      }

      button,
      a.button {
        align-items: center;
        background: var(--navy);
        border: 0;
        border-radius: 8px;
        color: #ffffff;
        cursor: pointer;
        display: inline-flex;
        font: inherit;
        font-size: 14px;
        font-weight: 800;
        justify-content: center;
        min-height: 48px;
        padding: 12px 16px;
        text-decoration: none;
      }

      button:disabled {
        background: #98a2b3;
        cursor: not-allowed;
      }

      .feedback {
        border-radius: 10px;
        display: none;
        font-size: 14px;
        font-weight: 700;
        margin: 18px 0 0;
        padding: 12px;
      }

      .feedback[data-kind="error"],
      .feedback[data-kind="success"] {
        display: block;
      }

      .feedback[data-kind="error"] {
        background: #fcebeb;
        color: var(--danger);
      }

      .feedback[data-kind="success"] {
        background: var(--success-soft);
        color: var(--success);
      }

      .success-actions {
        display: none;
        margin-top: 18px;
      }

      main[data-state="success"] .success-actions {
        display: grid;
      }

      main[data-state="success"] form,
      main[data-state="invalid"] form {
        display: none;
      }
    </style>
  </head>
  <body>
    <main id="card" data-state="loading">
      <div class="brand" aria-label="Verifact">
        <div class="mark">V</div>
        <div>
          <div class="brand-name">Verifact</div>
          <div class="brand-line">The red. The blue. The truth.</div>
        </div>
      </div>

      <h1>Reset your password</h1>
      <p id="intro">Enter a new password for your Verifact account.</p>

      <form id="reset-form" hidden>
        <label>
          New password
          <input id="new-password" type="password" autocomplete="new-password" minlength="8" required />
        </label>

        <label>
          Confirm password
          <input id="confirm-password" type="password" autocomplete="new-password" minlength="8" required />
        </label>

        <button id="submit-button" type="submit">Update password</button>
      </form>

      <div id="feedback" class="feedback" role="status" aria-live="polite"></div>
    </main>

    <script src="https://cdn.jsdelivr.net/npm/@supabase/supabase-js@2"></script>
    <script>
      const SUPABASE_URL = __SUPABASE_PUBLIC_URL__;
      const SUPABASE_ANON_KEY = __SUPABASE_PUBLIC_ANON_KEY__;
      const card = document.getElementById("card");
      const form = document.getElementById("reset-form");
      const feedback = document.getElementById("feedback");
      const intro = document.getElementById("intro");
      const newPassword = document.getElementById("new-password");
      const confirmPassword = document.getElementById("confirm-password");
      const submitButton = document.getElementById("submit-button");

      function setFeedback(message, kind) {
        feedback.textContent = message;
        feedback.dataset.kind = kind;
      }

      function showInvalidLink() {
        card.dataset.state = "invalid";
        intro.textContent = "This reset link could not be used. Request a new password reset email from Verifact and try again.";
        setFeedback("We could not verify this reset link. Please request a new password reset email.", "error");
      }

      async function initializeResetSession() {
        const hashParams = new URLSearchParams(window.location.hash.replace(/^#/, ""));
        const accessToken = hashParams.get("access_token") || "";
        const refreshToken = hashParams.get("refresh_token") || "";
        const type = hashParams.get("type") || "";

        if (!accessToken || !refreshToken || type !== "recovery") {
          showInvalidLink();
          return;
        }

        if (!window.supabase || !window.supabase.createClient) {
          card.dataset.state = "invalid";
          setFeedback("Password reset is temporarily unavailable. Please try again.", "error");
          return;
        }

        const supabaseClient = window.supabase.createClient(SUPABASE_URL, SUPABASE_ANON_KEY);
        let sessionResult;

        try {
          sessionResult = await supabaseClient.auth.setSession({
            access_token: accessToken,
            refresh_token: refreshToken,
          });
        } catch {
          showInvalidLink();
          return;
        }

        if (sessionResult.error) {
          showInvalidLink();
          return;
        }

        if (window.history && window.history.replaceState) {
          window.history.replaceState(null, "", "/reset-password");
        }

        card.dataset.state = "ready";
        form.hidden = false;

        form.addEventListener("submit", async (event) => {
          event.preventDefault();
          setFeedback("", "");

          const nextPassword = newPassword.value;
          const confirmedPassword = confirmPassword.value;

          if (nextPassword.length < 8) {
            setFeedback("Password must be at least 8 characters.", "error");
            return;
          }

          if (nextPassword !== confirmedPassword) {
            setFeedback("Passwords must match.", "error");
            return;
          }

          submitButton.disabled = true;
          submitButton.textContent = "Updating...";

          try {
            const { error: updateError } = await supabaseClient.auth.updateUser({ password: nextPassword });

            if (updateError) {
              setFeedback("We could not update your password. Please try again.", "error");
              return;
            }
          } catch {
            setFeedback("We could not update your password. Please try again.", "error");
            return;
          } finally {
            submitButton.disabled = false;
            submitButton.textContent = "Update password";
          }

          card.dataset.state = "success";
          intro.textContent = "";
          setFeedback("Password updated successfully. You can now return to the Verifact app to sign in.", "success");
        });
      }

      initializeResetSession();
    </script>
  </body>
</html>
""".replace("__SUPABASE_PUBLIC_URL__", json.dumps(SUPABASE_PUBLIC_URL)).replace(
    "__SUPABASE_PUBLIC_ANON_KEY__",
    json.dumps(SUPABASE_PUBLIC_ANON_KEY),
)

LEGAL_PAGE_STYLE = """
  :root {
    --ink: #172033;
    --muted: #667085;
    --border: #e4e7ec;
    --surface: #ffffff;
    --soft: #f5f7fa;
    --navy: #0d1b3e;
    --blue: #185fa5;
  }

  * {
    box-sizing: border-box;
  }

  body {
    background: var(--soft);
    color: var(--ink);
    font-family: Inter, ui-sans-serif, system-ui, -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
    line-height: 1.6;
    margin: 0;
  }

  .page {
    margin: 0 auto;
    max-width: 850px;
    min-height: 100vh;
    padding: 28px 20px 44px;
  }

  header,
  footer {
    align-items: center;
    display: flex;
    flex-wrap: wrap;
    gap: 12px;
    justify-content: space-between;
  }

  .brand {
    color: var(--navy);
    font-size: 20px;
    font-weight: 800;
    text-decoration: none;
  }

  nav {
    display: flex;
    flex-wrap: wrap;
    gap: 12px;
  }

  a {
    color: var(--blue);
    font-weight: 700;
    text-decoration: none;
  }

  a:hover {
    text-decoration: underline;
  }

  main {
    background: var(--surface);
    border: 1px solid var(--border);
    border-radius: 14px;
    margin: 28px 0;
    padding: 28px;
  }

  h1 {
    color: var(--navy);
    font-size: clamp(32px, 8vw, 48px);
    letter-spacing: 0;
    line-height: 1.05;
    margin: 0 0 10px;
  }

  h2 {
    color: var(--navy);
    font-size: 20px;
    margin: 30px 0 10px;
  }

  p,
  li {
    color: var(--muted);
    font-size: 15px;
  }

  ul {
    padding-left: 22px;
  }

  .updated {
    color: var(--muted);
    font-size: 13px;
    font-weight: 800;
    margin-top: 0;
  }

  footer {
    color: var(--muted);
    font-size: 14px;
  }

  @media (max-width: 560px) {
    .page {
      padding: 20px 14px 36px;
    }

    main {
      border-radius: 10px;
      padding: 22px;
    }

    header,
    footer {
      align-items: flex-start;
      flex-direction: column;
    }
  }
"""


def build_legal_page(title: str, description: str, content: str) -> str:
    return f"""<!DOCTYPE html>
<html lang="en">
  <head>
    <meta charset="utf-8" />
    <meta name="viewport" content="width=device-width, initial-scale=1" />
    <meta name="robots" content="index, follow" />
    <meta name="description" content="{description}" />
    <title>{title} | Verifact</title>
    <style>{LEGAL_PAGE_STYLE}</style>
  </head>
  <body>
    <div class="page">
      <header>
        <a class="brand" href="/">Verifact</a>
        <nav aria-label="Legal pages">
          <a href="/about">About</a>
          <a href="/privacy">Privacy</a>
          <a href="/personal-privacy">Personal Privacy</a>
          <a href="/terms">Terms</a>
          <a href="/copyright">Copyright</a>
          <a href="/community-guidelines">Guidelines</a>
        </nav>
      </header>
      <main>
        {content}
      </main>
      <footer>
        <span>&copy; 2026 PennyFloat</span>
        <a href="mailto:support@factfight.com">support@factfight.com</a>
      </footer>
    </div>
  </body>
</html>"""


ABOUT_PAGE_STYLE = """
  :root {
    --navy: #0d1b3e;
    --navy-2: #12306f;
    --surface: #ffffff;
    --soft: #f5f7fa;
    --ink: #172033;
    --muted: #667085;
    --border: #e4e7ec;
    --green: #1d9e75;
    --red: #e24b4a;
    --amber: #ef9f27;
    --purple: #534ab7;
  }

  * {
    box-sizing: border-box;
  }

  body {
    background: var(--soft);
    color: var(--ink);
    font-family: ui-sans-serif, system-ui, -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
    line-height: 1.6;
    margin: 0;
  }

  a {
    color: inherit;
  }

  .hero {
    background: linear-gradient(135deg, var(--navy), var(--navy-2));
    color: #ffffff;
    padding: 34px 20px 42px;
  }

  .container {
    margin: 0 auto;
    max-width: 850px;
    width: 100%;
  }

  .nav {
    align-items: center;
    display: flex;
    flex-wrap: wrap;
    gap: 14px;
    justify-content: space-between;
    margin-bottom: 44px;
  }

  .brand {
    color: #ffffff;
    font-size: 22px;
    font-weight: 700;
    text-decoration: none;
  }

  .nav-links {
    display: flex;
    flex-wrap: wrap;
    gap: 12px;
  }

  .nav-links a {
    color: #dbe7ff;
    font-size: 14px;
    font-weight: 600;
    text-decoration: none;
  }

  .nav-links a:hover {
    color: #ffffff;
    text-decoration: underline;
  }

  .badge {
    background: rgba(255, 255, 255, 0.12);
    border: 1px solid rgba(255, 255, 255, 0.22);
    border-radius: 999px;
    color: #e7eefc;
    display: inline-flex;
    font-size: 14px;
    font-weight: 600;
    margin-bottom: 18px;
    padding: 7px 12px;
  }

  h1 {
    color: #ffffff;
    font-size: clamp(42px, 10vw, 72px);
    letter-spacing: 0;
    line-height: 0.95;
    margin: 0 0 12px;
  }

  .slogan {
    color: #ffffff;
    font-size: clamp(22px, 6vw, 34px);
    font-weight: 700;
    line-height: 1.15;
    margin: 0 0 16px;
  }

  .intro {
    color: #dbe7ff;
    font-size: 18px;
    margin: 0;
    max-width: 720px;
  }

  main {
    padding: 34px 20px 0;
  }

  .card {
    background: var(--surface);
    border: 1px solid var(--border);
    border-radius: 16px;
    margin-bottom: 18px;
    padding: 24px;
  }

  h2 {
    color: var(--navy);
    font-size: 24px;
    line-height: 1.2;
    margin: 0 0 10px;
  }

  p {
    color: var(--muted);
    font-size: 16px;
    margin: 0 0 14px;
  }

  p:last-child {
    margin-bottom: 0;
  }

  ul {
    margin: 0;
    padding-left: 22px;
  }

  li {
    color: var(--muted);
    font-size: 16px;
    margin: 8px 0;
  }

  .pill-row {
    display: flex;
    flex-wrap: wrap;
    gap: 8px;
    margin-top: 14px;
  }

  .pill {
    border-radius: 999px;
    color: #ffffff;
    display: inline-flex;
    font-size: 13px;
    font-weight: 700;
    padding: 6px 10px;
  }

  .green {
    background: var(--green);
  }

  .red {
    background: var(--red);
  }

  .amber {
    background: var(--amber);
  }

  .purple {
    background: var(--purple);
  }

  .contact {
    background: #eef4ff;
    border-color: #d8e4ff;
  }

  .contact a {
    color: var(--navy);
    font-weight: 700;
  }

  footer {
    color: var(--muted);
    font-size: 14px;
    margin: 28px auto 0;
    max-width: 850px;
    padding: 22px 20px 34px;
  }

  @media (max-width: 560px) {
    .hero {
      padding-top: 24px;
    }

    .nav {
      align-items: flex-start;
      flex-direction: column;
      margin-bottom: 34px;
    }

    main {
      padding-top: 24px;
    }

    .card {
      border-radius: 12px;
      padding: 20px;
    }
  }
"""


ABOUT_PAGE_HTML = f"""<!DOCTYPE html>
<html lang="en">
  <head>
    <meta charset="utf-8" />
    <meta name="viewport" content="width=device-width, initial-scale=1" />
    <meta name="robots" content="index, follow" />
    <meta
      name="description"
      content="About Verifact, a community-powered claim verification app with AI-assisted source review and community voting."
    />
    <title>About Verifact | The red. The blue. The truth.</title>
    <style>{ABOUT_PAGE_STYLE}</style>
  </head>
  <body>
    <header class="hero">
      <div class="container">
        <nav class="nav" aria-label="Verifact pages">
          <a class="brand" href="/">Verifact</a>
          <div class="nav-links">
            <a href="/privacy">Privacy</a>
            <a href="/terms">Terms</a>
            <a href="/copyright">Copyright</a>
            <a href="/community-guidelines">Guidelines</a>
          </div>
        </nav>
        <span class="badge">Owned and operated by PennyFloat</span>
        <h1>About Verifact</h1>
        <p class="slogan">The red. The blue. The truth.</p>
        <p class="intro">Verifact is a community-powered claim verification app built to help people review claims, sources, and evidence with more transparency.</p>
      </div>
    </header>

    <main>
      <div class="container">
        <section class="card">
          <h2>What Verifact is</h2>
          <p>Verifact helps users submit claims, review source links, add evidence, and vote on whether a claim appears True, Fake, or Not Sure.</p>
          <p>The product is designed for community-powered claim verification. It is not an official fact-checking authority, and it does not promise guaranteed truth.</p>
        </section>

        <section class="card">
          <h2>How the review process works</h2>
          <ul>
            <li>Users submit claims with source links or context.</li>
            <li>Community members can add evidence and review what supports or challenges a claim.</li>
            <li>AI-assisted source review can help summarize source quality and risk signals.</li>
            <li>AI does not make the final decision. Human and community voting remains the deciding signal in Verifact's review flow.</li>
          </ul>
          <div class="pill-row" aria-label="Verifact review signals">
            <span class="pill green">True</span>
            <span class="pill red">Fake</span>
            <span class="pill amber">Not Sure</span>
            <span class="pill purple">AI-assisted review</span>
          </div>
        </section>

        <section class="card">
          <h2>Neutral by design</h2>
          <p>Verifact is built for political neutrality. The same rules apply to claims from any person, party, organization, source, or viewpoint.</p>
          <p>The goal is to make the review process more transparent: show the claim, show the evidence, show the source context, and let the community participate.</p>
        </section>

        <section class="card contact">
          <h2>Contact</h2>
          <p>Verifact is owned and operated by PennyFloat.</p>
          <p>For support, app review questions, or public page requests, contact <a href="mailto:support@factfight.com">support@factfight.com</a>.</p>
        </section>
      </div>
    </main>

    <footer>
      <div class="container">&copy; 2026 PennyFloat</div>
    </footer>
  </body>
</html>"""


PRIVACY_POLICY_HTML = build_legal_page(
    "FactFight Privacy Policy",
    "Privacy Policy for FactFight, operated by MD Media LLC.",
    """
        <h1>Privacy Policy</h1>
        <p class="updated">Last updated July 18, 2026 &middot; Operated by MD Media LLC (Houston, Texas)</p>
        <p>FactFight is a community-powered claim verification service, available through our mobile application and website at factfight.com, operated by MD Media LLC ("FactFight," "we," "us"). This policy explains how we collect, use, store, share, and protect information for accounts, claims, evidence, voting, reports, moderation, and AI-assisted analysis.</p>
        <p><strong>We do not sell your personal information.</strong></p>

        <h2>1. Information We Collect</h2>
        <ul>
          <li>Account information: email address, username, display name, avatar, authentication identifiers, and account status.</li>
          <li>User-generated content: claim text, source URLs, uploaded images and their metadata, evidence notes, votes, reports, and comments.</li>
          <li>Moderation, safety, reputation, and trust signals: report status, visibility status, vote history, badges, and account-action records.</li>
          <li>Technical information: IP-derived network data, device or browser details, app version, diagnostics, cookies, local storage, and service logs.</li>
        </ul>

        <h2>2. How We Use Information</h2>
        <ul>
          <li>To create and authenticate accounts, maintain profiles and reputation, provide support, and process account-deletion requests.</li>
          <li>To publish and operate claims, evidence, images, votes, reports, and public contributor pages.</li>
          <li>To run moderation and safety systems, investigate abuse, enforce policies, prevent spam and fraud, and protect service integrity.</li>
          <li>To operate diagnostics, reliability monitoring, security checks, and product improvements.</li>
        </ul>
        <p>We do not use your personal information for third-party advertising, and we do not sell it.</p>

        <h2>3. AI-Assisted Analysis</h2>
        <p>We use AI-assisted systems (including OpenAI) to classify, summarize, moderate, and evaluate claims, source URLs, uploaded evidence, and related context, and to detect prohibited content. To do this, we may send claim text, source URLs, and related content to our AI provider for processing. AI outputs are preliminary risk signals that may be wrong, incomplete, outdated, or biased, and never make the final community verdict.</p>

        <h2>4. Service Providers (Sub-Processors)</h2>
        <p>We share information with a limited set of service providers that process it only to provide services to us, under confidentiality and security obligations, and not for their own marketing or resale:</p>
        <ul>
          <li>Supabase — authentication, database, and file storage.</li>
          <li>Render — backend and API hosting.</li>
          <li>Vercel — website hosting.</li>
          <li>OpenAI — AI-assisted analysis and content moderation.</li>
          <li>Resend — transactional and account email delivery.</li>
          <li>Expo, Apple, and Google — mobile app delivery and distribution.</li>
        </ul>

        <h2>5. Legal Requests and Disclosures</h2>
        <p>We do not sell your personal information. However, we may disclose information when required to do so by valid legal process — including a lawful court order, subpoena, or government request — or where we believe in good faith that disclosure is necessary to comply with the law, protect the rights, property, or safety of FactFight, our users, or the public, or to prevent imminent harm. In responding to any such request, we disclose only the information legally required and no more. Where permitted by law, we will make reasonable efforts to notify the affected user before disclosing their information, unless we are legally prohibited from doing so or the request relates to an emergency.</p>

        <h2>6. Your Privacy Rights</h2>
        <p>Depending on where you live, you may have rights over your personal information, including the rights to access, correct, delete, or receive a copy of it, and to object to or restrict certain processing.</p>
        <p><strong>California (CCPA/CPRA):</strong> You have the right to know what personal information we collect, to access and delete it, and to opt out of the sale or sharing of personal information. We do not sell or share your personal information as those terms are defined under California law. We will not discriminate against you for exercising your rights.</p>
        <p><strong>EEA/UK (GDPR):</strong> You have the rights of access, rectification, erasure, restriction, portability, and objection, and the right to lodge a complaint with your data protection authority. We process personal information to provide the service, to comply with law, and for our legitimate interests in safety, security, and integrity.</p>
        <p>To exercise any right, contact support@factfight.com from the email address associated with your account. We may need to verify your identity before acting on a request.</p>

        <h2>7. Data Retention</h2>
        <p>We retain personal information for as long as needed to operate the service, comply with law, resolve disputes, prevent abuse, and preserve verification history. Public contributions such as claims, evidence, and votes may be retained in anonymized form after account deletion to preserve the integrity of the verification record.</p>

        <h2>8. Account Deletion</h2>
        <p>You may request account deletion from within the app or by contacting support@factfight.com from the email address associated with your account. When you delete your account, your personal information is removed, but your public contributions may remain, attributed to an anonymous user, to preserve verification history. Some records may be retained or anonymized when needed for safety, legal compliance, fraud prevention, or dispute resolution.</p>

        <h2>9. Children's Privacy</h2>
        <p>FactFight is not directed to children under 13, and you must be at least 13 years old to use it. We do not knowingly collect personal information from children under 13. If we learn that we have collected personal information from a child under 13, we will delete it. If you believe a child under 13 has provided us information, contact support@factfight.com.</p>

        <h2>10. Public Content</h2>
        <p>Claims, evidence, votes, usernames, public profile details, and reputation signals may be visible to other users when submitted to public or community-facing features. Do not submit private personal information in public claims or evidence.</p>

        <h2>11. Changes to This Policy</h2>
        <p>We may update this policy. Material changes will be communicated through the Service. Continued use after changes constitutes acceptance.</p>

        <h2>12. Contact</h2>
        <p>For privacy questions, requests, or to exercise your rights, contact <a href="mailto:support@factfight.com">support@factfight.com</a>. MD Media LLC, Houston, Texas.</p>
    """,
)

PERSONAL_PRIVACY_HTML = build_legal_page(
    "FactFight Personal Privacy",
    "Personal privacy promise for FactFight, operated by MD Media LLC.",
    """
        <h1>Personal Privacy</h1>
        <p class="updated">Last updated July 18, 2026 &middot; Operated by MD Media LLC (Houston, Texas)</p>
        <p>FactFight and MD Media LLC take personal user privacy seriously. We do not sell, rent, trade, or voluntarily provide personal user information to unrelated third parties for marketing, advertising, data brokerage, or commercial resale.</p>

        <h2>Our Privacy Promise</h2>
        <p>We will not provide personal user information to a third party unless a limited exception applies. Limited exceptions include a valid court order, subpoena, warrant, legally binding government request, regulatory requirement, user consent, or a situation where disclosure is necessary to protect users, investigate abuse, prevent fraud, secure the service, or comply with applicable law.</p>

        <h2>Service Providers</h2>
        <p>FactFight uses trusted service providers — including Supabase (authentication, database, storage), Render (backend hosting), Vercel (website hosting), OpenAI (AI-assisted analysis and moderation), Resend (email delivery), and Expo, Apple, and Google (app delivery) — to operate the service. These providers may process limited information only as needed to provide services to FactFight, and are not allowed to use personal user information for their own marketing or resale.</p>

        <h2>Legal Requests and Disclosures</h2>
        <p>We may disclose information when required to do so by valid legal process — including a lawful court order, subpoena, or government request — or where we believe in good faith that disclosure is necessary to comply with the law, protect the rights, property, or safety of FactFight, our users, or the public, or to prevent imminent harm. In responding to any such request, we disclose only the information legally required and no more. Where permitted by law, we will make reasonable efforts to notify the affected user before disclosing their information, unless we are legally prohibited from doing so or the request relates to an emergency.</p>

        <h2>Public Content</h2>
        <p>Claims, evidence, votes, reports, usernames, profile details, and reputation signals may be visible inside FactFight when users submit them to public or community-facing features. This page does not make public submissions private.</p>

        <h2>Contact</h2>
        <p>Questions about personal privacy may be sent to <a href="mailto:support@factfight.com">support@factfight.com</a>.</p>
    """,
)

TERMS_OF_SERVICE_HTML = build_legal_page(
    "FactFight Terms of Use",
    "Terms of Use for FactFight, operated by MD Media LLC.",
    """
        <h1>Terms of Use</h1>
        <p class="updated">Last updated July 18, 2026 &middot; Operated by MD Media LLC (Houston, Texas)</p>

        <h2>1. Acceptance of These Terms</h2>
        <p>By creating an account, logging in, or otherwise using FactFight (the "Service"), whether through our mobile application or website at factfight.com, you agree to these Terms of Use ("Terms"). If you do not agree, do not use the Service. If you are using the Service on behalf of an organization, you represent that you have authority to bind that organization.</p>
        <p>You must be at least 13 years old to use FactFight.</p>

        <h2>2. What FactFight Is — and Is Not</h2>
        <p>FactFight is a platform that allows users to post claims, submit evidence, vote, and participate in community verification of statements, including statements about political and public matters. FactFight provides tools for users to evaluate claims collectively, assisted by automated analysis.</p>
        <p>FactFight is not a publisher, journalist, or arbiter of truth. FactFight does not author claims. Claims, evidence, comments, and votes are created and submitted by users. Verdicts, scores, and labels displayed on the Service are the automated and aggregated output of community voting and AI-assisted analysis — they are informational only, may be incorrect, and do not represent statements of fact, opinion, or endorsement by FactFight or MD Media LLC.</p>

        <h2>3. User-Generated Content and Section 230</h2>
        <p>The Service hosts content provided by users. Under Section 230 of the Communications Decency Act (47 U.S.C. § 230), FactFight is a provider of an interactive computer service and is not the publisher or speaker of user-provided content. You — not FactFight — are solely responsible for the claims, evidence, and other content you submit.</p>
        <p>FactFight does not endorse, guarantee, or assume responsibility for any user content. Any reliance you place on claims or verdicts is at your own risk.</p>

        <h2>4. Verdicts, Scores, and Accuracy — Disclaimer</h2>
        <p>Verdicts (including labels such as "True," "Fake," "Disputed," or similar), source-quality scores, accuracy scores, and AI-generated analysis are produced by a combination of community voting and automated systems. These outputs:</p>
        <ul>
          <li>May be inaccurate, incomplete, outdated, or wrong;</li>
          <li>Are not professional, legal, financial, medical, or journalistic advice;</li>
          <li>Are not statements of fact by FactFight;</li>
          <li>Should never be relied upon as the sole basis for any decision.</li>
        </ul>
        <p>Always consult original sources and qualified professionals. FactFight expressly disclaims liability for any action taken in reliance on any verdict, score, or content on the Service.</p>

        <h2>5. Prohibited Content and Conduct — Zero Tolerance</h2>
        <p>FactFight has zero tolerance for objectionable content and abusive users. You may not post, and you may not use the Service to distribute, content that:</p>
        <ul>
          <li>Constitutes harassment, threats, incitement to violence, or hate speech;</li>
          <li>Is defamatory, libelous, or knowingly false about an identifiable person;</li>
          <li>Is sexually explicit, or exploits or endangers minors;</li>
          <li>Infringes intellectual property or privacy rights;</li>
          <li>Constitutes spam, fraud, impersonation, or manipulation of votes or verdicts;</li>
          <li>Violates any applicable law, including election, campaign-finance, and defamation law.</li>
        </ul>
        <p>FactFight reviews reported content and acts on violations within 24 hours, including removing content and terminating accounts. FactFight may remove any content or suspend or terminate any account at its sole discretion.</p>

        <h2>6. Integrity of the Service</h2>
        <p>You may not: operate multiple accounts to influence outcomes; use bots or automated means to vote or post; fabricate sources or evidence; manipulate reputation, scores, or verdicts; or interfere with the Service's operation. Violations are grounds for immediate termination.</p>

        <h2>7. Your Content — License and Responsibility</h2>
        <p>You retain ownership of content you submit. You grant FactFight a worldwide, non-exclusive, royalty-free license to host, display, reproduce, and distribute your content in connection with operating and promoting the Service.</p>
        <p>Claims become part of the public record. You may delete a claim within three (3) hours of posting. After three hours, or once a verdict is finalized, claims are permanent and may not be deleted. If you delete your account, your personal information is removed but your public contributions remain, attributed to an anonymous user.</p>
        <p>You represent that you have all rights necessary to submit your content and that it does not violate these Terms or any law.</p>

        <h2>8. Indemnification</h2>
        <p>You agree to indemnify, defend, and hold harmless FactFight, MD Media LLC, and their officers, members, employees, and agents from any claims, damages, liabilities, losses, and expenses (including reasonable attorneys' fees) arising from: (a) your content; (b) your use of the Service; (c) your violation of these Terms; or (d) your violation of any law or third-party right.</p>

        <h2>9. Disclaimer of Warranties</h2>
        <p>THE SERVICE IS PROVIDED "AS IS" AND "AS AVAILABLE," WITHOUT WARRANTIES OF ANY KIND, EXPRESS OR IMPLIED, INCLUDING MERCHANTABILITY, FITNESS FOR A PARTICULAR PURPOSE, ACCURACY, AND NON-INFRINGEMENT. FACTFIGHT DOES NOT WARRANT THAT THE SERVICE OR ANY VERDICT WILL BE ACCURATE, UNINTERRUPTED, OR ERROR-FREE.</p>

        <h2>10. Limitation of Liability</h2>
        <p>TO THE MAXIMUM EXTENT PERMITTED BY LAW, FACTFIGHT AND MD MEDIA LLC WILL NOT BE LIABLE FOR ANY INDIRECT, INCIDENTAL, SPECIAL, CONSEQUENTIAL, OR PUNITIVE DAMAGES, OR ANY LOSS OF PROFITS OR DATA, ARISING FROM YOUR USE OF THE SERVICE. OUR TOTAL LIABILITY FOR ANY CLAIM WILL NOT EXCEED THE GREATER OF (A) THE AMOUNT YOU PAID US IN THE TWELVE MONTHS BEFORE THE CLAIM, OR (B) ONE HUNDRED U.S. DOLLARS ($100).</p>

        <h2>11. Dispute Resolution and Arbitration</h2>
        <p>Any dispute arising from these Terms or the Service will be resolved by binding individual arbitration administered by the American Arbitration Association (AAA) under its applicable Consumer Arbitration Rules, seated in Houston, Texas. You and FactFight waive any right to a jury trial or to participate in a class action, except where prohibited by law.</p>

        <h2>12. Governing Law</h2>
        <p>These Terms are governed by the laws of the State of Texas, without regard to conflict-of-laws principles.</p>

        <h2>13. DMCA / Copyright</h2>
        <p>FactFight responds to notices of alleged copyright infringement under the Digital Millennium Copyright Act. To submit a notice, contact our designated agent at <a href="mailto:support@factfight.com">support@factfight.com</a>.</p>

        <h2>14. Changes to These Terms</h2>
        <p>We may update these Terms. Material changes will be communicated through the Service. Continued use after changes constitutes acceptance.</p>

        <h2>15. Termination</h2>
        <p>We may suspend or terminate your access at any time, with or without notice, for any violation of these Terms or for any reason at our discretion.</p>

        <h2>16. Contact</h2>
        <p>Questions about these Terms may be sent to <a href="mailto:support@factfight.com">support@factfight.com</a>. MD Media LLC, Houston, Texas.</p>
    """,
)

COPYRIGHT_NOTICE_HTML = build_legal_page(
    "Verifact Copyright Notice",
    "Copyright notice and intellectual property policy for Verifact by PennyFloat.",
    """
        <h1>Copyright Notice</h1>
        <p class="updated">Effective June 10, 2026</p>
        <p>&copy; 2026 PennyFloat. All rights reserved.</p>

        <h2>Verifact Ownership</h2>
        <p>Verifact, including its name, product experience, software, interface designs, workflows, analysis systems, documentation, and original PennyFloat content, is owned by PennyFloat or its licensors and protected by intellectual property laws.</p>

        <h2>User-Generated Content Licensing</h2>
        <p>You retain ownership of content you submit to Verifact, including claims, evidence, images, comments, reports, and related materials, subject to rights held by third parties. By submitting content, you grant PennyFloat a worldwide, non-exclusive, royalty-free license to host, store, reproduce, display, analyze, moderate, format, and distribute that content as needed to operate, protect, and improve Verifact.</p>

        <h2>Your Responsibility</h2>
        <p>You are responsible for ensuring that your submissions do not infringe copyrights, trademarks, privacy rights, publicity rights, confidentiality obligations, or other rights. Do not upload screenshots, articles, images, videos, documents, or other materials unless you have the right to share them.</p>

        <h2>DMCA and Copyright Contact Process</h2>
        <p>If you believe content on Verifact infringes your copyright, contact <a href="mailto:support@factfight.com">support@factfight.com</a> with your name, contact email, a description of the copyrighted work, the location of the allegedly infringing content, and a statement that you believe the use is unauthorized.</p>

        <h2>Repeat Infringement and Abuse</h2>
        <p>PennyFloat may remove or restrict content and may suspend or terminate accounts that repeatedly infringe intellectual property rights or misuse the copyright reporting process.</p>

        <h2>AI Analysis Disclaimer</h2>
        <p>AI-assisted analysis may process user-generated content to classify, summarize, or compare claims and evidence. AI analysis does not create a legal determination of ownership, infringement, fair use, authorization, or factual accuracy.</p>
    """,
)

COMMUNITY_GUIDELINES_HTML = build_legal_page(
    "Verifact Community Guidelines",
    "Community Guidelines for Verifact by PennyFloat.",
    """
        <h1>Community Guidelines</h1>
        <p class="updated">Effective June 10, 2026</p>
        <p>These Community Guidelines keep Verifact focused on useful claim and evidence review. They apply to claims, evidence, images, reports, votes, usernames, profiles, and any other user-submitted content.</p>

        <h2>Respectful Discussion</h2>
        <p>Discuss claims, evidence, and reasoning without attacking people. Strong disagreement is allowed; harassment, threats, intimidation, and targeted abuse are not.</p>

        <h2>Not Allowed</h2>
        <ul>
          <li>Harassment, threats, bullying, doxxing, exploitation, or coordinated abuse.</li>
          <li>Explicit sexual content, exploitative content, illegal content, graphic violence for shock value, or content that promotes self-harm or violence.</li>
          <li>Hate speech or attacks based on race, ethnicity, national origin, religion, sex, gender identity, sexual orientation, disability, age, caste, or veteran status.</li>
          <li>Spam, scams, malware, malicious links, repetitive content, fake engagement, or coordinated manipulation.</li>
          <li>Impersonation of another person, organization, public official, platform, moderator, or PennyFloat representative.</li>
        </ul>

        <h2>Evidence Quality Expectations</h2>
        <p>Submit evidence that is relevant to the claim, clearly sourced when possible, and not intentionally misleading. Avoid cropped, edited, or out-of-context media unless you clearly explain the edit or context.</p>

        <h2>Moderation Actions</h2>
        <p>Content may be removed for policy violations. Verifact may also reduce visibility, add labels, limit features, preserve records, suspend accounts, or report serious issues when content violates these Guidelines, Terms, platform rules, or applicable law.</p>

        <h2>Contact</h2>
        <p>Questions or reports may be sent to <a href="mailto:support@factfight.com">support@factfight.com</a>.</p>
    """,
)

PUBLIC_SITE_URL = (
    os.environ.get("FACTFIGHT_PUBLIC_SITE_URL")
    or os.environ.get("VERIFACT_PUBLIC_SITE_URL")
    or "https://factfight.com"
).rstrip("/")
VERIFACT_APP_STORE_URL = os.environ.get("VERIFACT_APP_STORE_URL", "").strip()
VERIFACT_GOOGLE_PLAY_URL = os.environ.get("VERIFACT_GOOGLE_PLAY_URL", "").strip()
VERIFACT_DEFAULT_OG_IMAGE_URL = f"{PUBLIC_SITE_URL}/assets/icon/icon.png"

CLAIM_PAGE_STYLE = """
  :root {
    --bg: #08111f;
    --panel: #0f1b2d;
    --panel-strong: #14253d;
    --ink: #f6f8fb;
    --muted: #aeb9c9;
    --soft: #d8e0ec;
    --border: rgba(255, 255, 255, 0.12);
    --blue: #5ea2ff;
    --red: #ff6b6b;
    --green: #33d39f;
    --gold: #f6c65b;
    --shadow: rgba(0, 0, 0, 0.35);
  }

  * {
    box-sizing: border-box;
  }

  body {
    background: var(--bg);
    color: var(--ink);
    font-family: Inter, ui-sans-serif, system-ui, -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
    line-height: 1.6;
    margin: 0;
  }

  a {
    color: inherit;
  }

  .page {
    margin: 0 auto;
    max-width: 800px;
    min-height: 100vh;
    padding: 24px 18px 44px;
    width: 100%;
  }

  .topbar,
  .footer {
    align-items: center;
    display: flex;
    flex-wrap: wrap;
    gap: 12px;
    justify-content: space-between;
  }

  .brand {
    align-items: center;
    color: var(--ink);
    display: inline-flex;
    font-size: 19px;
    font-weight: 850;
    gap: 10px;
    text-decoration: none;
  }

  .mark {
    align-items: center;
    background: var(--ink);
    border-radius: 8px;
    color: var(--bg);
    display: inline-flex;
    font-size: 12px;
    font-weight: 900;
    height: 30px;
    justify-content: center;
    width: 30px;
  }

  .slogan-small {
    color: var(--muted);
    font-size: 13px;
    font-weight: 800;
  }

  .hero,
  .card,
  .download-card {
    background: var(--panel);
    border: 1px solid var(--border);
    border-radius: 18px;
    box-shadow: 0 18px 48px var(--shadow);
  }

  .hero {
    margin-top: 24px;
    overflow: hidden;
  }

  .hero-inner {
    padding: 28px;
  }

  .eyebrow-row {
    display: flex;
    flex-wrap: wrap;
    gap: 8px;
    margin-bottom: 18px;
  }

  .eyebrow {
    border: 1px solid var(--border);
    border-radius: 999px;
    color: var(--soft);
    display: inline-flex;
    font-size: 12px;
    font-weight: 850;
    line-height: 1;
    padding: 8px 10px;
  }

  h1 {
    color: var(--ink);
    font-size: clamp(32px, 8vw, 56px);
    letter-spacing: 0;
    line-height: 1.02;
    margin: 0;
  }

  .description {
    color: var(--soft);
    font-size: 17px;
    margin: 16px 0 0;
  }

  .claim-image {
    background: #050a13;
    display: block;
    max-height: 430px;
    object-fit: cover;
    width: 100%;
  }

  .actions {
    display: flex;
    flex-wrap: wrap;
    gap: 10px;
    margin-top: 22px;
  }

  .button,
  .download-button {
    align-items: center;
    border-radius: 10px;
    display: inline-flex;
    font-size: 14px;
    font-weight: 850;
    justify-content: center;
    min-height: 46px;
    padding: 12px 15px;
    text-decoration: none;
  }

  .button.primary {
    background: var(--ink);
    color: var(--bg);
  }

  .button.secondary,
  .download-button {
    background: var(--panel-strong);
    border: 1px solid var(--border);
    color: var(--ink);
  }

  .grid {
    display: grid;
    gap: 14px;
    grid-template-columns: repeat(3, minmax(0, 1fr));
    margin-top: 14px;
  }

  .stat {
    background: var(--panel-strong);
    border: 1px solid var(--border);
    border-radius: 14px;
    padding: 15px;
  }

  .stat span {
    color: var(--muted);
    display: block;
    font-size: 12px;
    font-weight: 800;
    margin-bottom: 4px;
  }

  .stat strong {
    color: var(--ink);
    display: block;
    font-size: 24px;
    line-height: 1;
  }

  .bar {
    background: rgba(255, 255, 255, 0.08);
    border-radius: 999px;
    height: 8px;
    margin-top: 10px;
    overflow: hidden;
  }

  .bar-fill {
    border-radius: inherit;
    display: block;
    height: 100%;
    min-width: 3px;
  }

  .true {
    background: var(--green);
  }

  .fake {
    background: var(--red);
  }

  .unsure {
    background: var(--gold);
  }

  .content {
    display: grid;
    gap: 14px;
    margin-top: 14px;
  }

  .card,
  .download-card {
    padding: 22px;
  }

  h2 {
    color: var(--ink);
    font-size: 18px;
    letter-spacing: 0;
    line-height: 1.25;
    margin: 0 0 10px;
  }

  p {
    color: var(--muted);
    font-size: 15px;
    margin: 0;
  }

  .source-link {
    color: var(--blue);
    display: inline-flex;
    font-weight: 850;
    margin-top: 10px;
    overflow-wrap: anywhere;
    text-decoration: none;
  }

  .source-link:hover {
    text-decoration: underline;
  }

  .downloads {
    display: grid;
    gap: 10px;
    grid-template-columns: repeat(2, minmax(0, 1fr));
    margin-top: 16px;
  }

  .download-button.disabled {
    color: var(--muted);
  }

  .tagline {
    color: var(--ink);
    font-size: 20px;
    font-weight: 900;
    margin-top: 18px;
  }

  .footer {
    color: var(--muted);
    font-size: 13px;
    margin-top: 22px;
  }

  @media (max-width: 640px) {
    .page {
      padding: 18px 12px 34px;
    }

    .hero,
    .card,
    .download-card {
      border-radius: 14px;
    }

    .hero-inner,
    .card,
    .download-card {
      padding: 20px;
    }

    .actions,
    .downloads,
    .grid {
      grid-template-columns: 1fr;
    }

    .button,
    .download-button {
      width: 100%;
    }
  }
"""


def compact_text(value: object, fallback: str = "") -> str:
    text = fallback if value is None else str(value)
    return " ".join(text.split())


def html_attr(value: object, fallback: str = "") -> str:
    text = compact_text(value, fallback)
    return escape(" ".join(text.split()), quote=True)


def html_body(value: object, fallback: str = "") -> str:
    return escape(compact_text(value, fallback), quote=False)


def truncate_text(value: str, max_length: int) -> str:
    if len(value) <= max_length:
        return value

    return value[: max_length - 3].rstrip() + "..."


def to_int(value: object) -> int:
    try:
        return int(value or 0)
    except (TypeError, ValueError):
        return 0


def format_count(value: object) -> str:
    return f"{to_int(value):,}"


def normalize_public_url(url: object) -> str:
    value = str(url or "").strip()

    if not value:
        return ""

    parsed = urlparse(value)
    if parsed.scheme in {"http", "https"} and parsed.netloc:
        return value

    if value.startswith("/"):
        return f"{PUBLIC_SITE_URL}{value}"

    return f"https://{value}"


def get_url_host(url: str) -> str:
    try:
        return (urlparse(url).hostname or "Source").removeprefix("www.")
    except Exception:
        return "Source"


def get_claim_image_url(claim: dict) -> str:
    for key in ("image_url", "thumbnail_url"):
        image_url = normalize_public_url(claim.get(key))
        if image_url:
            return image_url

    return ""


def get_claim_meta_description(claim: dict) -> str:
    description = compact_text(claim.get("description"), "Review this claim on Verifact.")
    return truncate_text(description, 180)


def get_vote_percent(count: int, total: int) -> str:
    if total <= 0:
        return "0"

    return str(round((count / total) * 100, 1))


def is_uuid(value: str) -> bool:
    try:
        UUID(value)
        return True
    except ValueError:
        return False


def build_download_button(label: str, url: str) -> str:
    if url:
        return f'<a class="download-button" href="{html_attr(url)}" rel="noopener noreferrer" target="_blank">{html_body(label)}</a>'

    return f'<span class="download-button disabled" aria-disabled="true">{html_body(label)} - Coming soon</span>'


def build_vote_stat(label: str, count: int, total: int, color_class: str) -> str:
    return f"""
            <div class="stat">
              <span>{html_body(label)}</span>
              <strong>{format_count(count)}</strong>
              <div class="bar" aria-hidden="true"><span class="bar-fill {color_class}" style="width: {get_vote_percent(count, total)}%"></span></div>
            </div>"""


def build_public_claim_page(claim: dict) -> str:
    claim_id = str(claim.get("id") or "")
    encoded_claim_id = quote(claim_id, safe="")
    title = html_body(claim.get("title"), "Verifact claim")
    meta_title = truncate_text(compact_text(claim.get("title"), "Verifact claim"), 95)
    description = html_body(claim.get("description"), "No description provided.")
    meta_description = get_claim_meta_description(claim)
    category = html_body(claim.get("category"), "Other")
    ai_summary = html_body(
        claim.get("ai_summary") or claim.get("source_support_summary") or claim.get("ai_reason"),
        "AI review is still pending. Community voting and evidence are still available.",
    )
    source_url = normalize_public_url(claim.get("source_url"))
    source_host = html_body(get_url_host(source_url))
    source_link = ""
    if source_url:
        source_link = f'<a class="source-link" href="{html_attr(source_url)}" rel="noopener noreferrer" target="_blank">{source_host}</a>'

    votes_true = to_int(claim.get("votes_true"))
    votes_fake = to_int(claim.get("votes_fake"))
    votes_unsure = to_int(claim.get("votes_unsure"))
    total_votes = to_int(claim.get("total_votes")) or votes_true + votes_fake + votes_unsure
    image_url = get_claim_image_url(claim)
    og_image_url = image_url or VERIFACT_DEFAULT_OG_IMAGE_URL
    public_url = f"{PUBLIC_SITE_URL}/claim/{encoded_claim_id}"
    app_url = f"verifact://claim/{encoded_claim_id}"
    image_html = ""
    if image_url:
        image_html = f'<img class="claim-image" src="{html_attr(image_url)}" alt="Image for this Verifact claim" loading="lazy" />'

    return f"""<!DOCTYPE html>
<html lang="en">
  <head>
    <meta charset="utf-8" />
    <meta name="viewport" content="width=device-width, initial-scale=1" />
    <meta name="robots" content="index, follow" />
    <meta name="description" content="{html_attr(meta_description)}" />
    <meta property="og:type" content="article" />
    <meta property="og:site_name" content="Verifact" />
    <meta property="og:url" content="{html_attr(public_url)}" />
    <meta property="og:title" content="{html_attr(meta_title)}" />
    <meta property="og:description" content="{html_attr(meta_description)}" />
    <meta property="og:image" content="{html_attr(og_image_url)}" />
    <meta name="twitter:card" content="summary_large_image" />
    <meta name="twitter:title" content="{html_attr(meta_title)}" />
    <meta name="twitter:description" content="{html_attr(meta_description)}" />
    <meta name="twitter:image" content="{html_attr(og_image_url)}" />
    <title>{title} | Verifact</title>
    <style>{CLAIM_PAGE_STYLE}</style>
  </head>
  <body>
    <div class="page">
      <header class="topbar">
        <a class="brand" href="/"><span class="mark">VF</span><span>Verifact</span></a>
        <span class="slogan-small">The red. The blue. The truth.</span>
      </header>

      <main>
        <article class="hero">
          {image_html}
          <div class="hero-inner">
            <div class="eyebrow-row">
              <span class="eyebrow">{category}</span>
              <span class="eyebrow">{format_count(total_votes)} total votes</span>
            </div>
            <h1>{title}</h1>
            <p class="description">{description}</p>
            <div class="actions">
              <a class="button primary" href="{html_attr(app_url)}">Open in Verifact App</a>
              <a class="button secondary" href="#download">Download Verifact</a>
            </div>
          </div>
        </article>

        <section class="grid" aria-label="Vote totals">
          {build_vote_stat("True", votes_true, total_votes, "true")}
          {build_vote_stat("Fake", votes_fake, total_votes, "fake")}
          {build_vote_stat("Not sure", votes_unsure, total_votes, "unsure")}
        </section>

        <div class="content">
          <section class="card">
            <h2>AI Summary</h2>
            <p>{ai_summary}</p>
          </section>

          <section class="card">
            <h2>Source</h2>
            <p>Review the original source connected to this claim.</p>
            {source_link}
          </section>

          <section id="download" class="download-card">
            <h2>Download Verifact</h2>
            <p>Read the claim on web or open it in the app when Verifact is installed.</p>
            <div class="downloads" aria-label="Download Verifact">
              {build_download_button("App Store", VERIFACT_APP_STORE_URL)}
              {build_download_button("Google Play", VERIFACT_GOOGLE_PLAY_URL)}
            </div>
            <p class="tagline">The red. The blue. The truth.</p>
          </section>
        </div>
      </main>

      <footer class="footer">
        <span>&copy; 2026 PennyFloat</span>
        <a href="/privacy">Privacy</a>
      </footer>
    </div>
  </body>
</html>"""


def build_public_claim_404_page() -> str:
    description = "This Verifact claim could not be found or is no longer public."
    return f"""<!DOCTYPE html>
<html lang="en">
  <head>
    <meta charset="utf-8" />
    <meta name="viewport" content="width=device-width, initial-scale=1" />
    <meta name="robots" content="noindex, nofollow" />
    <meta name="description" content="{description}" />
    <meta property="og:title" content="Claim not found | Verifact" />
    <meta property="og:description" content="{description}" />
    <meta property="og:image" content="{html_attr(VERIFACT_DEFAULT_OG_IMAGE_URL)}" />
    <meta name="twitter:card" content="summary_large_image" />
    <title>Claim not found | Verifact</title>
    <style>{CLAIM_PAGE_STYLE}</style>
  </head>
  <body>
    <div class="page">
      <header class="topbar">
        <a class="brand" href="/"><span class="mark">VF</span><span>Verifact</span></a>
        <span class="slogan-small">The red. The blue. The truth.</span>
      </header>
      <main>
        <section class="hero">
          <div class="hero-inner">
            <div class="eyebrow-row"><span class="eyebrow">404</span></div>
            <h1>Claim not found</h1>
            <p class="description">{description}</p>
            <div class="actions">
              <a class="button primary" href="/">Go to Verifact</a>
            </div>
          </div>
        </section>
      </main>
    </div>
  </body>
</html>"""


def build_public_claim_error_page() -> str:
    description = "Verifact could not load this claim page right now. Please try again later."
    return f"""<!DOCTYPE html>
<html lang="en">
  <head>
    <meta charset="utf-8" />
    <meta name="viewport" content="width=device-width, initial-scale=1" />
    <meta name="robots" content="noindex, nofollow" />
    <meta name="description" content="{description}" />
    <meta property="og:title" content="Claim temporarily unavailable | Verifact" />
    <meta property="og:description" content="{description}" />
    <meta property="og:image" content="{html_attr(VERIFACT_DEFAULT_OG_IMAGE_URL)}" />
    <meta name="twitter:card" content="summary_large_image" />
    <title>Claim temporarily unavailable | Verifact</title>
    <style>{CLAIM_PAGE_STYLE}</style>
  </head>
  <body>
    <div class="page">
      <header class="topbar">
        <a class="brand" href="/"><span class="mark">VF</span><span>Verifact</span></a>
        <span class="slogan-small">The red. The blue. The truth.</span>
      </header>
      <main>
        <section class="hero">
          <div class="hero-inner">
            <div class="eyebrow-row"><span class="eyebrow">Unavailable</span></div>
            <h1>Claim temporarily unavailable</h1>
            <p class="description">{description}</p>
          </div>
        </section>
      </main>
    </div>
  </body>
</html>"""


def fetch_public_claim_row(claim_id: str) -> dict | None:
    supabase = get_supabase_client()
    response = supabase.table("claims").select("*").eq("id", claim_id).limit(1).execute()
    rows = response.data or []

    if isinstance(rows, dict):
        return rows

    if not rows:
        return None

    return rows[0]

# PHASE 4 STEP 20
# PHASE 4 STEP 20C
ALLOWED_SOURCE_QUALITIES = {
    "official",
    "mainstream",
    "specialized",
    "social",
    "blog",
    "unknown",
}


# PHASE 4 STEP 20C
def normalize_source_quality(value):
    if value is None:
        return "unknown"

    normalized_value = str(value).strip().lower()

    if normalized_value in ALLOWED_SOURCE_QUALITIES:
        return normalized_value

    return "unknown"


# PHASE 5 STEP 1B
RANK_ORDER = [
    "New Scout",
    "Claim Checker",
    "Trusted Verifier",
    "Source Hunter",
    "Verifact Guardian",
]


def calculate_trust_tier(trust_score):
    score = max(0, min(100, float(trust_score or 50)))
    if score >= 75:
        return "HIGH_TRUST"
    if score >= 55:
        return "TRUSTED"
    if score >= 30:
        return "BASIC"
    return "LOW_TRUST"


def calculate_rank_title(trust_score):
    score = max(0, min(100, float(trust_score or 50)))
    if score >= 90:
        return "Verifact Guardian"
    if score >= 75:
        return "Source Hunter"
    if score >= 55:
        return "Trusted Verifier"
    if score >= 30:
        return "Claim Checker"
    return "New Scout"


def calculate_vote_weight(trust_tier):
    weights = {
        "LOW_TRUST": 0.75,
        "BASIC": 1.0,
        "TRUSTED": 1.2,
        "HIGH_TRUST": 1.4,
    }
    return weights.get(str(trust_tier or "BASIC").upper(), 1.0)


def resolve_display_rank(current_rank, highest_rank_achieved):
    current_index = RANK_ORDER.index(current_rank) if current_rank in RANK_ORDER else 0
    highest_index = RANK_ORDER.index(highest_rank_achieved) if highest_rank_achieved in RANK_ORDER else 0
    return RANK_ORDER[max(current_index, highest_index)]

app.add_middleware(
    CORSMiddleware,
    allow_origins=get_cors_allowed_origins(),
    allow_credentials=False,
    allow_methods=["GET", "POST", "PATCH", "DELETE", "OPTIONS"],
    allow_headers=["Accept", "Authorization", "Content-Type"],
)


ClaimType = Literal["FACTUAL", "OPINION", "SATIRE", "QUESTION", "PROMOTION", "UNCLEAR"]
AiStatus = Literal[
    "PENDING",
    "LOW_RISK",
    "MEDIUM_RISK",
    "HIGH_RISK",
    "NEEDS_MORE_EVIDENCE",
    "NOT_FACT_CHECKABLE",
    "ERROR",
]


# PHASE 4 STEP 27
AI_RATE_LIMIT_WINDOW_SECONDS = 60 * 60
AI_RATE_LIMIT_MAX_REQUESTS = 30
AI_CLAIM_COOLDOWN_SECONDS = 45
SOURCE_SCORE_RATE_LIMIT_MAX_REQUESTS = 120
RATE_LIMIT_BUCKETS: dict[str, list[float]] = {}
CLAIM_AI_COOLDOWNS: dict[str, float] = {}
# Every configured rate-limit window and the AI cooldown are well under an
# hour; an entry untouched for longer than this is stale regardless of which
# bucket it belongs to.
_STALE_ENTRY_TTL_SECONDS = 3600
_rate_limit_sweep_counter = 0
# These dicts previously kept one entry per (bucket, IP) or per claim_id for
# the lifetime of the process with no eviction - a visitor who is never seen
# again still left its entry in memory forever. On a public app with many
# distinct IPs and an ever-growing set of claim IDs, that grows without
# bound until the process runs out of memory and Render kills/restarts it
# (which is also why a restart made things "work again": it reset these
# dicts to empty). Sweeping periodically bounds their size to roughly what's
# been active in the last hour instead of everything since the last restart.
_RATE_LIMIT_SWEEP_INTERVAL = 200

# PHASE 6 STEP 1 — Duplicate claim detection tuning.
# Statuses a claim can have while it is still worth voting on. A duplicate is
# only useful to suggest if the user can still act on it, so we never surface
# finalized/locked claims. NOTE: the client creates claims with status "ACTIVE"
# (see services/claimService.ts), while the DB column default is "OPEN"; both are
# included so the feature works regardless of which path created the row.
DUPLICATE_VOTABLE_STATUSES = ["ACTIVE", "OPEN", "PENDING", "EARLY_VERDICT", "REVIEWING"]
# Similarity floor to be considered a candidate at all (spec: > 0.85).
DUPLICATE_SIMILARITY_THRESHOLD = 0.85
# At/above this, embeddings alone are trusted and the stance check is skipped.
DUPLICATE_STANCE_SKIP_SIMILARITY = 0.95
DUPLICATE_MATCH_LIMIT = 3
# Duplicate check fires on every submission; allow a generous per-IP budget.
DUPLICATE_CHECK_RATE_LIMIT_MAX_REQUESTS = 120

# SINGLE WRITE PATH — server-owned claim creation limits. The per-IP limit is
# an inexpensive first layer; the per-user rolling 24-hour count below is the
# durable limit shared by every Render worker.
CLAIMS_CREATE_RATE_LIMIT_MAX_REQUESTS = 30
CLAIMS_PER_DAY_LIMIT = max(1, int(os.environ.get("CLAIMS_PER_DAY_LIMIT", "20")))
CLAIM_TITLE_MAX_LENGTH = 160
CLAIM_DESCRIPTION_MAX_LENGTH = 2000

RESERVED_USERNAME_MESSAGE = (
    "This username is reserved. If you represent this person or organization, "
    "please apply for verification."
)
USERNAME_TAKEN_MESSAGE = "Username is already taken"
IDENTITY_ADMIN_ROLES = {"SUPER_ADMIN", "ADMIN", "MODERATOR"}
ROLE_ASSIGNMENT_PERMISSIONS = {
    "SUPER_ADMIN": {"ADMIN", "MODERATOR"},
    "ADMIN": {"MODERATOR"},
    "MODERATOR": set(),
}
INITIAL_ADMIN_ROLES = {
    "md.noithat@gmail.com": "SUPER_ADMIN",
    "delonnhmd@gmail.com": "ADMIN",
    "minhducmediallc@gmail.com": "MODERATOR",
}
INITIAL_ADMIN_ROLES_SEEDED = False


# PHASE 4 STEP 27
def get_client_ip(request: Request) -> str:
    forwarded_for = request.headers.get("x-forwarded-for", "")
    if forwarded_for:
        return forwarded_for.split(",", 1)[0].strip() or "unknown"

    return request.client.host if request.client else "unknown"


# PHASE 4 STEP 27
def _sweep_stale_rate_limit_state(now: float) -> None:
    stale_bucket_keys = [
        key
        for key, timestamps in RATE_LIMIT_BUCKETS.items()
        if not timestamps or now - timestamps[-1] > _STALE_ENTRY_TTL_SECONDS
    ]
    for key in stale_bucket_keys:
        RATE_LIMIT_BUCKETS.pop(key, None)

    stale_cooldown_keys = [
        claim_id
        for claim_id, last_request in CLAIM_AI_COOLDOWNS.items()
        if now - last_request > _STALE_ENTRY_TTL_SECONDS
    ]
    for claim_id in stale_cooldown_keys:
        CLAIM_AI_COOLDOWNS.pop(claim_id, None)


def enforce_rate_limit(request: Request, bucket: str, max_requests: int, window_seconds: int) -> None:
    global _rate_limit_sweep_counter
    now = monotonic()
    key = f"{bucket}:{get_client_ip(request)}"
    recent_requests = [
        timestamp
        for timestamp in RATE_LIMIT_BUCKETS.get(key, [])
        if now - timestamp < window_seconds
    ]

    if len(recent_requests) >= max_requests:
        print("[rate-limit] blocked:", bucket, get_client_ip(request), flush=True)
        raise HTTPException(status_code=429, detail="Too many actions. Please try again later.")

    recent_requests.append(now)
    RATE_LIMIT_BUCKETS[key] = recent_requests

    _rate_limit_sweep_counter += 1
    if _rate_limit_sweep_counter >= _RATE_LIMIT_SWEEP_INTERVAL:
        _rate_limit_sweep_counter = 0
        _sweep_stale_rate_limit_state(now)


# PHASE 4 STEP 27
def enforce_claim_ai_cooldown(claim_id: str) -> None:
    now = monotonic()
    last_request = CLAIM_AI_COOLDOWNS.get(claim_id)

    if last_request is not None and now - last_request < AI_CLAIM_COOLDOWN_SECONDS:
        print("[ai cooldown] blocked claim:", claim_id, flush=True)
        raise HTTPException(status_code=429, detail="AI pre-check was just run. Please try again later.")

    CLAIM_AI_COOLDOWNS[claim_id] = now


class AiPrecheckRequest(BaseModel):
    claim_id: str = ""
    title: str = ""
    description: str = ""
    source_url: str = ""
    category: str = ""


# PHASE 4 STEP 3
class AiPrecheckRetryRequest(BaseModel):
    claim_id: str = ""


# PHASE 6 STEP 1 — Duplicate detection request bodies.
class DuplicateCheckRequest(BaseModel):
    title: str = ""
    description: str = ""


# CONTENT SAFETY (NEW, additive) — objectionable-content gate at submission.
class ContentSafetyRequest(BaseModel):
    title: str = ""
    description: str = ""


# SINGLE WRITE PATH — author_id is deliberately absent. It always comes from
# the verified Supabase JWT, never from caller-controlled JSON.
class ClaimCreateRequest(BaseModel):
    title: str
    description: str
    category: str
    source_url: str
    video_url: str | None = None
    image_url: str | None = None
    image_path: str | None = None
    thumbnail_url: str | None = None
    sub_category: str | None = None
    politician_tag: str | None = None


class ClaimVoteRequest(BaseModel):
    vote_type: str


class ClaimReportRequest(BaseModel):
    reason: str
    note: str = ""


class ClaimEvidenceRequest(BaseModel):
    url: str
    note: str
    evidence_type: str = "ADDS_CONTEXT"


# PHASE 6 STEP 1 — Fire-and-forget embedding storage for a just-created claim.
class ClaimEmbedRequest(BaseModel):
    claim_id: str = ""


# PHASE 6 STEP 2 — Offline citation evidence submission (books/newspapers/journals/documents).
class CitationEvidenceRequest(BaseModel):
    claim_id: str = ""
    evidence_type: str = ""
    reference_type: str = ""
    citation: dict[str, Any] = {}
    note: str = ""


# PHASE 6 STEP 2 — Citation dispute submission.
class CitationDisputeRequest(BaseModel):
    reason: str = ""


# PHASE 5 STEP 1B
class AdminOverrideRequest(BaseModel):
    claim_id: str = ""
    new_status: str = ""
    reason: str = ""


# PHASE 5 STEP 2
class AdminReportActionRequest(BaseModel):
    status: str = "RESOLVED"
    admin_note: str = ""
    hide_target: bool = False


# PHASE 5 STEP 3
class AdminContentVisibilityRequest(BaseModel):
    target_type: str = "CLAIM"
    target_id: str = ""
    reason: str = ""


class AdminClaimActionRequest(BaseModel):
    claim_id: str = ""
    reason: str = ""
    featured: bool = True


class AdminUserActionRequest(BaseModel):
    user_id: str = ""
    reason: str = ""
    suspended: bool = True


# HIDE/UNHIDE CLAIMS (NEW, additive) — POST /admin/claims/{claim_id}/hide
class AdminHideClaimRequest(BaseModel):
    reason: str = ""


# MODERATION APPEALS (NEW, additive)
class AppealCreateRequest(BaseModel):
    action_type: str = ""
    claim_id: str | None = None
    notification_id: str | None = None
    appeal_text: str = ""


class AppealResolveRequest(BaseModel):
    decision: str = ""
    review_note: str = ""


class IdentityUsernameCheckRequest(BaseModel):
    username: str = ""


class VerificationRequestPayload(BaseModel):
    request_type: str = ""
    requested_name: str = ""
    official_email: str = ""
    official_website: str = ""
    social_links: list[Any] = []
    supporting_documents: list[Any] = []


class AdminVerificationDecisionRequest(BaseModel):
    decision_note: str = ""
    claimed_by_user_id: str | None = None


class AdminAssignRoleRequest(BaseModel):
    target_email: str = ""
    email: str = ""
    role: str = ""
    note: str = ""
    reason: str = ""


class AdminRoleStatusRequest(BaseModel):
    target_email: str = ""
    email: str = ""
    note: str = ""
    reason: str = ""


class ProfileEnsureRequest(BaseModel):
    username: str = ""
    display_name: str = ""


class ProfileUpdateRequest(BaseModel):
    username: str | None = None
    display_name: str | None = None
    avatar_url: str | None = None
    avatar_path: str | None = None
    bio: str | None = None
    profile_visibility: str | None = None


class MentionTagsRequest(BaseModel):
    target_type: Literal["claim", "evidence"]
    target_id: str = ""
    text: str = ""


# APPLE GUIDELINE 1.2 — user blocking (NEW)
class BlockUserRequest(BaseModel):
    source_claim_id: str | None = None


class AiPrecheckResponse(BaseModel):
    ok: bool
    claim_id: str
    # PHASE 4 STEP 7
    claim_type: ClaimType | None = None
    ai_confidence: float | None = None
    source_count: int | None = None
    source_quality: str | None = None
    # PHASE 4 STEP 9
    source_domain: str | None = None
    source_score: int | None = None
    source_reason: str | None = None
    # PHASE 4 STEP 21
    source_read_status: str | None = None
    source_page_title: str | None = None
    source_supports_claim: bool | None = None
    source_support_summary: str | None = None
    # PHASE 4 STEP 10
    evidence_used_count: int | None = None
    red_flags: list[str] | None = None
    ai_summary: str | None = None
    ai_status: AiStatus | None = None
    # PHASE 6 STEP 3 — natural-truth classification (NEW, optional)
    naturally_true_category: str | None = None
    verdict_signal: str | None = None
    error: str | None = None
    # PHASE 4 STEP 5B
    details: str | None = None
    hint: str | None = None
    supabase_updated: bool | None = None
    updated_claim: dict | None = None


def normalize_profile_username_value(value: Any) -> str:
    return str(value or "").strip().lstrip("@").lower()


def normalize_public_profile_slug_value(value: Any) -> str:
    normalized = re.sub(r"[^a-z0-9]+", "-", str(value or "").strip().lower())
    return normalized.strip("-")[:48]


def is_valid_profile_username(value: str) -> bool:
    return bool(re.fullmatch(r"[a-z0-9_]{3,20}", value))


def get_metadata_username_values(user: dict) -> set[str]:
    metadata = user.get("user_metadata") or user.get("raw_user_meta_data") or {}

    if not isinstance(metadata, dict):
        metadata = {}

    values: set[str] = set()
    for key in ("username", "displayName", "display_name", "full_name", "name"):
        normalized_value = normalize_profile_username_value(metadata.get(key))

        if normalized_value:
            values.add(normalized_value)

    return values


def fetch_auth_users_for_username_check() -> list[dict]:
    supabase_url = os.environ.get("SUPABASE_URL", "").rstrip("/")
    service_role_key = os.environ.get("SUPABASE_SERVICE_ROLE_KEY", "")

    if not supabase_url or not service_role_key:
        raise RuntimeError("Missing SUPABASE_URL or SUPABASE_SERVICE_ROLE_KEY on backend.")

    users: list[dict] = []

    for page in range(1, 21):
        url = f"{supabase_url}/auth/v1/admin/users?page={page}&per_page=1000"
        request = UrlRequest(
            url,
            headers={
                "apikey": service_role_key,
                "authorization": f"Bearer {service_role_key}",
            },
        )

        with urlopen(request, timeout=10) as response:
            payload = json.loads(response.read().decode("utf-8"))

        page_users = payload.get("users") if isinstance(payload, dict) else payload

        if not isinstance(page_users, list) or not page_users:
            break

        users.extend([user for user in page_users if isinstance(user, dict)])

        if len(page_users) < 1000:
            break

    return users


ADMIN_METRICS_CACHE_TTL_SECONDS = 300
_admin_metrics_cache_payload: dict[str, Any] | None = None
_admin_metrics_cache_expires_at = 0.0


def normalize_identity_key(value: Any) -> str:
    return re.sub(r"[\s_.-]+", "", str(value or "").strip().lower())


def normalize_admin_email(value: Any) -> str:
    return str(value or "").strip().lower()


def normalize_admin_role(value: Any) -> str:
    return str(value or "").strip().upper().replace(" ", "_")


def now_utc_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def get_first_row(result: Any) -> dict | None:
    rows = getattr(result, "data", None) or []

    if isinstance(rows, dict):
        return rows

    return rows[0] if rows else None


def is_missing_column_error(error: Exception, column_names: list[str]) -> bool:
    message = str(error).lower()
    return (
        any(column.lower() in message for column in column_names)
        and ("column" in message or "schema cache" in message or "could not find" in message)
    )


def execute_reserved_query(query_builder, optional_columns: list[str] | None = None):
    try:
        return query_builder("active").execute()
    except Exception as active_error:
        if optional_columns and is_missing_column_error(active_error, optional_columns):
            return None
        if not is_missing_column_error(active_error, ["active"]):
            raise

    try:
        return query_builder("is_active").execute()
    except Exception as is_active_error:
        if optional_columns and is_missing_column_error(is_active_error, optional_columns):
            return None
        if not is_missing_column_error(is_active_error, ["is_active"]):
            raise

    try:
        return query_builder(None).execute()
    except Exception as error:
        if optional_columns and is_missing_column_error(error, optional_columns):
            return None
        raise


def find_reserved_row_by_key(supabase: Any, table_name: str, normalized_key: str) -> dict | None:
    def build_query(active_column: str | None):
        query = supabase.table(table_name).select("*").eq("normalized_key", normalized_key).limit(1)

        if active_column:
            query = query.eq(active_column, True)

        return query

    result = execute_reserved_query(build_query)
    return get_first_row(result) if result else None


def find_reserved_row_by_alias(supabase: Any, table_name: str, normalized_key: str) -> dict | None:
    for alias_column in ("aliases", "alias_keys", "normalized_aliases"):
        def build_query(active_column: str | None, column: str = alias_column):
            query = supabase.table(table_name).select("*").contains(column, [normalized_key]).limit(1)

            if active_column:
                query = query.eq(active_column, True)

            return query

        result = execute_reserved_query(build_query, optional_columns=[alias_column])
        row = get_first_row(result) if result else None

        if row:
            return row

    return None


def find_reserved_identity_match(value: Any) -> dict | None:
    normalized_key = normalize_identity_key(value)

    if not normalized_key:
        return None

    supabase = get_supabase_client()
    targets = (
        ("reserved_people", "PERSON"),
        ("reserved_brands", "BRAND"),
    )

    for table_name, reserved_type in targets:
        row = find_reserved_row_by_key(supabase, table_name, normalized_key)

        if not row:
            row = find_reserved_row_by_alias(supabase, table_name, normalized_key)

        if row:
            return {
                "type": reserved_type,
                "table": table_name,
                "normalized_key": normalized_key,
                "row": row,
            }

    return None


def query_reserved_identity_table_for_username(
    supabase: Any,
    table_name: str,
    normalized_key: str,
) -> tuple[dict | None, str | None]:
    try:
        result = (
            supabase.table(table_name)
            .select("*")
            .eq("normalized_key", normalized_key)
            .eq("active", True)
            .limit(1)
            .execute()
        )
    except Exception as error:
        error_message = str(error)
        print(f"[identity check] {table_name} normalized_key query failed:", error_message, flush=True)
        return None, error_message

    if result.data:
        return get_first_row(result), None

    try:
        alias_result = (
            supabase.table(table_name)
            .select("*")
            .contains("aliases", [normalized_key])
            .eq("active", True)
            .limit(1)
            .execute()
        )
    except Exception as error:
        error_message = str(error)
        print(f"[identity check] {table_name} aliases query skipped:", error_message, flush=True)
        return None, None

    return get_first_row(alias_result), None


def is_system_reserved_username_row(row: dict | None) -> bool:
    if not row:
        return False

    category = str(row.get("category") or "").strip().lower()
    source_import = str(row.get("source_import") or "").strip().lower()
    return (
        category == "system reserved usernames"
        or source_import in {"reserved_system_usernames.txt", "032_reserved_system_usernames.sql"}
    )


def build_reserved_username_check_response(username: str) -> dict:
    normalized_key = normalize_identity_key(username)

    if not normalized_key:
        raise HTTPException(status_code=400, detail="Username is required.")

    print("[identity check] incoming username:", str(username or ""), flush=True)
    print("[identity check] normalized username:", normalized_key, flush=True)

    try:
        supabase = get_supabase_client()
    except Exception as error:
        error_message = str(error)
        print("[identity check] supabase client unavailable:", error_message, flush=True)
        print("[identity check] people_match false", flush=True)
        print("[identity check] brand_match false", flush=True)
        print("[identity check] error:", error_message, flush=True)
        return {
            "available": True,
            "reserved": False,
            "message": None,
            "normalized_key": normalized_key,
            "warning": "Reserved identity check temporarily unavailable.",
        }

    people_row, people_error = query_reserved_identity_table_for_username(
        supabase,
        "reserved_people",
        normalized_key,
    )
    brand_row, brand_error = query_reserved_identity_table_for_username(
        supabase,
        "reserved_brands",
        normalized_key,
    )
    error_messages = [message for message in [people_error, brand_error] if message]
    people_match = bool(people_row)
    brand_match = bool(brand_row)

    print("[identity check] people_match", people_match, flush=True)
    print("[identity check] brand_match", brand_match, flush=True)

    if error_messages:
        print("[identity check] error:", " | ".join(error_messages), flush=True)

    if people_match or brand_match:
        system_reserved_username = is_system_reserved_username_row(people_row) and not brand_match

        return {
            "available": False,
            "reserved": not system_reserved_username,
            "normalized_key": normalized_key,
            "message": USERNAME_TAKEN_MESSAGE if system_reserved_username else RESERVED_USERNAME_MESSAGE,
        }

    response = {
        "available": True,
        "reserved": False,
        "message": None,
        "normalized_key": normalized_key,
    }

    if error_messages:
        response["warning"] = "Reserved identity check temporarily unavailable."

    return response


def check_username_availability_for_save(username: str, excluded_user_id: str = "") -> dict:
    normalized_username = normalize_profile_username_value(username)

    if not is_valid_profile_username(normalized_username):
        return {
            "ok": True,
            "available": False,
            "reserved": False,
            "normalized_username": normalized_username,
            "message": "Username must be 3-20 characters.",
        }

    reserved_check = build_reserved_username_check_response(normalized_username)

    if not reserved_check.get("available"):
        return {
            "ok": True,
            "available": False,
            "reserved": bool(reserved_check.get("reserved")),
            "normalized_username": normalized_username,
            "message": reserved_check.get("message") or USERNAME_TAKEN_MESSAGE,
        }

    supabase = get_supabase_client()
    profile_query = (
        supabase.table("profiles")
        .select("id")
        .eq("username_normalized", normalized_username)
        .limit(1)
    )

    if excluded_user_id:
        profile_query = profile_query.neq("id", excluded_user_id)

    profile_result = profile_query.execute()

    if profile_result.data:
        return {
            "ok": True,
            "available": False,
            "reserved": False,
            "normalized_username": normalized_username,
            "message": USERNAME_TAKEN_MESSAGE,
        }

    for auth_user in fetch_auth_users_for_username_check():
        auth_user_id = str(auth_user.get("id") or "")

        if excluded_user_id and auth_user_id == excluded_user_id:
            continue

        if normalized_username in get_metadata_username_values(auth_user):
            return {
                "ok": True,
                "available": False,
                "reserved": False,
                "normalized_username": normalized_username,
                "message": USERNAME_TAKEN_MESSAGE,
            }

    return {
        "ok": True,
        "available": True,
        "reserved": False,
        "normalized_username": normalized_username,
        "message": "Username available",
    }


def read_auth_user_metadata(auth_user: Any) -> dict:
    metadata = getattr(auth_user, "user_metadata", None) or getattr(auth_user, "raw_user_meta_data", None) or {}
    return metadata if isinstance(metadata, dict) else {}


def read_auth_user_metadata_string(auth_user: Any, key: str) -> str:
    value = read_auth_user_metadata(auth_user).get(key)
    return value.strip() if isinstance(value, str) and value.strip() else ""


def generate_backend_profile_slug(username: str, user_id: str) -> str:
    base_slug = re.sub(r"[^a-z0-9]+", "-", str(username or "").strip().lower()).strip("-")[:48] or "user"
    suffix = user_id.replace("-", "")[-6:].lower()
    return f"{base_slug}-{suffix}"[:48] if suffix else base_slug


def generate_backend_default_username(seed: str) -> str:
    digest = hashlib.md5(str(seed or "user").encode("utf-8")).hexdigest()
    return f"user_{1000 + (int(digest[:8], 16) % 9000):04d}"


def generate_backend_fallback_username(email: str, user_id: str) -> str:
    return generate_backend_default_username(user_id or email or "user")


def is_backend_generated_placeholder_username(value: str | None) -> bool:
    normalized = normalize_profile_username_value(value or "")
    return bool(normalized and re.fullmatch(r"[a-z][a-z0-9]{2,12}_[a-f0-9]{6}", normalized))


def get_review_safe_backend_username(username: str | None, user_id: str | None) -> str:
    normalized = normalize_profile_username_value(username or "")

    if not is_valid_profile_username(normalized) or is_backend_generated_placeholder_username(normalized):
        return generate_backend_default_username(user_id or username or "user")

    return normalized


def get_review_safe_backend_display_name(display_name: str | None, username: str | None, user_id: str | None) -> str:
    trimmed_display_name = str(display_name or "").strip()

    if trimmed_display_name and not is_backend_generated_placeholder_username(trimmed_display_name):
        return trimmed_display_name

    return get_review_safe_backend_username(username, user_id)


def get_backend_preferred_username(auth_user: Any) -> str:
    email = str(getattr(auth_user, "email", "") or "")
    user_id = str(getattr(auth_user, "id", "") or "")
    metadata_username = read_auth_user_metadata_string(auth_user, "username")

    for value in (metadata_username, generate_backend_fallback_username(email, user_id)):
        normalized = normalize_profile_username_value(value)

        if is_valid_profile_username(normalized):
            return normalized

    return generate_backend_fallback_username(email, user_id)


MENTION_USERNAME_PATTERN = re.compile(r"@([A-Za-z0-9_]+)")


def normalize_mention_search_query(value: str) -> str:
    return re.sub(r"[%(),]", " ", str(value or "").strip().lstrip("@"))[:32].strip()


def extract_backend_mentions(text: str, limit: int) -> list[str]:
    usernames: list[str] = []
    seen: set[str] = set()

    for match in MENTION_USERNAME_PATTERN.finditer(text or ""):
        username = normalize_profile_username_value(match.group(1))

        if not is_valid_profile_username(username) or username in seen:
            continue

        seen.add(username)
        usernames.append(username)

        if len(usernames) >= limit:
            break

    return usernames


def safe_execute_table_query(query: Any, context: str) -> list[dict]:
    try:
        result = query.execute()
        return result.data or []
    except Exception as error:
        print(f"[mentions] {context} query warning:", error, flush=True)
        return []


def map_profile_mention_result(row: dict) -> dict | None:
    if row.get("is_suspended") or row.get("is_deleted"):
        return None

    user_id = str(row.get("id") or "")
    username = get_review_safe_backend_username(row.get("username"), user_id)

    if not user_id or not username:
        return None

    trust_score = row.get("trust_score") or 50
    rank_title = resolve_display_rank(calculate_rank_title(trust_score), row.get("highest_rank_achieved") or row.get("rank_title"))

    return {
        "type": "user",
        "id": user_id,
        "username": username,
        "display_name": get_review_safe_backend_display_name(row.get("display_name"), row.get("username"), user_id),
        "rank_title": rank_title,
        "avatar_url": row.get("avatar_url"),
        "verified": bool(row.get("verified")),
    }


def map_organization_mention_result(row: dict) -> dict | None:
    slug = normalize_profile_username_value(row.get("slug"))

    if not slug:
        return None

    return {
        "type": "organization",
        "id": str(row.get("id") or ""),
        "username": slug,
        "display_name": str(row.get("name") or slug),
        "avatar_url": row.get("avatar_url"),
        "verified": bool(row.get("verified")),
    }


def search_profile_mentions(supabase: Any, query: str, limit: int) -> list[dict]:
    select_fields = "id,username,display_name,avatar_url,verified,trust_score,rank_title,highest_rank_achieved,is_suspended,is_deleted"
    rows_by_id: dict[str, dict] = {}

    for column, search_value in (("username", normalize_profile_username_value(query)), ("display_name", query)):
        if not search_value:
            continue

        rows = safe_execute_table_query(
            supabase.table("profiles").select(select_fields).ilike(column, f"%{search_value}%").limit(limit),
            f"profile {column}",
        )

        for row in rows:
            row_id = str(row.get("id") or "")

            if row_id:
                rows_by_id[row_id] = row

    return [result for row in rows_by_id.values() if (result := map_profile_mention_result(row))]


def search_organization_mentions(supabase: Any, query: str, limit: int) -> list[dict]:
    rows_by_id: dict[str, dict] = {}

    for column, search_value in (("slug", normalize_profile_username_value(query)), ("name", query)):
        if not search_value:
            continue

        rows = safe_execute_table_query(
            supabase.table("organizations").select("id,name,slug,description,avatar_url,verified,created_at").ilike(column, f"%{search_value}%").limit(limit),
            f"organization {column}",
        )

        for row in rows:
            row_id = str(row.get("id") or "")

            if row_id:
                rows_by_id[row_id] = row

    return [result for row in rows_by_id.values() if (result := map_organization_mention_result(row))]


def fetch_mention_targets_by_username(supabase: Any, usernames: list[str]) -> tuple[dict[str, dict], dict[str, dict]]:
    profiles_by_username: dict[str, dict] = {}
    organizations_by_slug: dict[str, dict] = {}

    if not usernames:
        return profiles_by_username, organizations_by_slug

    profile_rows = safe_execute_table_query(
        supabase.table("profiles")
        .select("id,username,username_normalized,display_name,is_suspended,is_deleted")
        .in_("username_normalized", usernames),
        "profile exact",
    )

    if not profile_rows:
        profile_rows = safe_execute_table_query(
            supabase.table("profiles")
            .select("id,username,display_name,is_suspended,is_deleted")
            .in_("username", usernames),
            "profile exact fallback",
        )

    for row in profile_rows:
        if row.get("is_suspended") or row.get("is_deleted"):
            continue

        username = normalize_profile_username_value(row.get("username_normalized") or row.get("username"))

        if username:
            profiles_by_username[username] = row

    organization_rows = safe_execute_table_query(
        supabase.table("organizations").select("id,name,slug,verified").in_("slug", usernames),
        "organization exact",
    )

    for row in organization_rows:
        slug = normalize_profile_username_value(row.get("slug"))

        if slug:
            organizations_by_slug[slug] = row

    return profiles_by_username, organizations_by_slug


def insert_mention_tag_row(supabase: Any, table_name: str, payload: dict) -> bool:
    try:
        supabase.table(table_name).insert(payload).execute()
        return True
    except Exception as error:
        message = str(error).lower()

        if "duplicate" in message or "23505" in message:
            return False

        print("[mentions] tag insert warning:", error, flush=True)
        return False


def truncate_notification_body(value: Any, limit: int = 60) -> str:
    text = re.sub(r"\s+", " ", str(value or "")).strip()
    return text[:limit]


def insert_notification_record(
    supabase: Any,
    target_user_id: str,
    notification_type: str,
    title: str,
    body: str,
    claim_id: str | None = None,
) -> None:
    payload = {
        "user_id": target_user_id,
        "type": notification_type,
        "title": title,
        "body": body,
        "claim_id": claim_id,
    }

    try:
        supabase.table("notifications").insert(payload).execute()
    except Exception as error:
        print("[mentions] notification warning:", error, flush=True)


def sanitize_backend_bio(value: str | None) -> str | None:
    if value is None:
        return None

    return re.sub(r"\s+", " ", value.replace("<", "").replace(">", "")).strip()[:160]


def normalize_backend_profile_visibility(value: str | None) -> str:
    return "private" if value == "private" else "public"


def is_valid_backend_avatar_url(value: str | None) -> bool:
    if value is None or not value.strip():
        return True

    parsed = urlparse(value.strip())
    return parsed.scheme in {"http", "https"} and bool(parsed.netloc)


def build_profile_response(row: dict | None, updated_rows: int | None = None) -> dict:
    response = {"ok": True, "profile": row}

    if updated_rows is not None:
        response["updated_rows"] = updated_rows

    return response


def fetch_profile_row(supabase: Any, user_id: str) -> dict | None:
    result = supabase.table("profiles").select("*").eq("id", user_id).limit(1).execute()
    return get_first_row(result)


def lookup_public_profile_row(
    supabase: Any,
    identifier: str,
    username: str | None = None,
) -> dict | None:
    trimmed_identifier = str(identifier or "").strip()
    normalized_slug = normalize_public_profile_slug_value(trimmed_identifier)
    normalized_username = normalize_profile_username_value(username or trimmed_identifier)
    lookups: list[tuple[str, str]] = []

    if is_uuid(trimmed_identifier):
        lookups.append(("id", trimmed_identifier))

    if normalized_slug:
        lookups.append(("public_profile_slug", normalized_slug))

    if normalized_username:
        lookups.extend((
            ("username_normalized", normalized_username),
            ("username", normalized_username),
        ))

    seen: set[tuple[str, str]] = set()

    for column_name, value in lookups:
        if not value or (column_name, value) in seen:
            continue

        seen.add((column_name, value))

        try:
            result = (
                supabase.table("profiles")
                .select("*")
                .eq(column_name, value)
                .limit(1)
                .execute()
            )
        except Exception as error:
            if is_missing_column_error(error, [column_name]):
                print(f"[public profile] lookup skipped missing column {column_name}", flush=True)
                continue

            print("[public profile] lookup warning:", error, flush=True)
            continue

        row = get_first_row(result)

        if row:
            return row

    return None


def build_public_profile_api_row(row: dict) -> dict:
    user_id = str(row.get("id") or "")
    username = get_review_safe_backend_username(row.get("username"), user_id)
    display_name = get_review_safe_backend_display_name(row.get("display_name"), row.get("username"), user_id)
    trust_score = row.get("trust_score") or 50
    current_rank = calculate_rank_title(trust_score)
    rank_title = resolve_display_rank(
        current_rank,
        row.get("highest_rank_achieved") or row.get("rank_title"),
    )

    return {
        "id": user_id,
        "username": username,
        "display_name": display_name,
        "avatar_url": row.get("avatar_url"),
        "bio": row.get("bio"),
        "public_profile_slug": row.get("public_profile_slug") or generate_backend_profile_slug(username, user_id),
        "profile_visibility": normalize_backend_profile_visibility(row.get("profile_visibility")),
        "trust_score": trust_score,
        "rank_title": rank_title,
        "highest_rank_achieved": row.get("highest_rank_achieved") or rank_title,
        "reputation_points": row.get("reputation_points") or row.get("reputation_score") or 0,
        "monthly_reputation_points": row.get("monthly_reputation_points") or 0,
        "badge_list": row.get("badge_list") if isinstance(row.get("badge_list"), list) else [],
        "evidence_count": row.get("evidence_count") or 0,
        "correct_votes": row.get("correct_votes") or 0,
        "created_at": row.get("created_at"),
        "is_deleted": bool(row.get("is_deleted")),
        "deleted_at": row.get("deleted_at"),
    }


PROFILE_ACTIVITY_LIMIT = 100
PUBLIC_CLAIM_STATUSES = {
    "PENDING",
    "ACTIVE",
    "EARLY_VERDICT",
    "FINALIZED_TRUE",
    "FINALIZED_FAKE",
    "INSUFFICIENT_DATA",
    "LOCKED",
    "OPEN",
    "VOTING_CLOSED",
    "COMMUNITY_TRUE",
    "COMMUNITY_FAKE",
    "NEEDS_MORE_EVIDENCE",
}


def is_missing_relation_error(error: Exception) -> bool:
    message = str(error).lower()
    return any(marker in message for marker in ("42p01", "pgrst205", "relation", "table")) and any(
        marker in message for marker in ("does not exist", "schema cache", "could not find")
    )


def is_public_profile_claim(row: dict | None) -> bool:
    if not row:
        return False

    return not any((row.get("is_deleted"), row.get("is_hidden"), row.get("hidden"))) and str(
        row.get("safety_status") or ""
    ).upper() == "APPROVED"


def get_public_claim_verdict(status: Any) -> str | None:
    normalized = str(status or "").upper()
    if normalized in {"FINALIZED_TRUE", "COMMUNITY_TRUE"}:
        return "TRUE"
    if normalized in {"FINALIZED_FAKE", "COMMUNITY_FAKE"}:
        return "FAKE"
    if normalized in {"INSUFFICIENT_DATA", "NEEDS_MORE_EVIDENCE"}:
        return "NEEDS_MORE_EVIDENCE"
    return None


def safe_public_count(value: Any) -> int:
    try:
        return max(0, int(value or 0))
    except (TypeError, ValueError):
        return 0


def profile_activity_is_blocked(supabase: Any, viewer_id: str | None, profile_id: str) -> bool:
    if not viewer_id or viewer_id == profile_id:
        return False

    for blocker_id, blocked_id in ((viewer_id, profile_id), (profile_id, viewer_id)):
        try:
            result = (
                supabase.table("user_blocks")
                .select("id")
                .eq("blocker_id", blocker_id)
                .eq("blocked_id", blocked_id)
                .limit(1)
                .execute()
            )
        except Exception as error:
            print("[profile activity] block lookup failed:", str(error), flush=True)
            raise HTTPException(status_code=503, detail="Could not load this profile right now.")

        if get_first_row(result):
            return True

    return False


def get_public_profile_activity_context(request: Request, identifier: str) -> tuple[Any, dict]:
    supabase = get_supabase_client()
    profile = lookup_public_profile_row(supabase, identifier)

    if not profile or profile.get("is_deleted"):
        raise HTTPException(status_code=404, detail="Contributor profile unavailable.")

    viewer_id = get_optional_authenticated_user_id(request)
    profile_id = str(profile.get("id") or "")

    if not profile_id or profile_activity_is_blocked(supabase, viewer_id, profile_id):
        raise HTTPException(status_code=404, detail="Contributor profile unavailable.")

    return supabase, profile


def fetch_public_profile_posts(supabase: Any, profile_id: str) -> list[dict]:
    try:
        result = (
            supabase.table("claims")
            .select(
                "id,title,description,image_url,thumbnail_url,category,status,votes_true,votes_fake,"
                "votes_unsure,total_votes,created_at,is_deleted,is_hidden,hidden,safety_status"
            )
            .eq("author_id", profile_id)
            .order("created_at", desc=True)
            .limit(PROFILE_ACTIVITY_LIMIT)
            .execute()
        )
    except Exception as error:
        print("[profile activity] posts failed:", str(error), flush=True)
        raise HTTPException(status_code=503, detail="Could not load profile posts right now.")

    posts: list[dict] = []
    for row in result.data or []:
        if not is_public_profile_claim(row):
            continue

        description = re.sub(r"\s+", " ", str(row.get("description") or "")).strip()
        posts.append(
            {
                "id": str(row.get("id") or ""),
                "title": str(row.get("title") or "Untitled claim")[:200],
                "description_preview": description[:280],
                "image_url": row.get("image_url"),
                "thumbnail_url": row.get("thumbnail_url"),
                "category": row.get("category"),
                "status": str(row.get("status") or "") if row.get("status") else None,
                "final_verdict": get_public_claim_verdict(row.get("status")),
                "vote_totals": {
                    "true": safe_public_count(row.get("votes_true")),
                    "fake": safe_public_count(row.get("votes_fake")),
                    "unsure": safe_public_count(row.get("votes_unsure")),
                    "total": safe_public_count(row.get("total_votes")),
                },
                "created_at": row.get("created_at"),
            }
        )

    return posts


def fetch_claims_for_public_activity(supabase: Any, claim_ids: list[str]) -> dict[str, dict]:
    valid_ids = sorted({claim_id for claim_id in claim_ids if is_uuid(claim_id)})
    if not valid_ids:
        return {}

    try:
        result = (
            supabase.table("claims")
            .select("id,title,is_deleted,is_hidden,hidden,safety_status")
            .in_("id", valid_ids)
            .execute()
        )
    except Exception as error:
        print("[profile activity] related claims failed:", str(error), flush=True)
        raise HTTPException(status_code=503, detail="Could not load profile activity right now.")

    return {
        str(row.get("id")): row
        for row in (result.data or [])
        if is_public_profile_claim(row)
    }


def is_removed_public_reply(row: dict) -> bool:
    status = str(row.get("status") or "").upper()
    return bool(
        row.get("is_deleted")
        or row.get("deleted_at")
        or row.get("removed")
        or row.get("is_removed")
        or row.get("hidden")
        or row.get("is_hidden")
        or status in {"DELETED", "REMOVED", "HIDDEN", "REJECTED"}
    )


def fetch_public_profile_replies(supabase: Any, profile_id: str) -> list[dict]:
    rows: list[dict] = []

    # The current repository has no comments migration yet. Supporting both
    # conventional relation names keeps this endpoint stable without exposing
    # raw rows or inventing a public votes fallback.
    for table_name in ("comments", "replies"):
        try:
            result = (
                supabase.table(table_name)
                .select("*")
                .eq("user_id", profile_id)
                .order("created_at", desc=True)
                .limit(PROFILE_ACTIVITY_LIMIT)
                .execute()
            )
            rows = list(result.data or [])
            break
        except Exception as error:
            if is_missing_relation_error(error):
                continue
            print(f"[profile activity] {table_name} failed:", str(error), flush=True)
            raise HTTPException(status_code=503, detail="Could not load profile replies right now.")

    visible_rows = [row for row in rows if not is_removed_public_reply(row)]
    claim_ids = [str(row.get("claim_id") or "") for row in visible_rows]
    claims = fetch_claims_for_public_activity(supabase, claim_ids)
    replies: list[dict] = []

    for row in visible_rows:
        claim_id = str(row.get("claim_id") or "")
        claim = claims.get(claim_id)
        text = next(
            (
                str(value).strip()
                for value in (row.get("text"), row.get("content"), row.get("body"), row.get("comment_text"))
                if value is not None and str(value).strip()
            ),
            "",
        )
        reply_id = str(row.get("id") or "")

        if not claim or not reply_id or not text:
            continue

        replies.append(
            {
                "id": reply_id,
                "text": text[:2000],
                "claim_id": claim_id,
                "claim_title": str(claim.get("title") or "Untitled claim")[:200],
                "created_at": row.get("created_at"),
                "reply_count": safe_public_count(row.get("reply_count") or row.get("replies_count")),
                "helpful_count": safe_public_count(row.get("helpful_count")),
                "anchor": f"reply-{reply_id}",
            }
        )

    return replies


def is_public_evidence_row(row: dict) -> bool:
    if row.get("hidden") or row.get("is_hidden"):
        return False

    status = str(row.get("status") or row.get("moderation_status") or "").upper()
    return status not in {"REJECTED", "REMOVED", "HIDDEN", "DELETED", "BLOCKED"}


def fetch_public_profile_evidence(supabase: Any, profile_id: str) -> list[dict]:
    try:
        result = (
            supabase.table("evidence")
            .select("*")
            .eq("user_id", profile_id)
            .order("created_at", desc=True)
            .limit(PROFILE_ACTIVITY_LIMIT)
            .execute()
        )
    except Exception as error:
        print("[profile activity] evidence failed:", str(error), flush=True)
        raise HTTPException(status_code=503, detail="Could not load profile evidence right now.")

    visible_rows = [row for row in (result.data or []) if is_public_evidence_row(row)]
    claim_ids = [str(row.get("claim_id") or "") for row in visible_rows]
    claims = fetch_claims_for_public_activity(supabase, claim_ids)
    evidence_items: list[dict] = []

    for row in visible_rows:
        claim_id = str(row.get("claim_id") or "")
        claim = claims.get(claim_id)
        evidence_id = str(row.get("id") or "")

        if not claim or not evidence_id:
            continue

        source_url = str(row.get("url") or "").strip()
        source_domain = urlparse(source_url).hostname if source_url else None
        evidence_items.append(
            {
                "id": evidence_id,
                "evidence_type": str(row.get("evidence_type") or "UNCLEAR"),
                "note": str(row.get("note") or "")[:2000],
                "source_url": source_url or None,
                "source_domain": source_domain,
                "image_url": row.get("image_url"),
                "thumbnail_url": row.get("thumbnail_url"),
                "claim_id": claim_id,
                "claim_title": str(claim.get("title") or "Untitled claim")[:200],
                "helpful_count": safe_public_count(row.get("helpful_count")),
                "created_at": row.get("created_at"),
            }
        )

    return evidence_items


def build_public_profile_summary(supabase: Any, profile: dict) -> dict:
    profile_id = str(profile.get("id") or "")
    private_profile = normalize_backend_profile_visibility(profile.get("profile_visibility")) == "private"
    posts = fetch_public_profile_posts(supabase, profile_id)
    replies = fetch_public_profile_replies(supabase, profile_id)
    evidence_items = fetch_public_profile_evidence(supabase, profile_id)
    correct_votes = safe_public_count(profile.get("correct_votes"))
    incorrect_votes = safe_public_count(profile.get("incorrect_votes"))
    finalized_votes = correct_votes + incorrect_votes

    try:
        vote_result = (
            supabase.table("votes")
            .select("id", count="exact")
            .eq("user_id", profile_id)
            .limit(1)
            .execute()
        )
        total_votes = safe_public_count(getattr(vote_result, "count", None))
    except Exception as error:
        print("[profile activity] aggregate vote count warning:", str(error), flush=True)
        total_votes = finalized_votes

    public_profile = build_public_profile_api_row(profile)
    public_profile.pop("trust_score", None)
    public_profile.pop("correct_votes", None)
    public_profile.pop("deleted_at", None)
    public_profile["bio"] = None if private_profile else public_profile.get("bio")
    public_profile["badge_list"] = [] if private_profile else public_profile.get("badge_list", [])
    public_profile["reputation_points"] = 0 if private_profile else safe_public_count(public_profile.get("reputation_points"))
    public_profile["monthly_reputation_points"] = 0 if private_profile else safe_public_count(public_profile.get("monthly_reputation_points"))

    return {
        "profile": public_profile,
        "counts": {
            "claims": len(posts),
            "replies": len(replies),
            "evidence": len(evidence_items),
        },
        "voting": {
            "total_votes": total_votes,
            "finalized_votes": finalized_votes,
            "accuracy_percentage": round((correct_votes / finalized_votes) * 100, 1) if finalized_votes else None,
        },
    }


def fetch_private_vote_history(supabase: Any, user_id: str) -> list[dict]:
    try:
        vote_result = (
            supabase.table("votes")
            .select("claim_id,vote_type,created_at")
            .eq("user_id", user_id)
            .order("created_at", desc=True)
            .limit(PROFILE_ACTIVITY_LIMIT)
            .execute()
        )
    except Exception as error:
        print("[profile activity] private votes failed:", str(error), flush=True)
        raise HTTPException(status_code=503, detail="Could not load voting history right now.")

    votes = list(vote_result.data or [])
    claim_ids = [str(row.get("claim_id") or "") for row in votes]
    claims = fetch_claims_for_public_activity(supabase, claim_ids)
    history: list[dict] = []

    for vote in votes:
        claim_id = str(vote.get("claim_id") or "")
        claim = claims.get(claim_id)
        vote_type = str(vote.get("vote_type") or "").upper()

        if not claim or vote_type not in {"TRUE", "FAKE", "UNSURE"}:
            continue

        verdict = get_public_claim_verdict(claim.get("status"))
        history.append(
            {
                "claim_id": claim_id,
                "claim_title": str(claim.get("title") or "Untitled claim")[:200],
                "vote_type": vote_type,
                "claim_status": claim.get("status"),
                "final_verdict": verdict,
                "result": "PENDING" if verdict is None else "MATCHED" if vote_type == ("UNSURE" if verdict == "NEEDS_MORE_EVIDENCE" else verdict) else "DID_NOT_MATCH",
                "voted_at": vote.get("created_at"),
            }
        )

    return history


def get_authenticated_identity(request: Request) -> dict:
    authorization = request.headers.get("authorization", "")

    if not authorization.lower().startswith("bearer "):
        raise HTTPException(status_code=401, detail="Authentication required.")

    access_token = authorization.split(" ", 1)[1].strip()

    if not access_token:
        raise HTTPException(status_code=401, detail="Authentication required.")

    try:
        auth_response = get_supabase_client().auth.get_user(access_token)
        auth_user = getattr(auth_response, "user", None)
    except Exception as error:
        print("[identity auth] token validation failed:", str(error), flush=True)
        raise HTTPException(status_code=401, detail="Authentication required.")

    user_id = str(getattr(auth_user, "id", "") or "")
    email = normalize_admin_email(getattr(auth_user, "email", "") or "")

    if not user_id:
        raise HTTPException(status_code=401, detail="Authentication required.")

    return {
        "id": user_id,
        "email": email,
        "user": auth_user,
    }


def fetch_admin_user_by_email(supabase: Any, email: str) -> dict | None:
    result = (
        supabase.table("admin_users")
        .select("*")
        .eq("email", normalize_admin_email(email))
        .limit(1)
        .execute()
    )
    return get_first_row(result)


def get_admin_active_value(admin_user: dict | None) -> bool:
    if not admin_user:
        return False

    return bool(admin_user.get("active", admin_user.get("is_active", True)))


def sanitize_admin_user(admin_user: dict) -> dict:
    return {
        "email": normalize_admin_email(admin_user.get("email")),
        "role": normalize_admin_role(admin_user.get("role")),
        "active": get_admin_active_value(admin_user),
        "created_at": admin_user.get("created_at"),
        "updated_at": admin_user.get("updated_at"),
    }


def get_admin_action_target_email(payload: Any) -> str:
    return normalize_admin_email(getattr(payload, "target_email", "") or getattr(payload, "email", ""))


def get_admin_action_note(payload: Any) -> str:
    return str(getattr(payload, "note", "") or getattr(payload, "reason", "") or "").strip()


def get_admin_role_by_email_rpc(supabase: Any, email: str) -> str | None:
    try:
        result = supabase.rpc("get_admin_role_by_email", {"email": normalize_admin_email(email)}).execute()
    except Exception as error:
        print("[identity admin] get_admin_role_by_email rpc skipped:", str(error), flush=True)
        return None

    data = getattr(result, "data", None)

    if isinstance(data, str):
        role = normalize_admin_role(data)
        return role if role in IDENTITY_ADMIN_ROLES else None

    if isinstance(data, list) and data:
        first_value = data[0]

        if isinstance(first_value, str):
            role = normalize_admin_role(first_value)
            return role if role in IDENTITY_ADMIN_ROLES else None

        if isinstance(first_value, dict):
            role = normalize_admin_role(first_value.get("role"))
            return role if role in IDENTITY_ADMIN_ROLES else None

    if isinstance(data, dict):
        role = normalize_admin_role(data.get("role"))
        return role if role in IDENTITY_ADMIN_ROLES else None

    return None


def can_assign_admin_role_rpc(supabase: Any, actor_email: str, target_role: str) -> bool | None:
    params_variants = (
        {"actor_email": normalize_admin_email(actor_email), "target_role": normalize_admin_role(target_role)},
        {"p_actor_email": normalize_admin_email(actor_email), "p_target_role": normalize_admin_role(target_role)},
    )

    for params in params_variants:
        try:
            result = supabase.rpc("can_assign_admin_role", params).execute()
        except Exception as error:
            print("[identity admin] can_assign_admin_role rpc skipped:", str(error), flush=True)
            continue

        data = getattr(result, "data", None)

        if isinstance(data, bool):
            return data

        if isinstance(data, list) and data:
            first_value = data[0]

            if isinstance(first_value, bool):
                return first_value

            if isinstance(first_value, dict):
                for value in first_value.values():
                    if isinstance(value, bool):
                        return value

        if isinstance(data, dict):
            for value in data.values():
                if isinstance(value, bool):
                    return value

    return None


def actor_can_assign_role(supabase: Any, actor: dict, target_role: str) -> bool:
    local_allowed = target_role in ROLE_ASSIGNMENT_PERMISSIONS.get(actor["role"], set())

    if not local_allowed:
        return False

    rpc_result = can_assign_admin_role_rpc(supabase, actor["email"], target_role)

    if rpc_result is False:
        return False

    return True


def insert_with_optional_columns(
    supabase: Any,
    table_name: str,
    payload: dict,
    optional_columns: set[str],
):
    try:
        return supabase.table(table_name).insert(payload).execute()
    except Exception as error:
        if not is_missing_column_error(error, list(optional_columns)):
            raise

    stripped_payload = {
        key: value
        for key, value in payload.items()
        if key not in optional_columns
    }
    return supabase.table(table_name).insert(stripped_payload).execute()


def update_with_optional_columns(
    supabase: Any,
    table_name: str,
    payload: dict,
    match_column: str,
    match_value: Any,
    optional_columns: set[str],
):
    try:
        return (
            supabase.table(table_name)
            .update(payload)
            .eq(match_column, match_value)
            .execute()
        )
    except Exception as error:
        if not is_missing_column_error(error, list(optional_columns)):
            raise

    stripped_payload = {
        key: value
        for key, value in payload.items()
        if key not in optional_columns
    }
    return (
        supabase.table(table_name)
        .update(stripped_payload)
        .eq(match_column, match_value)
        .execute()
    )


def insert_admin_user_row(supabase: Any, payload: dict):
    try:
        return insert_with_optional_columns(
            supabase,
            "admin_users",
            payload,
            {"created_at", "updated_at"},
        )
    except Exception as error:
        if not is_missing_column_error(error, ["active"]):
            raise

    next_payload = {
        ("is_active" if key == "active" else key): value
        for key, value in payload.items()
    }
    return insert_with_optional_columns(
        supabase,
        "admin_users",
        next_payload,
        {"created_at", "updated_at"},
    )


def update_admin_user_row(supabase: Any, target_email: str, payload: dict):
    try:
        return update_with_optional_columns(
            supabase,
            "admin_users",
            payload,
            "email",
            normalize_admin_email(target_email),
            {"updated_at"},
        )
    except Exception as error:
        if not is_missing_column_error(error, ["active"]):
            raise

    next_payload = {
        ("is_active" if key == "active" else key): value
        for key, value in payload.items()
    }
    return update_with_optional_columns(
        supabase,
        "admin_users",
        next_payload,
        "email",
        normalize_admin_email(target_email),
        {"updated_at"},
    )


def insert_admin_role_history(
    supabase: Any,
    target_email: str,
    old_role: str | None,
    new_role: str,
    actor: dict,
    reason: str = "",
) -> None:
    payload = {
        "target_email": normalize_admin_email(target_email),
        "old_role": old_role,
        "new_role": new_role,
        "changed_by_user_id": actor.get("id"),
        "changed_by_email": normalize_admin_email(actor.get("email")),
        "reason": reason.strip() or None,
        "created_at": now_utc_iso(),
    }

    try:
        supabase.table("admin_role_history").insert(payload).execute()
    except Exception as error:
        print("[identity admin] role history insert failed:", str(error), flush=True)
        raise HTTPException(status_code=500, detail="Could not record role history right now.")


def record_admin_role_change_entry(
    supabase: Any,
    target_email: str,
    old_role: str | None,
    new_role: str,
    actor: dict,
    note: str = "",
    action: str = "ROLE_ASSIGNED",
) -> None:
    normalized_target_email = normalize_admin_email(target_email)
    actor_email = normalize_admin_email(actor.get("email"))
    params_variants = (
        {
            "target_email": normalized_target_email,
            "old_role": old_role,
            "new_role": new_role,
            "actor_email": actor_email,
            "note": note.strip() or None,
        },
        {
            "p_target_email": normalized_target_email,
            "p_old_role": old_role,
            "p_new_role": new_role,
            "p_actor_email": actor_email,
            "p_note": note.strip() or None,
        },
        {
            "target_email": normalized_target_email,
            "old_role": old_role,
            "new_role": new_role,
            "changed_by_email": actor_email,
            "reason": note.strip() or None,
        },
    )

    for params in params_variants:
        try:
            supabase.rpc("record_admin_role_change", params).execute()
            return
        except Exception as error:
            print("[identity admin] record_admin_role_change rpc skipped:", str(error), flush=True)

    insert_admin_role_history(
        supabase,
        normalized_target_email,
        old_role,
        new_role,
        actor,
        f"{action}: {note.strip()}" if note.strip() else action,
    )


def insert_identity_audit_log(
    supabase: Any,
    actor: dict,
    action: str,
    target_type: str,
    target_id: str | None = None,
    metadata: dict | None = None,
) -> None:
    payload = {
        "actor_user_id": actor.get("id"),
        "actor_email": normalize_admin_email(actor.get("email")),
        "action": action,
        "target_type": target_type,
        "target_id": target_id,
        "metadata": metadata or {},
        "created_at": now_utc_iso(),
    }

    try:
        supabase.table("identity_audit_logs").insert(payload).execute()
    except Exception as error:
        print("[identity admin] audit insert failed:", str(error), flush=True)
        raise HTTPException(status_code=500, detail="Could not record admin action right now.")


def insert_identity_audit_log_best_effort(
    supabase: Any,
    actor: dict,
    action: str,
    target_type: str,
    target_id: str | None = None,
    metadata: dict | None = None,
) -> None:
    try:
        insert_identity_audit_log(supabase, actor, action, target_type, target_id, metadata)
    except HTTPException as error:
        print("[identity audit] best-effort audit skipped:", error.detail, flush=True)


def ensure_initial_admin_roles(supabase: Any) -> None:
    global INITIAL_ADMIN_ROLES_SEEDED

    if INITIAL_ADMIN_ROLES_SEEDED:
        return

    system_actor = {"id": None, "email": "system"}

    for email, role in INITIAL_ADMIN_ROLES.items():
        normalized_email = normalize_admin_email(email)
        existing = fetch_admin_user_by_email(supabase, normalized_email)

        if existing:
            continue

        now_iso = now_utc_iso()
        try:
            insert_admin_user_row(supabase, {
                "email": normalized_email,
                "role": role,
                "active": True,
                "created_at": now_iso,
                "updated_at": now_iso,
            })
            record_admin_role_change_entry(supabase, normalized_email, None, role, system_actor, "Initial role seed")
            insert_identity_audit_log_best_effort(
                supabase,
                system_actor,
                "ADMIN_ROLE_INITIAL_SEED",
                "ADMIN_USER",
                normalized_email,
                {"new_role": role},
            )
        except HTTPException:
            raise
        except Exception as error:
            print("[identity admin] initial admin seed failed:", str(error), flush=True)
            raise HTTPException(status_code=500, detail="Could not prepare admin access right now.")

    INITIAL_ADMIN_ROLES_SEEDED = True


def get_current_user_email(request: Request) -> str:
    identity = get_authenticated_identity(request)
    email = normalize_admin_email(identity.get("email"))

    if not email:
        raise HTTPException(status_code=401, detail="Authentication required.")

    return email


def require_admin_role(request: Request, allowed_roles: set[str] | None = None) -> dict:
    identity = get_authenticated_identity(request)
    email = normalize_admin_email(identity.get("email"))

    if not email:
        raise HTTPException(status_code=401, detail="Authentication required.")

    identity["email"] = email
    supabase = get_supabase_client()
    ensure_initial_admin_roles(supabase)
    admin_user = fetch_admin_user_by_email(supabase, identity["email"])

    if not admin_user:
        raise HTTPException(status_code=403, detail="Admin access required.")

    if not get_admin_active_value(admin_user):
        raise HTTPException(status_code=403, detail="Admin access required.")

    role = normalize_admin_role(admin_user.get("role")) or get_admin_role_by_email_rpc(supabase, identity["email"])

    if role not in IDENTITY_ADMIN_ROLES:
        raise HTTPException(status_code=403, detail="Admin access required.")

    if allowed_roles and role not in allowed_roles:
        raise HTTPException(status_code=403, detail="You do not have permission for this action.")

    return {
        **identity,
        "role": role,
        "admin_user": admin_user,
    }


def require_identity_admin(request: Request, allowed_roles: set[str] | None = None) -> dict:
    return require_admin_role(request, allowed_roles)


def upsert_admin_user_role(
    supabase: Any,
    target_email: str,
    new_role: str,
    actor: dict,
    note: str,
) -> dict:
    normalized_email = normalize_admin_email(target_email)
    existing = fetch_admin_user_by_email(supabase, normalized_email)
    old_role = normalize_admin_role(existing.get("role")) if existing else None
    now_iso = now_utc_iso()

    if existing:
        result = update_admin_user_row(supabase, normalized_email, {
            "role": new_role,
            "active": True,
            "updated_at": now_iso,
        })
        row = get_first_row(result) or {"email": normalized_email, "role": new_role, "active": True}
    else:
        result = insert_admin_user_row(supabase, {
            "email": normalized_email,
            "role": new_role,
            "active": True,
            "created_at": now_iso,
            "updated_at": now_iso,
        })
        row = get_first_row(result) or {"email": normalized_email, "role": new_role, "active": True}

    record_admin_role_change_entry(supabase, normalized_email, old_role, new_role, actor, note)
    insert_identity_audit_log(
        supabase,
        actor,
        "ADMIN_ROLE_ASSIGNED",
        "ADMIN_USER",
        normalized_email,
        {"old_role": old_role, "new_role": new_role, "note": note.strip() or None},
    )
    return row


def set_admin_user_active(
    supabase: Any,
    target_email: str,
    active: bool,
    actor: dict,
    note: str,
) -> dict:
    normalized_email = normalize_admin_email(target_email)
    existing = fetch_admin_user_by_email(supabase, normalized_email)

    if not existing:
        raise HTTPException(status_code=404, detail="Admin user not found.")

    role = normalize_admin_role(existing.get("role"))
    now_iso = now_utc_iso()
    result = update_admin_user_row(supabase, normalized_email, {
        "active": active,
        "updated_at": now_iso,
    })
    row = get_first_row(result) or {**existing, "active": active, "updated_at": now_iso}
    action = "ADMIN_ENABLED" if active else "ADMIN_DISABLED"

    record_admin_role_change_entry(
        supabase,
        normalized_email,
        role,
        role,
        actor,
        note,
        action,
    )
    insert_identity_audit_log(
        supabase,
        actor,
        action,
        "ADMIN_USER",
        normalized_email,
        {"role": role, "active": active, "note": note.strip() or None},
    )
    return row


def get_verification_request_row(supabase: Any, request_id: str) -> dict:
    result = supabase.table("verification_requests").select("*").eq("id", request_id).limit(1).execute()
    row = get_first_row(result)

    if not row:
        raise HTTPException(status_code=404, detail="Verification request not found.")

    return row


def get_verification_claimed_user_id(verification_row: dict, payload: AdminVerificationDecisionRequest) -> str | None:
    return (
        payload.claimed_by_user_id
        or verification_row.get("claimed_by_user_id")
        or verification_row.get("requester_user_id")
        or verification_row.get("user_id")
    )


def update_reserved_identity_verified(
    supabase: Any,
    verification_row: dict,
    claimed_by_user_id: str | None,
) -> tuple[str, dict]:
    request_type = str(verification_row.get("request_type") or "").strip().upper()
    requested_name = str(verification_row.get("requested_name") or "")
    normalized_key = normalize_identity_key(requested_name)

    if request_type not in {"PERSON", "BRAND"} or not normalized_key:
        raise HTTPException(status_code=400, detail="Verification request is missing identity details.")

    table_name = "reserved_people" if request_type == "PERSON" else "reserved_brands"
    existing = find_reserved_row_by_key(supabase, table_name, normalized_key)

    if not existing:
        raise HTTPException(status_code=404, detail="Reserved identity record not found.")

    update_payload = {
        "verified": True,
        "claimed_by_user_id": claimed_by_user_id,
        "updated_at": now_utc_iso(),
    }
    result = (
        supabase.table(table_name)
        .update(update_payload)
        .eq("normalized_key", normalized_key)
        .execute()
    )
    row = get_first_row(result) or existing
    return table_name, row


def get_excel_header_map(row: tuple[Any, ...]) -> dict[str, int]:
    return {
        str(value or "").strip(): index
        for index, value in enumerate(row)
        if str(value or "").strip()
    }


def get_excel_cell(row: tuple[Any, ...], header_map: dict[str, int], header: str) -> str:
    index = header_map.get(header)

    if index is None or index >= len(row):
        return ""

    value = row[index]
    return str(value or "").strip()


def find_excel_header_row(sheet: Any, required_headers: list[str]) -> tuple[int, dict[str, int]]:
    max_scan_row = min(sheet.max_row, 30)

    for row_number in range(1, max_scan_row + 1):
        row = tuple(cell.value for cell in sheet[row_number])
        header_map = get_excel_header_map(row)

        if all(header in header_map for header in required_headers):
            return row_number, header_map

    raise ValueError("The Excel file does not match the expected template.")


def build_reserved_rows_from_sheet(sheet: Any, config: dict) -> list[dict]:
    header_row_number, header_map = find_excel_header_row(sheet, config["required_headers"])
    rows: list[dict] = []
    seen_keys: set[str] = set()

    for row in sheet.iter_rows(min_row=header_row_number + 1, values_only=True):
        display_value = get_excel_cell(row, header_map, config["display_header"])
        category_value = get_excel_cell(row, header_map, config["category_header"])
        raw_key = get_excel_cell(row, header_map, config["normalized_header"])
        normalized_key = normalize_identity_key(raw_key or display_value)

        if not display_value or not normalized_key or normalized_key in seen_keys:
            continue

        seen_keys.add(normalized_key)
        rows.append({
            config["name_column"]: display_value,
            config["category_column"]: category_value or None,
            "normalized_key": normalized_key,
            "active": True,
            "updated_at": now_utc_iso(),
        })

    return rows


def upsert_reserved_rows(supabase: Any, table_name: str, rows: list[dict]) -> int:
    if not rows:
        return 0

    try:
        result = supabase.table(table_name).upsert(rows, on_conflict="normalized_key").execute()
        return len(result.data or rows)
    except Exception as error:
        if not is_missing_column_error(error, ["active", "updated_at"]):
            raise

    stripped_rows = [
        {key: value for key, value in row.items() if key not in {"active", "updated_at"}}
        for row in rows
    ]
    result = supabase.table(table_name).upsert(stripped_rows, on_conflict="normalized_key").execute()
    return len(result.data or stripped_rows)


def get_reserved_import_configs() -> dict[str, dict]:
    return {
        "Master Clean List": {
            "type": "people",
            "table": "reserved_people",
            "required_headers": [
                "Display Name",
                "Source Tab Category",
                "Normalized Value (No Spaces/Lowercase)",
            ],
            "display_header": "Display Name",
            "category_header": "Source Tab Category",
            "normalized_header": "Normalized Value (No Spaces/Lowercase)",
            "name_column": "display_name",
            "category_column": "category",
        },
        "Master Clean Brand List": {
            "type": "brands",
            "table": "reserved_brands",
            "required_headers": [
                "Brand Display Name",
                "Source Industry Tab",
                "Normalized Key Value",
            ],
            "display_header": "Brand Display Name",
            "category_header": "Source Industry Tab",
            "normalized_header": "Normalized Key Value",
            "name_column": "brand_name",
            "category_column": "industry",
        },
        "Master Clean Org List": {
            "type": "organizations",
            "table": "reserved_brands",
            "required_headers": [
                "Organization Display Name",
                "Source Category Tab",
                "Normalized Key Token",
            ],
            "display_header": "Organization Display Name",
            "category_header": "Source Category Tab",
            "normalized_header": "Normalized Key Token",
            "name_column": "brand_name",
            "category_column": "industry",
        },
    }


GENERIC_RESERVED_IMPORT_CONFIGS = {
    "people": {
        "action": "IMPORT_RESERVED_PEOPLE",
        "table": "reserved_people",
        "preferred_sheet": "Master Clean List",
        "required_columns": [
            "Display Name",
            "Normalized Value (No Spaces/Lowercase)",
            "Source Tab Category",
        ],
        "display_columns": ["Display Name"],
        "normalized_columns": ["Normalized Value (No Spaces/Lowercase)"],
        "category_columns": ["Source Tab Category"],
        "aliases_columns": ["Aliases", "Alias", "Optional aliases", "Optional Aliases"],
        "field_map": {
            "display_name": "Display Name",
            "normalized_key": "Normalized Value (No Spaces/Lowercase)",
            "category": "Source Tab Category",
            "aliases": "Aliases",
            "source_import": "source_import",
        },
    },
    "brands": {
        "action": "IMPORT_RESERVED_BRANDS",
        "table": "reserved_brands",
        "preferred_sheet": "Master Clean Brand List",
        "required_columns": [
            "Brand Display Name",
            "Normalized Key Value",
            "Source Industry Tab",
        ],
        "display_columns": ["Brand Display Name", "Organization Display Name"],
        "normalized_columns": ["Normalized Key Value", "Normalized Key Token"],
        "category_columns": ["Source Industry Tab", "Source Category Tab"],
        "website_columns": ["Website", "Official Website", "URL", "Url"],
        "aliases_columns": ["Aliases", "Alias", "Optional aliases", "Optional Aliases"],
        "field_map": {
            "brand_name": "Brand Display Name",
            "normalized_key": "Normalized Key Value",
            "industry": "Source Industry Tab",
            "website": "Website",
            "aliases": "Aliases",
            "source_import": "source_import",
        },
    },
}


def normalize_excel_header(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip()).lower()


def build_excel_column_lookup(columns: Any) -> dict[str, str]:
    return {
        normalize_excel_header(column): str(column)
        for column in columns
        if str(column or "").strip()
    }


def resolve_excel_column(column_lookup: dict[str, str], candidates: list[str]) -> str | None:
    for candidate in candidates:
        column = column_lookup.get(normalize_excel_header(candidate))

        if column:
            return column

    return None


def dataframe_has_required_columns(dataframe: Any, required_columns: list[str]) -> bool:
    column_lookup = build_excel_column_lookup(dataframe.columns)
    return all(resolve_excel_column(column_lookup, [column]) for column in required_columns)


def dataframe_matches_reserved_import_config(dataframe: Any, config: dict) -> bool:
    column_lookup = build_excel_column_lookup(dataframe.columns)
    return bool(
        resolve_excel_column(column_lookup, config["display_columns"])
        and resolve_excel_column(column_lookup, config["normalized_columns"])
        and resolve_excel_column(column_lookup, config["category_columns"])
    )


def find_reserved_identity_dataframe(workbook_sheets: dict[str, Any], config: dict) -> tuple[str, Any]:
    preferred_sheet = config["preferred_sheet"]

    if preferred_sheet in workbook_sheets and dataframe_matches_reserved_import_config(
        workbook_sheets[preferred_sheet],
        config,
    ):
        return preferred_sheet, workbook_sheets[preferred_sheet]

    for sheet_name, dataframe in workbook_sheets.items():
        if dataframe_matches_reserved_import_config(dataframe, config):
            return sheet_name, dataframe

    raise HTTPException(
        status_code=400,
        detail=(
            f"Could not find a worksheet with required columns: "
            f"{', '.join(config['required_columns'])}."
        ),
    )


def normalize_excel_string(value: Any) -> str:
    if value is None:
        return ""

    try:
        import pandas as pd

        if pd.isna(value):
            return ""
    except Exception:
        pass

    return str(value).strip()


def split_aliases(value: Any) -> list[str]:
    raw_value = normalize_excel_string(value)

    if not raw_value:
        return []

    aliases = [
        alias.strip()
        for alias in re.split(r"[,;\n\r|]+", raw_value)
        if alias.strip()
    ]
    deduped_aliases: list[str] = []
    seen_aliases: set[str] = set()

    for alias in aliases:
        normalized_alias = normalize_identity_key(alias)

        if not normalized_alias or normalized_alias in seen_aliases:
            continue

        seen_aliases.add(normalized_alias)
        deduped_aliases.append(alias)

    return deduped_aliases


def merge_aliases(existing_aliases: list[str], next_aliases: list[str]) -> list[str]:
    merged: list[str] = []
    seen_aliases: set[str] = set()

    for alias in [*existing_aliases, *next_aliases]:
        normalized_alias = normalize_identity_key(alias)

        if not normalized_alias or normalized_alias in seen_aliases:
            continue

        seen_aliases.add(normalized_alias)
        merged.append(alias)

    return merged


def get_dataframe_value(row: Any, column_name: str | None) -> str:
    if not column_name:
        return ""

    return normalize_excel_string(row.get(column_name))


def build_generic_reserved_import_rows(
    dataframe: Any,
    config: dict,
    filename: str,
    sheet_name: str,
) -> tuple[list[dict], int, list[str]]:
    column_lookup = build_excel_column_lookup(dataframe.columns)
    display_column = resolve_excel_column(column_lookup, config["display_columns"])
    normalized_column = resolve_excel_column(column_lookup, config["normalized_columns"])
    category_column = resolve_excel_column(column_lookup, config["category_columns"])
    website_column = resolve_excel_column(column_lookup, config.get("website_columns", []))
    aliases_column = resolve_excel_column(column_lookup, config.get("aliases_columns", []))
    rows_by_key: dict[str, dict] = {}
    duplicates = 0
    errors: list[str] = []
    source_import = f"{filename}:{sheet_name}"

    for index, row in dataframe.iterrows():
        row_number = int(index) + 2
        display_value = get_dataframe_value(row, display_column)
        raw_normalized_key = get_dataframe_value(row, normalized_column)
        normalized_key = normalize_identity_key(raw_normalized_key or display_value)

        if not display_value and not normalized_key:
            continue

        if not normalized_key:
            errors.append(f"Row {row_number}: missing normalized key and display name.")
            continue

        payload = {
            "normalized_key": normalized_key,
            "aliases": split_aliases(get_dataframe_value(row, aliases_column)),
            "source_import": source_import,
        }

        if config["table"] == "reserved_people":
            payload["display_name"] = display_value
            payload["category"] = get_dataframe_value(row, category_column) or None
        else:
            payload["brand_name"] = display_value
            payload["industry"] = get_dataframe_value(row, category_column) or None
            payload["website"] = get_dataframe_value(row, website_column) or None

        if normalized_key in rows_by_key:
            duplicates += 1
            existing_payload = rows_by_key[normalized_key]
            existing_payload["aliases"] = merge_aliases(existing_payload.get("aliases") or [], payload["aliases"])

            for key, value in payload.items():
                if key == "aliases":
                    continue

                if value not in ("", None):
                    existing_payload[key] = value

            continue

        rows_by_key[normalized_key] = payload

    return list(rows_by_key.values()), duplicates, errors


def fetch_existing_reserved_keys(supabase: Any, table_name: str, normalized_keys: list[str]) -> set[str]:
    existing_keys: set[str] = set()

    for index in range(0, len(normalized_keys), 500):
        batch = normalized_keys[index:index + 500]

        if not batch:
            continue

        result = (
            supabase.table(table_name)
            .select("normalized_key")
            .in_("normalized_key", batch)
            .execute()
        )

        for row in result.data or []:
            normalized_key = str(row.get("normalized_key") or "").strip()

            if normalized_key:
                existing_keys.add(normalized_key)

    return existing_keys


def upsert_generic_reserved_rows(supabase: Any, table_name: str, rows: list[dict]) -> None:
    for index in range(0, len(rows), 500):
        batch = rows[index:index + 500]

        if not batch:
            continue

        supabase.table(table_name).upsert(batch, on_conflict="normalized_key").execute()


def run_generic_reserved_identity_import(
    supabase: Any,
    actor: dict,
    upload_file: UploadFile,
    contents: bytes,
    import_type: str,
) -> dict:
    normalized_type = str(import_type or "").strip().lower()
    config = GENERIC_RESERVED_IMPORT_CONFIGS.get(normalized_type)

    if not config:
        raise HTTPException(status_code=400, detail="Import type must be people or brands.")

    filename = upload_file.filename or "reserved-identities.xlsx"

    if not filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Please upload an Excel .xlsx file.")

    if not contents:
        raise HTTPException(status_code=400, detail="Excel file is required.")

    try:
        import pandas as pd
    except ImportError:
        raise HTTPException(status_code=500, detail="Excel import is not configured right now.")

    try:
        workbook_sheets = pd.read_excel(BytesIO(contents), sheet_name=None, dtype=str)
    except Exception as error:
        print("[generic reserved import] pandas read failed:", str(error), flush=True)
        raise HTTPException(status_code=400, detail="Could not read the Excel file.")

    sheet_name, dataframe = find_reserved_identity_dataframe(workbook_sheets, config)
    rows, duplicates, row_errors = build_generic_reserved_import_rows(dataframe, config, filename, sheet_name)

    if row_errors:
        return {
            "success": False,
            "inserted": 0,
            "updated": 0,
            "duplicates": duplicates,
            "errors": row_errors[:50],
        }

    if not rows:
        raise HTTPException(status_code=400, detail="No importable rows were found in the Excel file.")

    normalized_keys = [row["normalized_key"] for row in rows]

    try:
        existing_keys = fetch_existing_reserved_keys(supabase, config["table"], normalized_keys)
        updated = sum(1 for key in normalized_keys if key in existing_keys)
        inserted = len(rows) - updated
        upsert_generic_reserved_rows(supabase, config["table"], rows)
        imported_at = now_utc_iso()
        insert_identity_audit_log(
            supabase,
            actor,
            config["action"],
            "RESERVED_IMPORT",
            filename,
            {
                "filename": filename,
                "sheet_name": sheet_name,
                "type": normalized_type,
                "rows_imported": len(rows),
                "rows_inserted": inserted,
                "rows_updated": updated,
                "duplicates": duplicates,
                "imported_at": imported_at,
                "admin_email": actor["email"],
            },
        )
    except HTTPException:
        raise
    except Exception as error:
        print("[generic reserved import] failed:", str(error), flush=True)
        raise HTTPException(status_code=503, detail="Could not import reserved identities right now.")

    return {
        "success": True,
        "inserted": inserted,
        "updated": updated,
        "duplicates": duplicates,
        "errors": [],
    }


@app.get("/", response_class=HTMLResponse)
def home():
    landing_path = Path(__file__).resolve().parents[1] / "client" / "landing" / "index.html"

    if landing_path.exists():
        return FileResponse(landing_path)

    return HTMLResponse(content=ABOUT_PAGE_HTML, status_code=200)


@app.get("/about", response_class=HTMLResponse)
def about_page():
    return HTMLResponse(content=ABOUT_PAGE_HTML, status_code=200)


@app.get("/privacy", response_class=HTMLResponse)
def privacy_page():
    return HTMLResponse(content=PRIVACY_POLICY_HTML, status_code=200)


@app.get("/personal-privacy", response_class=HTMLResponse)
def personal_privacy_page():
    return HTMLResponse(content=PERSONAL_PRIVACY_HTML, status_code=200)


@app.get("/terms", response_class=HTMLResponse)
def terms_page():
    return HTMLResponse(content=TERMS_OF_SERVICE_HTML, status_code=200)


@app.get("/copyright", response_class=HTMLResponse)
def copyright_page():
    return HTMLResponse(content=COPYRIGHT_NOTICE_HTML, status_code=200)


@app.get("/community-guidelines", response_class=HTMLResponse)
def community_guidelines_page():
    return HTMLResponse(content=COMMUNITY_GUIDELINES_HTML, status_code=200)


@app.get("/assets/icon/icon.png")
def app_icon():
    icon_path = Path(__file__).resolve().parents[1] / "assets" / "icon" / "icon.png"

    if not icon_path.exists():
        return HTMLResponse(content="", status_code=404)

    return FileResponse(icon_path, media_type="image/png")


@app.get("/claim/{claim_id}", response_class=HTMLResponse)
def public_claim_page(claim_id: str):
    normalized_claim_id = claim_id.strip()

    if not is_uuid(normalized_claim_id):
        return HTMLResponse(content=build_public_claim_404_page(), status_code=404)

    try:
        claim = fetch_public_claim_row(normalized_claim_id)
    except Exception as error:
        print("[public claim] fetch failed:", normalized_claim_id, error, flush=True)
        return HTMLResponse(content=build_public_claim_error_page(), status_code=503)

    # This public/SEO page reads via service role (bypasses RLS), so guard the
    # moderation flags explicitly: legacy `hidden`, the feed-gate `is_hidden`,
    # and soft-deleted `is_deleted` all must 404 so removed claims are never
    # served or indexed.
    if not claim or claim.get("hidden") or claim.get("is_hidden") or claim.get("is_deleted"):
        return HTMLResponse(content=build_public_claim_404_page(), status_code=404)

    return HTMLResponse(content=build_public_claim_page(claim), status_code=200)


@app.get("/health")
def health():
    # PHASE 4 STEP 21B
    return {"ok": True, "service": "Verifact backend", "version": "phase-4-step-21b"}


@app.get("/public-profile/{identifier}")
def public_profile(identifier: str, request: Request, username: str = ""):
    enforce_rate_limit(request, "public_profile", 180, AI_RATE_LIMIT_WINDOW_SECONDS)
    supabase = get_supabase_client()
    profile_row = lookup_public_profile_row(supabase, identifier, username)

    if not profile_row:
        return {
            "ok": False,
            "reason": "not_found",
            "message": "Contributor profile unavailable.",
        }

    return {
        "ok": True,
        "profile": build_public_profile_api_row(profile_row),
    }


@app.get("/profiles/{username}/posts")
def public_profile_posts(username: str, request: Request):
    enforce_rate_limit(request, "profile_posts", 180, AI_RATE_LIMIT_WINDOW_SECONDS)
    supabase, profile = get_public_profile_activity_context(request, username)
    posts = fetch_public_profile_posts(supabase, str(profile.get("id") or ""))
    return {"ok": True, "posts": posts, "count": len(posts)}


@app.get("/profiles/{username}/replies")
def public_profile_replies(username: str, request: Request):
    enforce_rate_limit(request, "profile_replies", 180, AI_RATE_LIMIT_WINDOW_SECONDS)
    supabase, profile = get_public_profile_activity_context(request, username)
    replies = fetch_public_profile_replies(supabase, str(profile.get("id") or ""))
    return {"ok": True, "replies": replies, "count": len(replies)}


@app.get("/profiles/{username}/evidence")
def public_profile_evidence(username: str, request: Request):
    enforce_rate_limit(request, "profile_evidence", 180, AI_RATE_LIMIT_WINDOW_SECONDS)
    supabase, profile = get_public_profile_activity_context(request, username)
    evidence_items = fetch_public_profile_evidence(supabase, str(profile.get("id") or ""))
    return {"ok": True, "evidence": evidence_items, "count": len(evidence_items)}


@app.get("/profiles/{username}/summary")
def public_profile_summary(username: str, request: Request):
    enforce_rate_limit(request, "profile_summary", 180, AI_RATE_LIMIT_WINDOW_SECONDS)
    supabase, profile = get_public_profile_activity_context(request, username)
    return {"ok": True, **build_public_profile_summary(supabase, profile)}


@app.get("/profiles/me/votes")
def private_profile_votes(request: Request):
    user_id = get_authenticated_user_id(request)
    history = fetch_private_vote_history(get_supabase_client(), user_id)
    return {"ok": True, "votes": history, "count": len(history)}


@app.get("/search/mentions")
def search_mentions(request: Request, q: str = "", limit: int = 8):
    enforce_rate_limit(request, "search_mentions", 180, AI_RATE_LIMIT_WINDOW_SECONDS)
    query = normalize_mention_search_query(q)
    safe_limit = max(1, min(8, int(limit or 8)))

    if not query:
        return {"results": []}

    supabase = get_supabase_client()
    results = [
        *search_profile_mentions(supabase, query, safe_limit),
        *search_organization_mentions(supabase, query, safe_limit),
    ]
    results.sort(key=lambda item: (not bool(item.get("verified")), str(item.get("username") or "")))

    return {"results": results[:safe_limit]}


@app.post("/mentions/tags")
def save_mention_tags(request: Request, payload: MentionTagsRequest):
    authenticated_user_id = get_authenticated_user_id(request)
    target_id = str(payload.target_id or "").strip()

    if not target_id:
        raise HTTPException(status_code=400, detail="Missing target id.")

    mention_limit = 5 if payload.target_type == "claim" else 3
    usernames = extract_backend_mentions(payload.text, mention_limit)

    if not usernames:
        return {"ok": True, "created": 0}

    supabase = get_supabase_client()
    actor_profile = fetch_profile_row(supabase, authenticated_user_id) or {}
    actor_username = get_review_safe_backend_username(actor_profile.get("username"), authenticated_user_id)
    claim_id = target_id
    notification_body = truncate_notification_body(payload.text)

    if payload.target_type == "claim":
        claim_rows = safe_execute_table_query(
            supabase.table("claims").select("id,title").eq("id", target_id).limit(1),
            "claim target",
        )
        claim_row = (claim_rows or [None])[0]

        if claim_row:
            notification_body = truncate_notification_body(claim_row.get("title"))

    if payload.target_type == "evidence":
        evidence_rows = safe_execute_table_query(
            supabase.table("evidence").select("id,claim_id,note").eq("id", target_id).limit(1),
            "evidence target",
        )
        evidence_row = (evidence_rows or [None])[0]

        if not evidence_row:
            raise HTTPException(status_code=404, detail="Evidence not found.")

        claim_id = str(evidence_row.get("claim_id") or "")
        notification_body = truncate_notification_body(evidence_row.get("note"))

    profiles_by_username, organizations_by_slug = fetch_mention_targets_by_username(supabase, usernames)
    created_count = 0

    for username in usernames:
        profile_row = profiles_by_username.get(username)

        if profile_row:
            tagged_user_id = str(profile_row.get("id") or "")
            table_name = "claim_tags" if payload.target_type == "claim" else "evidence_tags"
            tag_payload = {
                "tagged_user_id": tagged_user_id,
                "tagged_username": username,
                **({"claim_id": target_id} if payload.target_type == "claim" else {"evidence_id": target_id}),
            }

            if insert_mention_tag_row(supabase, table_name, tag_payload):
                created_count += 1

                if tagged_user_id != authenticated_user_id:
                    insert_notification_record(
                        supabase,
                        tagged_user_id,
                        "mention_claim" if payload.target_type == "claim" else "mention_evidence",
                        (
                            f"@{actor_username} mentioned you"
                            if payload.target_type == "claim"
                            else f"@{actor_username} mentioned you in evidence"
                        ),
                        notification_body,
                        claim_id or None,
                    )

        organization_row = organizations_by_slug.get(username)

        if organization_row:
            table_name = "claim_tags" if payload.target_type == "claim" else "evidence_tags"
            tag_payload = {
                "tagged_org_id": str(organization_row.get("id") or ""),
                "tagged_username": username,
                **({"claim_id": target_id} if payload.target_type == "claim" else {"evidence_id": target_id}),
            }

            if insert_mention_tag_row(supabase, table_name, tag_payload):
                created_count += 1

    return {"ok": True, "created": created_count}


@app.post("/identity/check-username")
def identity_check_username(payload: IdentityUsernameCheckRequest):
    try:
        return build_reserved_username_check_response(payload.username)
    except HTTPException:
        raise
    except Exception as error:
        print("[identity check] failed:", str(error), flush=True)
        normalized_key = normalize_identity_key(payload.username)
        return {
            "available": True,
            "reserved": False,
            "message": None,
            "normalized_key": normalized_key,
            "warning": "Reserved identity check temporarily unavailable.",
        }


@app.get("/auth/username-availability")
def username_availability(username: str, request: Request):
    normalized_username = normalize_profile_username_value(username)

    if not is_valid_profile_username(normalized_username):
        return {
            "ok": True,
            "available": False,
            "normalized_username": normalized_username,
            "message": "Username must be 3-20 characters.",
        }

    excluded_user_id = ""
    authorization = request.headers.get("authorization", "")

    if authorization.lower().startswith("bearer "):
        try:
            excluded_user_id = get_authenticated_user_id(request)
        except HTTPException:
            excluded_user_id = ""

    try:
        return check_username_availability_for_save(normalized_username, excluded_user_id)
    except Exception as error:
        print("[username availability] failed:", str(error), flush=True)
        raise HTTPException(status_code=503, detail="Could not check username availability right now.")


@app.post("/profile/ensure")
def ensure_backend_profile(payload: ProfileEnsureRequest, request: Request):
    identity = get_authenticated_identity(request)
    auth_user = identity["user"]
    user_id = identity["id"]
    supabase = get_supabase_client()

    try:
        existing_profile = fetch_profile_row(supabase, user_id)

        if existing_profile:
            if existing_profile.get("is_deleted"):
                raise HTTPException(status_code=403, detail="This account has been deleted.")

            updates: dict[str, Any] = {}
            email_confirmed = bool(getattr(auth_user, "email_confirmed_at", None))

            if email_confirmed and not existing_profile.get("verified"):
                updates["verified"] = True

            if not existing_profile.get("public_profile_slug"):
                updates["public_profile_slug"] = generate_backend_profile_slug(existing_profile.get("username"), user_id)

            if updates:
                updates["updated_at"] = now_utc_iso()
                update_result = (
                    supabase.table("profiles")
                    .update(updates)
                    .eq("id", user_id)
                    .execute()
                )
                existing_profile = get_first_row(update_result) or {**existing_profile, **updates}

            return build_profile_response(existing_profile)

        preferred_username = normalize_profile_username_value(payload.username) or get_backend_preferred_username(auth_user)

        if not is_valid_profile_username(preferred_username):
            raise HTTPException(status_code=400, detail="Username must be 3-20 characters.")

        availability = check_username_availability_for_save(preferred_username, user_id)

        if not availability.get("available"):
            raise HTTPException(
                status_code=409,
                detail=availability.get("message") or "Username is not available.",
            )

        now_iso = now_utc_iso()
        insert_result = (
            supabase.table("profiles")
            .insert({
                "id": user_id,
                "username": preferred_username,
                "display_name": preferred_username,
                "public_profile_slug": generate_backend_profile_slug(preferred_username, user_id),
                "profile_visibility": "public",
                "verified": bool(getattr(auth_user, "email_confirmed_at", None)),
                "created_at": now_iso,
                "updated_at": now_iso,
            })
            .execute()
        )
        inserted_profile = get_first_row(insert_result) or fetch_profile_row(supabase, user_id)
        return build_profile_response(inserted_profile)
    except HTTPException:
        raise
    except Exception as error:
        print("[profile ensure] failed:", str(error), flush=True)
        raise HTTPException(status_code=503, detail="Could not save your profile right now.")


@app.patch("/profile")
def update_backend_profile(payload: ProfileUpdateRequest, request: Request):
    identity = get_authenticated_identity(request)
    user_id = identity["id"]
    supabase = get_supabase_client()
    raw_updates = payload.dict(exclude_unset=True)
    updates: dict[str, Any] = {}

    if "username" in raw_updates:
        normalized_username = normalize_profile_username_value(payload.username)

        if not is_valid_profile_username(normalized_username):
            raise HTTPException(status_code=400, detail="Username must be 3-20 characters.")

        try:
            availability = check_username_availability_for_save(normalized_username, user_id)
        except Exception as error:
            print("[profile update] username check failed:", str(error), flush=True)
            raise HTTPException(status_code=503, detail="Could not check username right now.")

        if not availability.get("available"):
            raise HTTPException(
                status_code=409,
                detail=availability.get("message") or "Username is not available.",
            )

        updates["username"] = normalized_username
        updates["display_name"] = normalized_username
        updates["public_profile_slug"] = generate_backend_profile_slug(normalized_username, user_id)

    if "display_name" in raw_updates and "username" not in raw_updates:
        display_name = str(payload.display_name or "").strip()[:80]
        updates["display_name"] = display_name or None

    if "avatar_url" in raw_updates:
        avatar_url = str(payload.avatar_url or "").strip() or None

        if not is_valid_backend_avatar_url(avatar_url):
            raise HTTPException(status_code=400, detail="Avatar URL must be a valid URL.")

        updates["avatar_url"] = avatar_url

    if "avatar_path" in raw_updates:
        updates["avatar_path"] = str(payload.avatar_path or "").strip() or None

    if "bio" in raw_updates:
        updates["bio"] = sanitize_backend_bio(payload.bio)

    if "profile_visibility" in raw_updates:
        updates["profile_visibility"] = normalize_backend_profile_visibility(payload.profile_visibility)

    if not updates:
        return build_profile_response(fetch_profile_row(supabase, user_id))

    updates["updated_at"] = now_utc_iso()

    try:
        result = (
            supabase.table("profiles")
            .update(updates)
            .eq("id", user_id)
            .execute()
        )
        updated_rows = len(result.data or [])
        supabase_error = getattr(result, "error", None)
        print(
            "[profile update] database update:",
            {
                "user_id": user_id,
                "fields": sorted(updates.keys()),
                "updated_rows": updated_rows,
                "supabase_error": str(supabase_error) if supabase_error else None,
            },
            flush=True,
        )
        row = get_first_row(result) or fetch_profile_row(supabase, user_id)

        if not row:
            raise HTTPException(status_code=404, detail="Profile not found.")

        return build_profile_response(row, updated_rows)
    except HTTPException:
        raise
    except Exception as error:
        print("[profile update] failed:", {"user_id": user_id, "supabase_error": str(error)}, flush=True)
        raise HTTPException(status_code=503, detail="Could not update your profile right now.")


@app.post("/verification/request")
def create_verification_request(payload: VerificationRequestPayload, request: Request):
    request_type = payload.request_type.strip().upper()
    requested_name = payload.requested_name.strip()

    if request_type not in {"PERSON", "BRAND"}:
        raise HTTPException(status_code=400, detail="Verification type must be PERSON or BRAND.")

    if not requested_name:
        raise HTTPException(status_code=400, detail="Requested name is required.")

    requester = {"id": None, "email": None}
    authorization = request.headers.get("authorization", "")

    if authorization.lower().startswith("bearer "):
        requester = get_authenticated_identity(request)

    now_iso = now_utc_iso()
    row_payload = {
        "request_type": request_type,
        "requested_name": requested_name,
        "normalized_key": normalize_identity_key(requested_name),
        "official_email": payload.official_email.strip() or None,
        "official_website": payload.official_website.strip() or None,
        "social_links": payload.social_links or [],
        "supporting_documents": payload.supporting_documents or [],
        "status": "Pending",
        "requester_user_id": requester.get("id"),
        "created_at": now_iso,
        "updated_at": now_iso,
    }
    supabase = get_supabase_client()

    try:
        result = supabase.table("verification_requests").insert(row_payload).execute()
        row = get_first_row(result)
        insert_identity_audit_log_best_effort(
            supabase,
            requester,
            "VERIFICATION_REQUEST_CREATED",
            "VERIFICATION_REQUEST",
            str(row.get("id")) if row else None,
            {"request_type": request_type, "requested_name": requested_name},
        )
        return {"ok": True, "request": row}
    except Exception as error:
        print("[verification request] failed:", str(error), flush=True)
        raise HTTPException(status_code=503, detail="Could not submit verification request right now.")


@app.post("/admin/reserved/import")
async def admin_reserved_import(request: Request, file: UploadFile = File(...)):
    actor = require_identity_admin(request, {"SUPER_ADMIN", "ADMIN"})

    try:
        from openpyxl import load_workbook
    except ImportError:
        raise HTTPException(status_code=500, detail="Excel import is not configured right now.")

    contents = await file.read()

    if not contents:
        raise HTTPException(status_code=400, detail="Excel file is required.")

    try:
        workbook = load_workbook(BytesIO(contents), read_only=True, data_only=True)
    except Exception as error:
        print("[reserved import] workbook load failed:", str(error), flush=True)
        raise HTTPException(status_code=400, detail="Could not read the Excel file.")

    configs = get_reserved_import_configs()
    supabase = get_supabase_client()
    imported: dict[str, int] = {}

    try:
        for sheet_name, config in configs.items():
            if sheet_name not in workbook.sheetnames:
                continue

            rows = build_reserved_rows_from_sheet(workbook[sheet_name], config)
            imported[config["type"]] = upsert_reserved_rows(supabase, config["table"], rows)

        if not imported:
            raise HTTPException(status_code=400, detail="No supported reserved identity sheet was found.")

        insert_identity_audit_log(
            supabase,
            actor,
            "RESERVED_IDENTITIES_IMPORTED",
            "RESERVED_IMPORT",
            file.filename,
            {"filename": file.filename, "imported": imported},
        )
        return {"ok": True, "filename": file.filename, "imported": imported}
    except HTTPException:
        raise
    except ValueError as error:
        raise HTTPException(status_code=400, detail=str(error))
    except Exception as error:
        print("[reserved import] failed:", str(error), flush=True)
        raise HTTPException(status_code=503, detail="Could not import reserved identities right now.")


@app.post("/admin/import/reserved-identities")
async def admin_import_reserved_identities(
    request: Request,
    file: UploadFile = File(...),
    import_type: str = Form(..., alias="type"),
):
    actor = require_admin_role(request, {"SUPER_ADMIN", "ADMIN"})
    contents = await file.read()
    return run_generic_reserved_identity_import(
        get_supabase_client(),
        actor,
        file,
        contents,
        import_type,
    )


@app.get("/admin/verification-requests")
def admin_list_verification_requests(request: Request, status: str = "Pending", limit: int = 100):
    require_identity_admin(request)
    safe_limit = max(1, min(200, int(limit or 100)))
    request_status = status.strip() if status else "Pending"
    supabase = get_supabase_client()

    try:
        query = (
            supabase.table("verification_requests")
            .select("*")
            .order("created_at", desc=True)
            .limit(safe_limit)
        )

        if request_status.upper() != "ALL":
            query = query.eq("status", request_status)

        result = query.execute()
        return {"ok": True, "requests": result.data or []}
    except Exception as error:
        print("[admin verification] list failed:", str(error), flush=True)
        raise HTTPException(status_code=503, detail="Could not load verification requests right now.")


@app.post("/admin/verification-requests/{request_id}/approve")
def admin_approve_verification_request(
    request_id: str,
    payload: AdminVerificationDecisionRequest,
    request: Request,
):
    actor = require_identity_admin(request, {"SUPER_ADMIN", "ADMIN"})
    supabase = get_supabase_client()
    now_iso = now_utc_iso()

    try:
        verification_row = get_verification_request_row(supabase, request_id)
        claimed_by_user_id = get_verification_claimed_user_id(verification_row, payload)
        reserved_table, reserved_row = update_reserved_identity_verified(
            supabase,
            verification_row,
            claimed_by_user_id,
        )
        update_payload = {
            "status": "Approved",
            "reviewed_by_email": actor["email"],
            "reviewed_at": now_iso,
            "decision_note": payload.decision_note.strip() or None,
            "claimed_by_user_id": claimed_by_user_id,
            "updated_at": now_iso,
        }
        result = (
            supabase.table("verification_requests")
            .update(update_payload)
            .eq("id", request_id)
            .execute()
        )
        updated_request = get_first_row(result)
        insert_identity_audit_log(
            supabase,
            actor,
            "VERIFICATION_REQUEST_APPROVED",
            "VERIFICATION_REQUEST",
            request_id,
            {
                "reserved_table": reserved_table,
                "reserved_identity_id": reserved_row.get("id"),
                "claimed_by_user_id": claimed_by_user_id,
            },
        )
        return {"ok": True, "request": updated_request, "reserved_identity": reserved_row}
    except HTTPException:
        raise
    except Exception as error:
        print("[admin verification] approve failed:", str(error), flush=True)
        raise HTTPException(status_code=503, detail="Could not approve verification request right now.")


@app.post("/admin/verification-requests/{request_id}/reject")
def admin_reject_verification_request(
    request_id: str,
    payload: AdminVerificationDecisionRequest,
    request: Request,
):
    actor = require_identity_admin(request)
    supabase = get_supabase_client()
    now_iso = now_utc_iso()

    try:
        get_verification_request_row(supabase, request_id)
        result = (
            supabase.table("verification_requests")
            .update({
                "status": "Rejected",
                "reviewed_by_email": actor["email"],
                "reviewed_at": now_iso,
                "decision_note": payload.decision_note.strip() or None,
                "updated_at": now_iso,
            })
            .eq("id", request_id)
            .execute()
        )
        updated_request = get_first_row(result)
        insert_identity_audit_log(
            supabase,
            actor,
            "VERIFICATION_REQUEST_REJECTED",
            "VERIFICATION_REQUEST",
            request_id,
            {"decision_note": payload.decision_note.strip() or None},
        )
        return {"ok": True, "request": updated_request}
    except HTTPException:
        raise
    except Exception as error:
        print("[admin verification] reject failed:", str(error), flush=True)
        raise HTTPException(status_code=503, detail="Could not reject verification request right now.")


@app.get("/admin/users")
def admin_list_identity_users(request: Request):
    require_admin_role(request, {"SUPER_ADMIN", "ADMIN"})
    supabase = get_supabase_client()

    try:
        result = supabase.table("admin_users").select("*").order("email").execute()
        return {"ok": True, "users": [sanitize_admin_user(row) for row in result.data or []]}
    except Exception as error:
        print("[identity admin] list users failed:", str(error), flush=True)
        raise HTTPException(status_code=503, detail="Could not load admin users right now.")


@app.get("/admin/me")
def admin_me(request: Request):
    actor = require_admin_role(request)
    return {
        "email": actor["email"],
        "role": actor["role"],
        "active": get_admin_active_value(actor["admin_user"]),
    }


@app.post("/admin/users/assign-role")
def admin_assign_identity_role(payload: AdminAssignRoleRequest, request: Request):
    actor = require_admin_role(request, {"SUPER_ADMIN", "ADMIN"})
    target_email = get_admin_action_target_email(payload)
    new_role = normalize_admin_role(payload.role)
    note = get_admin_action_note(payload)

    if not target_email or "@" not in target_email:
        raise HTTPException(status_code=400, detail="A valid email is required.")

    if new_role == "SUPER_ADMIN" or new_role not in {"ADMIN", "MODERATOR"}:
        raise HTTPException(status_code=403, detail="You do not have permission to assign that role.")

    supabase = get_supabase_client()

    if not actor_can_assign_role(supabase, actor, new_role):
        raise HTTPException(status_code=403, detail="You do not have permission to assign that role.")

    try:
        row = upsert_admin_user_role(supabase, target_email, new_role, actor, note)
        return {"ok": True, "user": sanitize_admin_user(row)}
    except HTTPException:
        raise
    except Exception as error:
        print("[identity admin] assign role failed:", str(error), flush=True)
        raise HTTPException(status_code=503, detail="Could not assign admin role right now.")


@app.post("/admin/users/disable")
def admin_disable_identity_user(payload: AdminRoleStatusRequest, request: Request):
    actor = require_admin_role(request, {"SUPER_ADMIN"})
    target_email = get_admin_action_target_email(payload)
    note = get_admin_action_note(payload)

    if not target_email or "@" not in target_email:
        raise HTTPException(status_code=400, detail="A valid email is required.")

    if target_email == actor["email"]:
        raise HTTPException(status_code=400, detail="You cannot disable your own admin account.")

    supabase = get_supabase_client()

    try:
        row = set_admin_user_active(supabase, target_email, False, actor, note)
        return {"ok": True, "user": sanitize_admin_user(row)}
    except HTTPException:
        raise
    except Exception as error:
        print("[identity admin] disable user failed:", str(error), flush=True)
        raise HTTPException(status_code=503, detail="Could not disable admin user right now.")


@app.post("/admin/users/enable")
def admin_enable_identity_user(payload: AdminRoleStatusRequest, request: Request):
    actor = require_admin_role(request, {"SUPER_ADMIN"})
    target_email = get_admin_action_target_email(payload)
    note = get_admin_action_note(payload)

    if not target_email or "@" not in target_email:
        raise HTTPException(status_code=400, detail="A valid email is required.")

    supabase = get_supabase_client()

    try:
        row = set_admin_user_active(supabase, target_email, True, actor, note)
        return {"ok": True, "user": sanitize_admin_user(row)}
    except HTTPException:
        raise
    except Exception as error:
        print("[identity admin] enable user failed:", str(error), flush=True)
        raise HTTPException(status_code=503, detail="Could not enable admin user right now.")


@app.get("/auth/callback", response_class=HTMLResponse)
def auth_callback_page():
    return HTMLResponse(content=AUTH_CALLBACK_HTML, status_code=200)


@app.get("/auth/confirmed", response_class=HTMLResponse)
def auth_confirmed_page():
    return HTMLResponse(content=AUTH_CALLBACK_HTML, status_code=200)


@app.get("/reset-password", response_class=HTMLResponse)
def reset_password_page():
    return HTMLResponse(content=RESET_PASSWORD_HTML, status_code=200)


@app.get("/auth/reset-password", response_class=HTMLResponse)
def auth_reset_password_page():
    return HTMLResponse(content=RESET_PASSWORD_HTML, status_code=200)


# PHASE 5 STEP 1B
@app.get("/leaderboard")
def leaderboard(request: Request, type: str = "monthly", limit: int = 20):
    enforce_rate_limit(request, "leaderboard", 120, AI_RATE_LIMIT_WINDOW_SECONDS)
    leaderboard_type = "monthly" if type != "alltime" else "alltime"
    # PHASE 5 STEP 1C
    safe_limit = max(1, min(50, int(limit or 50)))
    supabase = get_supabase_client()
    month_start = datetime.now(timezone.utc).replace(day=1, hour=0, minute=0, second=0, microsecond=0).isoformat()

    if leaderboard_type == "monthly":
        reset_result = (
            supabase.table("profiles")
            .update({
                "monthly_reputation_points": 0,
                "monthly_reset_at": datetime.now(timezone.utc).isoformat(),
            })
            .lt("monthly_reset_at", month_start)
            .execute()
        )
        print("[leaderboard] monthly reset rows:", len(reset_result.data or []), flush=True)
        order_column = "monthly_reputation_points"
    else:
        order_column = "reputation_points"

    result = (
        supabase.table("profiles")
        .select("id, username, display_name, avatar_url, public_profile_slug, trust_score, trust_tier, rank_title, highest_rank_achieved, reputation_points, monthly_reputation_points, badge_list")
        .order(order_column, desc=True)
        .order("trust_score", desc=True)
        .limit(safe_limit)
        .execute()
    )

    rows = result.data or []
    users = []
    for index, row in enumerate(rows):
        user_id = row.get("id")
        safe_username = get_review_safe_backend_username(row.get("username"), user_id)
        trust_score = row.get("trust_score") or 50
        current_rank = calculate_rank_title(trust_score)
        display_rank = resolve_display_rank(current_rank, row.get("highest_rank_achieved") or row.get("rank_title"))
        badges = row.get("badge_list") or []
        badge_count = len(badges) if isinstance(badges, list) else 0
        users.append({
            "rank_position": index + 1,
            "id": user_id,
            "username": safe_username,
            "display_name": get_review_safe_backend_display_name(row.get("display_name"), row.get("username"), user_id),
            "avatar_url": row.get("avatar_url"),
            "public_profile_slug": row.get("public_profile_slug") or generate_backend_profile_slug(safe_username, str(user_id or "")),
            "rank_title": display_rank,
            "current_rank_title": current_rank,
            "highest_rank_achieved": row.get("highest_rank_achieved") or display_rank,
            "reputation_points": row.get("reputation_points") or 0,
            "monthly_reputation_points": row.get("monthly_reputation_points") or 0,
            "badge_list": badges if isinstance(badges, list) else [],
            "badge_count": badge_count,
            "trust_score": trust_score,
        })

    next_reset = datetime.now(timezone.utc).replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    next_reset_month = 1 if next_reset.month == 12 else next_reset.month + 1
    next_reset_year = next_reset.year + 1 if next_reset.month == 12 else next_reset.year
    next_reset = next_reset.replace(year=next_reset_year, month=next_reset_month)

    return {
        "ok": True,
        "type": leaderboard_type,
        "limit": safe_limit,
        "next_monthly_reset_at": next_reset.isoformat(),
        "users": users,
    }


# PHASE 5 STEP 1D
@app.get("/profile/reputation-events")
def profile_reputation_events(request: Request, limit: int = 50):
    authenticated_user_id = get_authenticated_user_id(request)
    safe_limit = max(1, min(50, int(limit or 50)))
    supabase = get_supabase_client()

    result = (
        supabase.table("reputation_events")
        .select("event_type, points_delta, trust_delta, badge_unlocked, rank_before, rank_after, reason, created_at, claim_id")
        .eq("user_id", authenticated_user_id)
        .order("created_at", desc=True)
        .limit(safe_limit)
        .execute()
    )

    return {
        "ok": True,
        "limit": safe_limit,
        "events": result.data or [],
    }


# PHASE 5 STEP 2
# PHASE 5 STEP 4
@app.delete("/account")
def delete_account(request: Request):
    """Account deletion = anonymization, never destruction.

    GDPR Article 17 / CCPA: personal data erased, public contributions
    preserved as legitimate public interest record per Article 17(3)(d).
    Claims, votes, and evidence keep their foreign keys to the (anonymized)
    profile row; only PII is scrubbed. The profile row itself is never
    deleted, so no cascade toward claims can ever fire.

    PHASE 6 STEP 3 inspection notes (behavior unchanged, documented):
    - PII scrubbed on profiles: username, display_name, avatar_url, bio,
      public_profile_slug (profile_visibility forced private). email is NOT a
      profiles column — it lives in auth.users and is scrubbed via the GoTrue
      admin API below. No other PII columns exist on profiles (remaining
      columns are ids, timestamps, and reputation/trust metrics, which are
      intentionally preserved).
    - Kept untouched: id, created_at, reputation/accuracy scores, vote
      history, claim authorship foreign keys.
    """
    authenticated_user_id = get_authenticated_user_id(request)
    supabase = get_supabase_client()
    deleted_username = f"deleted_{authenticated_user_id.replace('-', '')[-8:]}"
    now_iso = datetime.now(timezone.utc).isoformat()

    try:
        result = (
            supabase.table("profiles")
            .update({
                "is_deleted": True,
                "deleted_at": now_iso,
                "username": deleted_username,
                "display_name": "Deleted User",
                "avatar_url": None,
                "bio": None,
                "public_profile_slug": None,
                "profile_visibility": "private",
                "updated_at": now_iso,
            })
            .eq("id", authenticated_user_id)
            .execute()
        )
    except Exception as error:
        print("[account delete] profile anonymize failed:", str(error), flush=True)
        raise HTTPException(status_code=500, detail="Could not delete account right now.")

    if not result.data:
        raise HTTPException(status_code=404, detail="Account profile not found.")

    # PHASE 6 STEP 3 (NEW, additive): scrub auth-level PII and invalidate all
    # active sessions. Non-fatal by design — the profile anonymization above
    # already succeeded and must not be rolled back by an auth API hiccup.
    try:
        scrub_auth_user_and_sign_out(authenticated_user_id)
    except Exception as error:
        print("[account delete] auth scrub failed (non-fatal):", str(error), flush=True)

    # Existing keys unchanged for current clients; message added additively.
    return {
        "ok": True,
        "mode": "anonymized",
        "message": "Account deleted. Your contributions remain as part of the public record.",
    }


# ═══════════════════════════════════════════════════════════════════════
# APPLE GUIDELINE 1.2 — USER BLOCKING + EULA ACCEPTANCE (NEW, additive)
# Requires supabase/sql/038_user_blocks_eula.sql to be run first.
# Auth follows get_authenticated_user_id (same as DELETE /account); all
# writes use the service-role client (get_supabase_client), same as every
# other endpoint in this file.
# ═══════════════════════════════════════════════════════════════════════


def insert_block_moderation_report(
    supabase: Any,
    blocker_id: str,
    blocked_id: str,
    source_claim_id: str | None,
) -> None:
    """Developer notification (Apple req. 3b): every block surfaces in the
    same moderation view as reports (GET /admin/reports).

    Inspection notes on why the row looks like this:
    - reports.reason has a CHECK constraint (019_phase5_step2_launch_reports)
      allowing only the fixed enum — a literal "User blocked" reason would be
      rejected, so reason = HARASSMENT_OR_ABUSE and "User blocked" goes in
      the note, which the admin view already displays.
    - reports_one_target_check forces PROFILE-target rows to have
      claim_id NULL, so the source claim id rides along in the note text
      instead of the claim_id column.
    - Non-fatal by design: the block row is the safety feature; a failed
      notification row must not fail the block.
    """
    note = f"User blocked (blocked user {blocked_id})"

    if source_claim_id:
        note += f" — triggered from claim {source_claim_id}"

    try:
        supabase.table("reports").insert({
            "user_id": blocker_id,
            "target_type": "PROFILE",
            "profile_id": blocked_id,
            "claim_id": None,
            "evidence_id": None,
            "reason": "HARASSMENT_OR_ABUSE",
            "note": note,
            "status": "OPEN",
        }).execute()
    except Exception as error:
        # Most likely reports_unique_profile_user: the blocker already
        # reported this profile. The block itself succeeded; log and move on.
        print(
            "[block] moderation report row skipped:",
            blocker_id, "->", blocked_id, str(error),
            flush=True,
        )


@app.post("/api/users/{user_id}/block")
def block_user(user_id: str, request: Request, payload: BlockUserRequest | None = None):
    """Block a user. Idempotent — blocking an already-blocked user is 200."""
    blocker_id = get_authenticated_user_id(request)
    blocked_id = user_id.strip()

    if not is_uuid(blocked_id):
        raise HTTPException(status_code=404, detail="User not found.")

    if blocked_id == blocker_id:
        raise HTTPException(status_code=400, detail="You cannot block yourself.")

    source_claim_id = (payload.source_claim_id or "").strip() if payload else ""

    if source_claim_id and not is_uuid(source_claim_id):
        source_claim_id = ""

    supabase = get_supabase_client()

    target_result = (
        supabase.table("profiles").select("id").eq("id", blocked_id).limit(1).execute()
    )

    if not (target_result.data or []):
        raise HTTPException(status_code=404, detail="User not found.")

    existing_result = (
        supabase.table("user_blocks")
        .select("id")
        .eq("blocker_id", blocker_id)
        .eq("blocked_id", blocked_id)
        .limit(1)
        .execute()
    )

    if existing_result.data:
        # Already blocked — idempotent success, no duplicate moderation row.
        return {"ok": True, "blocked_id": blocked_id}

    try:
        supabase.table("user_blocks").insert({
            "blocker_id": blocker_id,
            "blocked_id": blocked_id,
            "source_claim_id": source_claim_id or None,
        }).execute()
    except Exception as error:
        # Unique-violation race (double tap) still means "blocked" — treat
        # as success; anything else is a real failure.
        message = str(error)

        if "user_blocks_blocker_id_blocked_id_key" not in message and "duplicate" not in message.lower():
            print("[block] insert failed:", blocker_id, "->", blocked_id, message, flush=True)
            raise HTTPException(status_code=500, detail="Could not block this user right now.")

        return {"ok": True, "blocked_id": blocked_id}

    insert_block_moderation_report(supabase, blocker_id, blocked_id, source_claim_id or None)
    print("[block] user blocked:", blocker_id, "->", blocked_id, flush=True)
    return {"ok": True, "blocked_id": blocked_id}


@app.delete("/api/users/{user_id}/block")
def unblock_user(user_id: str, request: Request):
    """Unblock a user. Idempotent — unblocking a non-blocked user is 200."""
    blocker_id = get_authenticated_user_id(request)
    blocked_id = user_id.strip()

    if not is_uuid(blocked_id):
        raise HTTPException(status_code=404, detail="User not found.")

    try:
        supabase = get_supabase_client()
        (
            supabase.table("user_blocks")
            .delete()
            .eq("blocker_id", blocker_id)
            .eq("blocked_id", blocked_id)
            .execute()
        )
    except Exception as error:
        print("[block] delete failed:", blocker_id, "->", blocked_id, str(error), flush=True)
        raise HTTPException(status_code=500, detail="Could not unblock this user right now.")

    return {"ok": True}


@app.get("/api/users/me/blocks")
def list_my_blocks(request: Request):
    """Ids only — the frontend caches this list in ClaimsContext."""
    blocker_id = get_authenticated_user_id(request)

    try:
        supabase = get_supabase_client()
        result = (
            supabase.table("user_blocks")
            .select("blocked_id")
            .eq("blocker_id", blocker_id)
            .execute()
        )
    except Exception as error:
        print("[block] list failed:", blocker_id, str(error), flush=True)
        raise HTTPException(status_code=500, detail="Could not load blocked users right now.")

    blocked_ids = [row.get("blocked_id") for row in (result.data or []) if row.get("blocked_id")]
    return {"blocked_ids": blocked_ids}


@app.post("/api/users/me/accept-terms")
def accept_terms(request: Request):
    """Record EULA acceptance (Apple req. 1). Fire-and-forget from the client.

    Missing profile row is not an error: on brand-new signups the profile can
    be created after email verification (ensure_backend_profile), so a no-op
    update here just means the timestamp lands on the next login instead.
    """
    authenticated_user_id = get_authenticated_user_id(request)
    now_iso = datetime.now(timezone.utc).isoformat()

    try:
        supabase = get_supabase_client()
        result = (
            supabase.table("profiles")
            .update({"terms_accepted_at": now_iso})
            .eq("id", authenticated_user_id)
            .execute()
        )
    except Exception as error:
        print("[terms] accept failed:", authenticated_user_id, str(error), flush=True)
        raise HTTPException(status_code=500, detail="Could not record terms acceptance right now.")

    if not (result.data or []):
        print("[terms] no profile row yet for:", authenticated_user_id, flush=True)

    return {"ok": True}


# ADMIN METRICS DASHBOARD (NEW, additive)
# Requires supabase/sql/045_admin_metrics_snapshot.sql to be run first.
def normalize_admin_metrics_payload(value: Any) -> dict[str, Any]:
    if isinstance(value, str):
        try:
            value = json.loads(value)
        except json.JSONDecodeError:
            value = {}

    if isinstance(value, list):
        value = value[0] if value else {}

    return value if isinstance(value, dict) else {}


@app.get("/admin/metrics")
def admin_metrics(request: Request):
    require_admin_user(request)
    global _admin_metrics_cache_payload, _admin_metrics_cache_expires_at

    now_monotonic = monotonic()
    if _admin_metrics_cache_payload is not None and now_monotonic < _admin_metrics_cache_expires_at:
        return {
            "ok": True,
            **_admin_metrics_cache_payload,
            "cached": True,
            "cache_expires_in_seconds": max(0, int(_admin_metrics_cache_expires_at - now_monotonic)),
        }

    now_utc = datetime.now(timezone.utc)
    today_start = now_utc.replace(hour=0, minute=0, second=0, microsecond=0)
    week_start = now_utc - timedelta(days=7)
    month_start = now_utc - timedelta(days=30)
    supabase = get_supabase_client()

    try:
        result = supabase.rpc(
            "admin_metrics_snapshot",
            {
                "today_start": today_start.isoformat(),
                "week_start": week_start.isoformat(),
                "month_start": month_start.isoformat(),
            },
        ).execute()
    except Exception as error:
        print("[admin metrics] snapshot failed:", str(error), flush=True)
        raise HTTPException(
            status_code=503,
            detail="Admin metrics are unavailable until supabase/sql/045_admin_metrics_snapshot.sql is applied.",
        )

    payload = normalize_admin_metrics_payload(result.data)

    if not payload:
        raise HTTPException(status_code=500, detail="Could not load admin metrics right now.")

    _admin_metrics_cache_payload = payload
    _admin_metrics_cache_expires_at = monotonic() + ADMIN_METRICS_CACHE_TTL_SECONDS

    return {
        "ok": True,
        **payload,
        "cached": False,
        "cache_expires_in_seconds": ADMIN_METRICS_CACHE_TTL_SECONDS,
    }


# PHASE 5 STEP 2
@app.get("/admin/reports")
def admin_list_reports(request: Request, status: str = "OPEN", limit: int = 50):
    require_admin_user(request)
    safe_limit = max(1, min(100, int(limit or 50)))
    report_status = status.strip().upper() if status else "OPEN"
    supabase = get_supabase_client()

    query = (
        supabase.table("reports")
        .select("id,target_type,claim_id,evidence_id,profile_id,user_id,reason,note,status,created_at,updated_at,resolved_at,admin_note")
        .order("created_at", desc=True)
        .limit(safe_limit)
    )

    if report_status != "ALL":
        query = query.eq("status", report_status)

    result = query.execute()
    reports = result.data or []
    supabase = get_supabase_client()

    for report in reports:
        target_type = report.get("target_type") or "CLAIM"
        target = None
        if target_type == "CLAIM" and report.get("claim_id"):
            target_result = supabase.table("claims").select("id,title,author_id,hidden,hidden_reason,is_featured,created_at").eq("id", report["claim_id"]).execute()
            target = (target_result.data or [None])[0]
        elif target_type == "EVIDENCE" and report.get("evidence_id"):
            target_result = supabase.table("evidence").select("id,note,url,user_id,hidden,hidden_reason,created_at").eq("id", report["evidence_id"]).execute()
            target = (target_result.data or [None])[0]
        elif target_type == "PROFILE" and report.get("profile_id"):
            target_result = supabase.table("profiles").select("id,username,display_name,profile_visibility,is_suspended,created_at").eq("id", report["profile_id"]).execute()
            target = (target_result.data or [None])[0]

        report["target"] = target

    return {"ok": True, "reports": reports}


# PHASE 5 STEP 2
@app.post("/admin/reports/{report_id}/resolve")
def admin_resolve_report(report_id: str, payload: AdminReportActionRequest, request: Request):
    admin_user_id = require_admin_user(request)
    supabase = get_supabase_client()
    next_status = payload.status.strip().upper() if payload.status else "RESOLVED"

    if next_status not in {"OPEN", "REVIEWING", "RESOLVED", "DISMISSED"}:
        raise HTTPException(status_code=400, detail="Unsupported report status.")

    report_result = supabase.table("reports").select("*").eq("id", report_id).execute()
    report_row = (report_result.data or [None])[0]

    if not report_row:
        raise HTTPException(status_code=404, detail="Report not found.")

    if payload.hide_target:
        target_type = report_row.get("target_type") or "CLAIM"
        if target_type == "CLAIM" and report_row.get("claim_id"):
            # is_hidden is the flag the restrictive RLS uses to remove a claim
            # from feeds/search (see 040) and that the admin hidden list + inline
            # hide use; set it alongside legacy `hidden` so every path agrees.
            supabase.table("claims").update({
                "is_hidden": True,
                "hidden": True,
                "hidden_reason": payload.admin_note.strip() or "Removed for violating community guidelines.",
                "hidden_at": datetime.now(timezone.utc).isoformat(),
                "hidden_by": admin_user_id,
                "updated_at": datetime.now(timezone.utc).isoformat(),
            }).eq("id", report_row["claim_id"]).execute()
        elif target_type == "EVIDENCE" and report_row.get("evidence_id"):
            supabase.table("evidence").update({
                "hidden": True,
                "hidden_reason": payload.admin_note.strip() or "Removed for violating community guidelines.",
                "hidden_at": datetime.now(timezone.utc).isoformat(),
                "hidden_by": admin_user_id,
                "updated_at": datetime.now(timezone.utc).isoformat(),
            }).eq("id", report_row["evidence_id"]).execute()
        elif target_type == "PROFILE" and report_row.get("profile_id"):
            supabase.table("profiles").update({
                "profile_visibility": "private",
                "updated_at": datetime.now(timezone.utc).isoformat(),
            }).eq("id", report_row["profile_id"]).execute()

    update_result = (
        supabase.table("reports")
        .update({
            "status": next_status,
            "resolved_at": datetime.now(timezone.utc).isoformat() if next_status in {"RESOLVED", "DISMISSED"} else None,
            "resolved_by": admin_user_id if next_status in {"RESOLVED", "DISMISSED"} else None,
            "admin_note": payload.admin_note.strip() or None,
            "updated_at": datetime.now(timezone.utc).isoformat(),
        })
        .eq("id", report_id)
        .execute()
    )

    return {"ok": True, "report": (update_result.data or [None])[0]}


# PHASE 5 STEP 3
@app.post("/admin/content/hide")
def admin_hide_content(payload: AdminContentVisibilityRequest, request: Request):
    admin_user_id = require_admin_user(request)
    target_type = payload.target_type.strip().upper()
    target_id = payload.target_id.strip()
    reason = payload.reason.strip() or "Removed for violating community guidelines."
    supabase = get_supabase_client()

    if target_type not in {"CLAIM", "EVIDENCE"} or not target_id:
        raise HTTPException(status_code=400, detail="Unsupported content target.")

    table_name = "claims" if target_type == "CLAIM" else "evidence"
    update_fields = {
        "hidden": True,
        "hidden_reason": reason,
        "hidden_at": datetime.now(timezone.utc).isoformat(),
        "hidden_by": admin_user_id,
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }

    # Claims use is_hidden for the feed-filtering RLS (040) and the admin hidden
    # list; the inline hide sets it too. Keep both flags in agreement. Evidence
    # has no is_hidden column, so only set it for claims.
    if target_type == "CLAIM":
        update_fields["is_hidden"] = True

    result = (
        supabase.table(table_name)
        .update(update_fields)
        .eq("id", target_id)
        .execute()
    )

    return {"ok": True, "target_type": target_type, "target_id": target_id, "updated": len(result.data or [])}


# PHASE 5 STEP 3
@app.post("/admin/content/restore")
def admin_restore_content(payload: AdminContentVisibilityRequest, request: Request):
    require_admin_user(request)
    target_type = payload.target_type.strip().upper()
    target_id = payload.target_id.strip()
    supabase = get_supabase_client()

    if target_type not in {"CLAIM", "EVIDENCE"} or not target_id:
        raise HTTPException(status_code=400, detail="Unsupported content target.")

    table_name = "claims" if target_type == "CLAIM" else "evidence"
    update_fields = {
        "hidden": False,
        "hidden_reason": None,
        "hidden_at": None,
        "hidden_by": None,
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }

    # Mirror the hide path: clear is_hidden for claims so the claim returns to
    # feeds. Evidence has no is_hidden column.
    if target_type == "CLAIM":
        update_fields["is_hidden"] = False

    result = (
        supabase.table(table_name)
        .update(update_fields)
        .eq("id", target_id)
        .execute()
    )

    return {"ok": True, "target_type": target_type, "target_id": target_id, "updated": len(result.data or [])}


@app.post("/admin/claims/delete")
def admin_delete_claim(payload: AdminClaimActionRequest, request: Request):
    admin_user_id = require_admin_user(request)
    claim_id = payload.claim_id.strip()

    if not claim_id:
        raise HTTPException(status_code=400, detail="claim_id is required")

    supabase = get_supabase_client()
    reason = payload.reason.strip() or "Violation of Verifact Terms of Use"
    now_iso = datetime.now(timezone.utc).isoformat()

    # SOFT DELETE (fix for the intermittent "admin action could not be
    # completed"). The old hard DELETE raised a ForeignKeyViolation whenever the
    # claim was referenced by a row with a NO-ACTION foreign key —
    # saved_claims.claim_id, moderation_appeals.claim_id,
    # user_blocks.source_claim_id, or another claim's canonical_claim_id — which
    # is exactly why it worked on some claims and failed on others. An UPDATE can
    # never hit a foreign key, so this is 100% reliable regardless of what
    # references the claim.
    #
    # - is_deleted (+ deleted_at/by/reason): the semantic "removed by moderation"
    #   marker and the anchor for a non-destructive Restore. We deliberately do
    #   NOT touch `status`, so the claim's real verdict/state is preserved and a
    #   Restore is lossless.
    # - is_hidden/hidden: reuse the ALREADY-LIVE restrictive RLS policy
    #   ("Hide hidden claims from public", migration 040) that removes the row
    #   from every feed, search and topic for non-admin, non-author readers.
    #   Migration 051 adds a second restrictive policy ("Hide deleted claims")
    #   that also hides is_deleted rows from the AUTHOR's own views (my-claims,
    #   saved, single claim); admins still see everything.
    # Single UPDATE, returning author_id + title so we can still notify the author.
    result = (
        supabase.table("claims")
        .update({
            "is_deleted": True,
            "deleted_at": now_iso,
            "deleted_by": admin_user_id,
            "deleted_reason": reason,
            "is_hidden": True,
            "hidden": True,
            "hidden_reason": reason,
            "hidden_at": now_iso,
            "updated_at": now_iso,
        })
        .eq("id", claim_id)
        .execute()
    )

    updated_rows = result.data or []

    if not updated_rows:
        raise HTTPException(status_code=404, detail="Claim not found.")

    author_id = updated_rows[0].get("author_id")
    title_snippet = str(updated_rows[0].get("title") or "")[:80]

    # MODERATION NOTIFICATION — tell the author why their claim was removed
    # (Apple 1.2 moderation transparency). The row still exists (soft delete),
    # so the notification can safely reference claim_id. Failure NEVER fails the
    # action.
    if author_id:
        try:
            supabase.table("notifications").insert({
                "user_id": author_id,
                "type": "claim_removed",
                "title": "Your claim was removed",
                "body": (
                    f'Your claim "{title_snippet}..." was removed by moderation. '
                    f"Reason: {reason}. Repeated violations may result in account "
                    "suspension. See the Terms of Use for our content rules."
                ),
                "claim_id": claim_id,
            }).execute()
        except Exception as error:
            print(f"[admin delete] notification failed: {error}", flush=True)

    return {"ok": True, "claim_id": claim_id, "deleted": len(updated_rows)}


@app.post("/admin/claims/lock-voting")
def admin_lock_claim_voting(payload: AdminClaimActionRequest, request: Request):
    require_admin_user(request)
    claim_id = payload.claim_id.strip()

    if not claim_id:
        raise HTTPException(status_code=400, detail="claim_id is required")

    now_iso = datetime.now(timezone.utc).isoformat()
    supabase = get_supabase_client()
    result = (
        supabase.table("claims")
        .update({
            "status": "LOCKED",
            "phase4_locked": True,
            "vote_accept_until": now_iso,
            "score_lock_at": now_iso,
            "verdict_reason": payload.reason.strip() or "Voting locked by admin.",
            "updated_at": now_iso,
        })
        .eq("id", claim_id)
        .execute()
    )

    return {"ok": True, "claim_id": claim_id, "updated": len(result.data or [])}


@app.post("/admin/claims/feature")
def admin_feature_claim(payload: AdminClaimActionRequest, request: Request):
    admin_user_id = require_admin_user(request)
    claim_id = payload.claim_id.strip()

    if not claim_id:
        raise HTTPException(status_code=400, detail="claim_id is required")

    now_iso = datetime.now(timezone.utc).isoformat()
    featured = bool(payload.featured)
    supabase = get_supabase_client()
    result = (
        supabase.table("claims")
        .update({
            "is_featured": featured,
            "featured_at": now_iso if featured else None,
            "featured_by": admin_user_id if featured else None,
            "updated_at": now_iso,
        })
        .eq("id", claim_id)
        .execute()
    )

    return {"ok": True, "claim_id": claim_id, "featured": featured, "updated": len(result.data or [])}


@app.post("/admin/users/suspend")
def admin_suspend_user(payload: AdminUserActionRequest, request: Request):
    admin_user_id = require_admin_user(request)
    target_user_id = payload.user_id.strip()

    if not target_user_id:
        raise HTTPException(status_code=400, detail="user_id is required")

    suspended = bool(payload.suspended)

    if suspended and target_user_id == admin_user_id:
        raise HTTPException(status_code=400, detail="Admins cannot suspend their own account.")

    now_iso = datetime.now(timezone.utc).isoformat()
    supabase = get_supabase_client()
    result = (
        supabase.table("profiles")
        .update({
            "is_suspended": suspended,
            "suspended_at": now_iso if suspended else None,
            "suspended_by": admin_user_id if suspended else None,
            "suspension_reason": payload.reason.strip() if suspended and payload.reason.strip() else None,
            "updated_at": now_iso,
        })
        .eq("id", target_user_id)
        .execute()
    )

    # MODERATION NOTIFICATION (NEW, additive) — same pattern as the claim
    # deletion notification above. Only on actual suspension (not
    # unsuspension), only when the update touched a row; failure NEVER fails
    # the suspension — response below is unchanged.
    if suspended and (result.data or []):
        reason = payload.reason.strip() or "Violation of Verifact Terms of Use"

        try:
            supabase.table("notifications").insert({
                "user_id": target_user_id,
                "type": "account_suspended",
                "title": "Your account has been suspended",
                "body": (
                    f"Your account has been suspended by moderation. Reason: {reason}. "
                    "If you believe this is a mistake, contact "
                    "support@factfight.com."
                ),
                "claim_id": None,
            }).execute()
        except Exception as error:
            print(f"[admin suspend] notification failed: {error}", flush=True)

    return {"ok": True, "user_id": target_user_id, "suspended": suspended, "updated": len(result.data or [])}


# ═══════════════════════════════════════════════════════════════════════
# HIDE/UNHIDE CLAIMS + ADMIN MANAGEMENT DASHBOARD (NEW, additive)
# Requires supabase/sql/040_hide_claims_admin_manage.sql to be run first.
# Inspection notes:
# - is_hidden is the NEW flag that the restrictive RLS policy filters out
#   of feeds/search/topics for non-admin readers. The legacy `hidden` flag
#   (020) only renders a "Content removed" box client-side, so these
#   endpoints set/clear BOTH flags to keep the two mechanisms in agreement.
#   The legacy /admin/content/hide and /admin/content/restore endpoints are
#   untouched.
# - Notifications follow the /admin/claims/delete pattern: only after a
#   successful update, and failure NEVER fails the moderation action.
# ═══════════════════════════════════════════════════════════════════════


@app.post("/admin/claims/{claim_id}/hide")
def admin_hide_claim(claim_id: str, payload: AdminHideClaimRequest, request: Request):
    admin_user_id = require_admin_user(request)
    target_claim_id = claim_id.strip()

    if not target_claim_id:
        raise HTTPException(status_code=400, detail="claim_id is required")

    reason = payload.reason.strip() or "Hidden by moderation."
    supabase = get_supabase_client()

    claim_result = (
        supabase.table("claims")
        .select("author_id,title,is_hidden")
        .eq("id", target_claim_id)
        .limit(1)
        .execute()
    )
    claim_row = (claim_result.data or [None])[0]

    if not claim_row:
        raise HTTPException(status_code=404, detail="Claim not found.")

    was_hidden = bool(claim_row.get("is_hidden"))
    now_iso = datetime.now(timezone.utc).isoformat()
    result = (
        supabase.table("claims")
        .update({
            "is_hidden": True,
            "hidden": True,
            "hidden_reason": reason,
            "hidden_at": now_iso,
            "hidden_by": admin_user_id,
            "updated_at": now_iso,
        })
        .eq("id", target_claim_id)
        .execute()
    )

    author_id = claim_row.get("author_id")

    # Notify only on an actual visibility change — re-hiding an already
    # hidden claim must not spam the author.
    if (result.data or []) and author_id and not was_hidden:
        title_snippet = str(claim_row.get("title") or "")[:80]

        try:
            supabase.table("notifications").insert({
                "user_id": author_id,
                "type": "claim_hidden",
                "title": "Your claim was hidden",
                "body": (
                    f'Your claim "{title_snippet}..." was hidden by moderation. '
                    f"Reason: {reason}. It is no longer visible to other users, "
                    "but it may be restored after review."
                ),
                "claim_id": target_claim_id,
            }).execute()
        except Exception as error:
            print(f"[admin hide] notification failed: {error}", flush=True)

    return {"ok": True, "claim_id": target_claim_id, "is_hidden": True, "updated": len(result.data or [])}


@app.post("/admin/claims/{claim_id}/unhide")
def admin_unhide_claim(claim_id: str, request: Request):
    require_admin_user(request)
    target_claim_id = claim_id.strip()

    if not target_claim_id:
        raise HTTPException(status_code=400, detail="claim_id is required")

    supabase = get_supabase_client()

    claim_result = (
        supabase.table("claims")
        .select("author_id,title,is_hidden,hidden")
        .eq("id", target_claim_id)
        .limit(1)
        .execute()
    )
    claim_row = (claim_result.data or [None])[0]

    if not claim_row:
        raise HTTPException(status_code=404, detail="Claim not found.")

    was_hidden = bool(claim_row.get("is_hidden")) or bool(claim_row.get("hidden"))
    # Unhide also lifts a soft delete: a claim can be hidden AND soft-deleted, so
    # making it visible again must clear is_deleted too, otherwise it would be
    # publicly visible yet still flagged "deleted" (invariant: visible ⟹ not
    # deleted). This makes Unhide double as Restore, which is what the moderation
    # UI already assumes ("reversible via Unhide").
    result = (
        supabase.table("claims")
        .update({
            "is_hidden": False,
            "hidden": False,
            "hidden_reason": None,
            "hidden_at": None,
            "hidden_by": None,
            "is_deleted": False,
            "deleted_at": None,
            "deleted_by": None,
            "deleted_reason": None,
            "updated_at": datetime.now(timezone.utc).isoformat(),
        })
        .eq("id", target_claim_id)
        .execute()
    )

    author_id = claim_row.get("author_id")

    # Notify only when the claim was actually hidden before this call.
    if (result.data or []) and author_id and was_hidden:
        title_snippet = str(claim_row.get("title") or "")[:80]

        try:
            supabase.table("notifications").insert({
                "user_id": author_id,
                "type": "claim_restored",
                "title": "Your claim was restored",
                "body": (
                    f'Your claim "{title_snippet}..." was reviewed and restored. '
                    "It is visible to other users again."
                ),
                "claim_id": target_claim_id,
            }).execute()
        except Exception as error:
            print(f"[admin unhide] notification failed: {error}", flush=True)

    return {"ok": True, "claim_id": target_claim_id, "is_hidden": False, "updated": len(result.data or [])}


@app.post("/admin/claims/{claim_id}/restore")
def admin_restore_claim(claim_id: str, request: Request):
    # Explicit reversal of a soft delete (POST /admin/claims/delete). Distinct
    # from Unhide only in intent + the notification wording; both clear
    # is_deleted + is_hidden. Kept as its own route so the moderation UI can
    # label deleted rows "Restore" and so a granted claim_removed appeal can
    # reuse it the way claim_hidden appeals reuse Unhide.
    require_admin_user(request)
    target_claim_id = claim_id.strip()

    if not target_claim_id:
        raise HTTPException(status_code=400, detail="claim_id is required")

    supabase = get_supabase_client()

    claim_result = (
        supabase.table("claims")
        .select("author_id,title,is_deleted")
        .eq("id", target_claim_id)
        .limit(1)
        .execute()
    )
    claim_row = (claim_result.data or [None])[0]

    if not claim_row:
        raise HTTPException(status_code=404, detail="Claim not found.")

    was_deleted = bool(claim_row.get("is_deleted"))
    result = (
        supabase.table("claims")
        .update({
            "is_deleted": False,
            "deleted_at": None,
            "deleted_by": None,
            "deleted_reason": None,
            "is_hidden": False,
            "hidden": False,
            "hidden_reason": None,
            "hidden_at": None,
            "hidden_by": None,
            "updated_at": datetime.now(timezone.utc).isoformat(),
        })
        .eq("id", target_claim_id)
        .execute()
    )

    author_id = claim_row.get("author_id")

    # Notify only when the claim was actually deleted before this call.
    if (result.data or []) and author_id and was_deleted:
        title_snippet = str(claim_row.get("title") or "")[:80]

        try:
            supabase.table("notifications").insert({
                "user_id": author_id,
                "type": "claim_restored",
                "title": "Your claim was restored",
                "body": (
                    f'Your claim "{title_snippet}..." was reviewed and restored. '
                    "It is visible to other users again."
                ),
                "claim_id": target_claim_id,
            }).execute()
        except Exception as error:
            print(f"[admin restore] notification failed: {error}", flush=True)

    return {"ok": True, "claim_id": target_claim_id, "is_deleted": False, "updated": len(result.data or [])}


# ═══════════════════════════════════════════════════════════════════════
# ADMIN QUEUE LISTS (NEW, additive) — read-only lists for the admin profile
# page. Both are service-role reads that bypass the hide RLS, and both reuse
# require_admin_user. They do NOT change any existing moderation flow: the
# frontend rows call the existing hide/unhide/delete endpoints.
# ═══════════════════════════════════════════════════════════════════════


def _usernames_for_ids(supabase: Any, user_ids: list[str]) -> dict[str, str | None]:
    unique_ids = sorted({uid for uid in user_ids if uid})

    if not unique_ids:
        return {}

    profiles_result = (
        supabase.table("profiles")
        .select("id,username")
        .in_("id", unique_ids)
        .execute()
    )

    return {
        profile["id"]: profile.get("username")
        for profile in (profiles_result.data or [])
    }


@app.get("/admin/claims/hidden")
def admin_list_hidden_claims(request: Request, limit: int = 50):
    require_admin_user(request)
    safe_limit = max(1, min(50, int(limit or 50)))
    supabase = get_supabase_client()

    # Service-role read bypasses the restrictive is_hidden RLS policy, so the
    # admin always sees hidden claims. Newest hidden first (hidden_at desc,
    # nulls last for legacy rows hidden before hidden_at was populated).
    result = (
        supabase.table("claims")
        .select(
            "id,title,description,author_id,is_hidden,is_deleted,deleted_reason,"
            "hidden_reason,hidden_at,votes_true,votes_fake,votes_unsure,created_at"
        )
        .eq("is_hidden", True)
        # Soft-deleted claims leave the Hidden list — they are removals, not
        # hides. They remain restorable from the admin dashboard (moderation).
        .eq("is_deleted", False)
        .order("hidden_at", desc=True)
        .order("created_at", desc=True)
        .limit(safe_limit)
        .execute()
    )
    claims = result.data or []
    usernames = _usernames_for_ids(supabase, [claim.get("author_id") for claim in claims])

    for claim in claims:
        author_id = claim.get("author_id")
        claim["claim_id"] = claim.get("id")
        claim["author_username"] = usernames.get(author_id or "")

    return {"ok": True, "claims": claims}


@app.get("/admin/claims/reported")
def admin_list_reported_claims(request: Request, limit: int = 50):
    require_admin_user(request)
    safe_limit = max(1, min(50, int(limit or 50)))
    supabase = get_supabase_client()

    # Actionable queue: unresolved CLAIM reports only. Grouped by claim below.
    reports_result = (
        supabase.table("reports")
        .select("id,claim_id,user_id,reason,note,status,created_at")
        .eq("target_type", "CLAIM")
        .in_("status", ["OPEN", "REVIEWING"])
        .order("created_at", desc=True)
        .limit(500)
        .execute()
    )
    reports = [report for report in (reports_result.data or []) if report.get("claim_id")]

    # Group reports by claim, preserving newest-first order of first appearance.
    grouped: dict[str, dict] = {}
    claim_order: list[str] = []

    for report in reports:
        claim_id = report["claim_id"]

        if claim_id not in grouped:
            grouped[claim_id] = {"claim_id": claim_id, "reports": []}
            claim_order.append(claim_id)

        grouped[claim_id]["reports"].append(report)

    # Keep only the most recently reported claims (reports are already newest
    # first, so claim_order is by latest report time).
    claim_ids = claim_order[:safe_limit]

    if not claim_ids:
        return {"ok": True, "claims": []}

    claims_result = (
        supabase.table("claims")
        .select("id,title,description,author_id,is_hidden,created_at")
        .in_("id", claim_ids)
        # A deleted claim leaves the actionable reported queue (its row is
        # skipped below when claim_row is missing).
        .eq("is_deleted", False)
        .execute()
    )
    claims_by_id = {claim["id"]: claim for claim in (claims_result.data or [])}

    reporter_ids = [report.get("user_id") for report in reports]
    author_ids = [claim.get("author_id") for claim in claims_by_id.values()]
    usernames = _usernames_for_ids(supabase, reporter_ids + author_ids)

    items = []

    for claim_id in claim_ids:
        claim_reports = grouped[claim_id]["reports"]
        claim_row = claims_by_id.get(claim_id)

        # A report may reference a claim that was hard-deleted; skip it.
        if not claim_row:
            continue

        author_id = claim_row.get("author_id")
        report_details = [
            {
                "reason": report.get("reason"),
                "note": report.get("note"),
                "reporter_id": report.get("user_id"),
                "reporter_username": usernames.get(report.get("user_id") or ""),
                "created_at": report.get("created_at"),
            }
            for report in claim_reports
        ]
        reasons = []
        for detail in report_details:
            reason = detail.get("reason")
            if reason and reason not in reasons:
                reasons.append(reason)

        items.append({
            "claim_id": claim_id,
            "title": claim_row.get("title"),
            "description": claim_row.get("description"),
            "author_id": author_id,
            "author_username": usernames.get(author_id or ""),
            "is_hidden": bool(claim_row.get("is_hidden")),
            "report_count": len(claim_reports),
            "latest_report_at": claim_reports[0].get("created_at") if claim_reports else None,
            "reasons": reasons,
            "reports": report_details,
        })

    return {"ok": True, "claims": items}


# ═══════════════════════════════════════════════════════════════════════
# MODERATION APPEALS (NEW, additive)
# Requires supabase/sql/044_moderation_appeals.sql to be run first.
# Inspection notes:
# - Suspended users CAN call /api/appeals: get_authenticated_user_id only
#   validates the token (verified — no is_suspended check anywhere in it),
#   so no exemption is needed; only require_admin_user blocks suspended
#   accounts. Do NOT add a suspended check to these user endpoints.
# - Grant reversal REUSES the existing endpoint functions directly:
#   admin_unhide_claim() for claim_hidden and admin_suspend_user(
#   suspended=False) for account_suspended — nothing reimplemented, and
#   their claim_restored notifications keep firing as they do today.
# - claim_removed is now REVERSIBLE: /admin/claims/delete soft-deletes (sets
#   is_deleted + is_hidden), so a granted claim_removed appeal COULD restore the
#   claim via admin_restore_claim(). It is not wired to do so automatically yet;
#   a granted claim_removed appeal still records the outcome and notifies the
#   user only. (Follow-up: reuse admin_restore_claim here the way claim_hidden
#   appeals reuse admin_unhide_claim.)
# - appeal_resolved notifications follow the existing fail-soft pattern:
#   failure never fails the resolution.
# ═══════════════════════════════════════════════════════════════════════

APPEAL_ACTION_TYPES = {"claim_hidden", "claim_removed", "account_suspended"}
APPEAL_TEXT_MIN_CHARS = 20
APPEAL_TEXT_MAX_CHARS = 500
APPEAL_DENIED_COOLDOWN_DAYS = 30


def raise_appeals_unavailable(error: Exception):
    print("[appeals] moderation_appeals unavailable:", str(error), flush=True)
    raise HTTPException(
        status_code=503,
        detail="Appeals are not available until the moderation appeals migration is applied.",
    )


@app.post("/api/appeals")
def create_appeal(payload: AppealCreateRequest, request: Request):
    user_id = get_authenticated_user_id(request)
    action_type = payload.action_type.strip().lower()
    appeal_text = payload.appeal_text.strip()
    claim_id = (payload.claim_id or "").strip() or None
    notification_id = (payload.notification_id or "").strip() or None

    if action_type not in APPEAL_ACTION_TYPES:
        raise HTTPException(status_code=400, detail="Unsupported appeal action type.")

    if len(appeal_text) < APPEAL_TEXT_MIN_CHARS or len(appeal_text) > APPEAL_TEXT_MAX_CHARS:
        raise HTTPException(
            status_code=400,
            detail=f"Appeal text must be {APPEAL_TEXT_MIN_CHARS}-{APPEAL_TEXT_MAX_CHARS} characters.",
        )

    if action_type == "account_suspended":
        claim_id = None

    supabase = get_supabase_client()

    # One appeal per decision: claim_id identifies hidden claims; notification_id
    # is the durable link for hard-deleted claims whose notification has no
    # claim_id.
    existing_query = (
        supabase.table("moderation_appeals")
        .select("id, status, created_at, reviewed_at")
        .eq("user_id", user_id)
        .eq("action_type", action_type)
    )

    if claim_id:
        existing_query = existing_query.eq("claim_id", claim_id)
    elif notification_id:
        existing_query = existing_query.eq("notification_id", notification_id)
    else:
        existing_query = existing_query.is_("claim_id", "null").is_("notification_id", "null")

    try:
        existing_rows = existing_query.execute().data or []
    except Exception as error:
        raise_appeals_unavailable(error)

    for row in existing_rows:
        if row.get("status") == "pending":
            raise HTTPException(status_code=409, detail="You already have a pending appeal for this decision.")

        if row.get("status") == "denied":
            try:
                denied_at = datetime.fromisoformat(
                    str(row.get("reviewed_at") or row.get("created_at")).replace("Z", "+00:00")
                )
            except ValueError:
                continue

            age_days = (datetime.now(timezone.utc) - denied_at).days

            if age_days < APPEAL_DENIED_COOLDOWN_DAYS:
                raise HTTPException(
                    status_code=409,
                    detail=(
                        "An appeal for this decision was denied recently. "
                        f"You can appeal again in {APPEAL_DENIED_COOLDOWN_DAYS - age_days} days."
                    ),
                )

    try:
        insert_result = (
            supabase.table("moderation_appeals")
            .insert({
                "user_id": user_id,
                "action_type": action_type,
                "claim_id": claim_id,
                "notification_id": notification_id,
                "appeal_text": appeal_text,
                "status": "pending",
            })
            .execute()
        )
    except Exception as error:
        raise_appeals_unavailable(error)
    appeal_row = (insert_result.data or [None])[0]

    if not appeal_row:
        raise HTTPException(status_code=500, detail="Could not submit the appeal.")

    return {"ok": True, "appeal": appeal_row}


@app.get("/api/appeals/mine")
def list_my_appeals(request: Request):
    user_id = get_authenticated_user_id(request)
    supabase = get_supabase_client()
    try:
        result = (
            supabase.table("moderation_appeals")
            .select("id, action_type, claim_id, notification_id, status, appeal_text, review_note, created_at, reviewed_at")
            .eq("user_id", user_id)
            .order("created_at", desc=True)
            .execute()
        )
    except Exception as error:
        raise_appeals_unavailable(error)

    return {"ok": True, "appeals": result.data or []}


@app.get("/admin/appeals")
def admin_list_appeals(request: Request, status: str = "all", limit: int = 100):
    require_admin_user(request)
    status_filter = (status or "all").strip().lower()
    supabase = get_supabase_client()
    query = (
        supabase.table("moderation_appeals")
        .select("*")
        .order("created_at", desc=True)
        .limit(max(1, min(limit, 200)))
    )

    if status_filter in {"pending", "granted", "denied"}:
        query = query.eq("status", status_filter)

    try:
        appeals = query.execute().data or []
    except Exception as error:
        raise_appeals_unavailable(error)

    # Attach usernames + claim titles for display; lookup failure only
    # degrades labels, never the list.
    try:
        user_ids = list({row["user_id"] for row in appeals if row.get("user_id")})
        claim_ids = list({row["claim_id"] for row in appeals if row.get("claim_id")})
        profiles_by_id: dict = {}
        claims_by_id: dict = {}

        if user_ids:
            profile_rows = (
                supabase.table("profiles").select("id, username, display_name").in_("id", user_ids).execute()
            )
            profiles_by_id = {row["id"]: row for row in profile_rows.data or []}

        if claim_ids:
            claim_rows = (
                supabase.table("claims").select("id, title, is_hidden").in_("id", claim_ids).execute()
            )
            claims_by_id = {row["id"]: row for row in claim_rows.data or []}

        for row in appeals:
            profile = profiles_by_id.get(row.get("user_id")) or {}
            claim = claims_by_id.get(row.get("claim_id")) or {}
            row["username"] = profile.get("username")
            row["display_name"] = profile.get("display_name")
            row["claim_title"] = claim.get("title")
            row["claim_is_hidden"] = claim.get("is_hidden")
    except Exception as error:
        print("[admin appeals] label lookup failed:", str(error), flush=True)

    return {"ok": True, "appeals": appeals}


@app.post("/admin/appeals/{appeal_id}/resolve")
def admin_resolve_appeal(appeal_id: str, payload: AppealResolveRequest, request: Request):
    admin_user_id = require_admin_user(request)
    decision = payload.decision.strip().lower()
    review_note = payload.review_note.strip()

    if decision not in {"granted", "denied"}:
        raise HTTPException(status_code=400, detail="decision must be 'granted' or 'denied'.")

    supabase = get_supabase_client()
    try:
        appeal_result = (
            supabase.table("moderation_appeals").select("*").eq("id", appeal_id.strip()).limit(1).execute()
        )
    except Exception as error:
        raise_appeals_unavailable(error)
    appeal = (appeal_result.data or [None])[0]

    if not appeal:
        raise HTTPException(status_code=404, detail="Appeal not found.")

    if appeal.get("status") != "pending":
        raise HTTPException(status_code=409, detail="This appeal was already resolved.")

    action_type = appeal.get("action_type")
    reversal_note = ""

    if decision == "granted":
        # Reverse the action by calling the EXISTING endpoint logic.
        if action_type == "claim_hidden" and appeal.get("claim_id"):
            admin_unhide_claim(str(appeal["claim_id"]), request)
        elif action_type == "account_suspended":
            admin_suspend_user(
                AdminUserActionRequest(user_id=str(appeal["user_id"]), suspended=False, reason=""),
                request,
            )
        elif action_type == "claim_removed":
            # Hard-deleted — cannot be restored; the grant is recorded only.
            reversal_note = " The original claim was permanently removed and cannot be restored."

    now_iso = datetime.now(timezone.utc).isoformat()
    try:
        update_result = (
            supabase.table("moderation_appeals")
            .update({
                "status": decision,
                "reviewed_by": admin_user_id,
                "review_note": review_note or None,
                "reviewed_at": now_iso,
            })
            .eq("id", appeal_id.strip())
            .eq("status", "pending")
            .execute()
        )
    except Exception as error:
        raise_appeals_unavailable(error)

    if not (update_result.data or []):
        raise HTTPException(status_code=409, detail="This appeal was already resolved.")

    # Notify the user (fail-soft, same pattern as the moderation actions).
    try:
        outcome_line = (
            "Your appeal was granted." if decision == "granted" else "Your appeal was denied."
        )

        if decision == "granted" and action_type == "claim_hidden":
            outcome_line += " Your claim has been restored."
        elif decision == "granted" and action_type == "account_suspended":
            outcome_line += " Your account has been unsuspended."
        elif decision == "granted" and action_type == "claim_removed":
            outcome_line += reversal_note

        body = outcome_line + (f" Reviewer note: {review_note}" if review_note else "")
        notification_claim_id = appeal.get("claim_id") if action_type != "claim_removed" else None

        supabase.table("notifications").insert({
            "user_id": appeal["user_id"],
            "type": "appeal_resolved",
            "title": "Your appeal was reviewed",
            "body": body,
            "claim_id": notification_claim_id,
        }).execute()
    except Exception as error:
        print(f"[admin appeals] notification failed: {error}", flush=True)

    return {"ok": True, "appeal_id": appeal_id, "status": decision}


@app.get("/admin/manage/users")
def admin_manage_users(request: Request, search: str = "", filter: str = "all"):
    require_admin_user(request)
    filter_mode = (filter or "all").strip().lower()

    if filter_mode not in {"all", "blocked", "suspended"}:
        filter_mode = "all"

    # Search runs inside admin_manage_users_search (SECURITY DEFINER,
    # service_role only) because emails live in auth.users, which PostgREST
    # does not expose. ilike wildcards typed by the admin are allowed on
    # purpose — this is an admin-only search box.
    search_query = (search or "").strip()[:80]
    supabase = get_supabase_client()
    result = supabase.rpc(
        "admin_manage_users_search",
        {"search_query": search_query, "filter_mode": filter_mode},
    ).execute()

    return {"ok": True, "users": result.data or []}


@app.get("/admin/manage/claims")
def admin_manage_claims(request: Request, search: str = "", filter: str = "all"):
    require_admin_user(request)
    filter_mode = (filter or "all").strip().lower()

    if filter_mode not in {"all", "hidden", "visible"}:
        filter_mode = "all"

    search_query = (search or "").strip()[:120]
    supabase = get_supabase_client()

    # Service-role read: bypasses the restrictive RLS policy, so hidden
    # claims are always visible to the dashboard.
    query = (
        supabase.table("claims")
        .select("id,title,author_id,is_hidden,is_deleted,deleted_reason,hidden,hidden_reason,hidden_at,votes_true,votes_fake,votes_unsure,created_at")
        .order("created_at", desc=True)
        .limit(50)
    )

    if search_query:
        query = query.ilike("title", f"%{search_query}%")

    if filter_mode == "hidden":
        query = query.eq("is_hidden", True)
    elif filter_mode == "visible":
        query = query.eq("is_hidden", False)

    result = query.execute()
    claims = result.data or []
    author_ids = sorted({claim["author_id"] for claim in claims if claim.get("author_id")})
    usernames: dict[str, str] = {}
    # TASK 4c — author suspension state so each claim row can show Ban/Unban.
    suspended_by_id: dict[str, bool] = {}

    if author_ids:
        profiles_result = (
            supabase.table("profiles")
            .select("id,username,is_suspended")
            .in_("id", author_ids)
            .execute()
        )
        for profile in (profiles_result.data or []):
            usernames[profile["id"]] = profile.get("username") or ""
            suspended_by_id[profile["id"]] = bool(profile.get("is_suspended"))

    for claim in claims:
        author_id = claim.get("author_id") or ""
        claim["author_username"] = usernames.get(author_id, None)
        claim["author_is_suspended"] = suspended_by_id.get(author_id, False)

    return {"ok": True, "claims": claims}


# PHASE 5 STEP 1C
@app.post("/admin/reputation/reset-monthly")
def admin_reset_monthly_reputation(request: Request):
    require_admin_user(request)
    supabase = get_supabase_client()
    now_iso = datetime.now(timezone.utc).isoformat()
    result = (
        supabase.table("profiles")
        .update({
            "monthly_reputation_points": 0,
            "monthly_reset_at": now_iso,
        })
        .execute()
    )

    print("[admin monthly reset] rows:", len(result.data or []), flush=True)

    return {
        "ok": True,
        "reset_at": now_iso,
        "rows": len(result.data or []),
    }


# PHASE 5 STEP 1B
@app.post("/admin/claims/override")
def admin_override_claim(payload: AdminOverrideRequest, request: Request):
    require_admin_user(request)
    claim_id = payload.claim_id.strip()
    new_status = payload.new_status.strip().upper()

    if not claim_id:
        raise HTTPException(status_code=400, detail="claim_id is required")

    if new_status not in {"FINALIZED_TRUE", "FINALIZED_FAKE", "INSUFFICIENT_DATA", "NEEDS_MORE_EVIDENCE"}:
        raise HTTPException(status_code=400, detail="Unsupported override status.")

    supabase = get_supabase_client()
    print("[admin override] claim_id:", claim_id, flush=True)
    print("[admin override] new_status:", new_status, flush=True)

    reverse_result = supabase.rpc("reverse_claim_reputation", {"target_claim_id": claim_id}).execute()
    print("[admin override] reverse rows:", reverse_result.data, flush=True)

    update_result = (
        supabase.table("claims")
        .update({
            "status": new_status,
            "verdict_reason": payload.reason.strip() or "Admin override applied.",
            "verdict_calculated_at": datetime.now(timezone.utc).isoformat(),
            "published_at": datetime.now(timezone.utc).isoformat(),
            "phase4_locked": True,
            "updated_at": datetime.now(timezone.utc).isoformat(),
        })
        .eq("id", claim_id)
        .execute()
    )

    print("[admin override] updated rows:", len(update_result.data or []), flush=True)

    process_result = supabase.rpc("process_claim_reputation", {"target_claim_id": claim_id}).execute()
    print("[admin override] process result:", process_result.data, flush=True)

    return {
        "ok": True,
        "claim_id": claim_id,
        "status": new_status,
    }


# VERDICT FORMULA v1 (NEW) — backend finalization point.
#
# Inspection finding: finalization in this app is client-triggered
# (claimService.ts -> finalize_expired_claim RPC); there was no backend
# finalization code path to insert into. This endpoint IS that path now:
# same guards as the RPC, verdict from services/verdict_engine.py
# (canonical formula; migration 043 mirrors it inside the RPC so the
# existing client flow uses v1 too). Already-finalized claims are never
# recomputed. Topic cluster stats are untouched.
@app.post("/claims/{claim_id}/finalize")
def finalize_claim_verdict_v1(claim_id: str, request: Request):
    get_authenticated_user_id(request)
    supabase = get_supabase_client()

    claim_result = (
        supabase.table("claims")
        .select("id, status, score_lock_at, expires_at")
        .eq("id", claim_id)
        .execute()
    )
    claim = (claim_result.data or [None])[0]

    if not claim:
        raise HTTPException(status_code=404, detail="Claim not found.")

    if claim.get("status") not in {"OPEN", "ACTIVE", "EARLY_VERDICT", "LOCKED", "VOTING_CLOSED"}:
        # Terminal claims keep their stored verdict — v1 applies to NEW
        # finalizations only.
        return {"ok": True, "claim_id": claim_id, "finalized": False, "status": claim.get("status")}

    lock_at = claim.get("score_lock_at") or claim.get("expires_at")

    if lock_at and datetime.fromisoformat(str(lock_at).replace("Z", "+00:00")) > datetime.now(timezone.utc):
        return {"ok": True, "claim_id": claim_id, "finalized": False, "status": claim.get("status")}

    verdict = compute_verdict(claim_id, supabase)
    final_status = map_verdict_to_claim_status(verdict["verdict"])
    now_iso = datetime.now(timezone.utc).isoformat()

    update_result = (
        supabase.table("claims")
        .update({
            "status": final_status,
            "verdict_reason": verdict["reason"],
            "combined_score": verdict["combined_score"],
            "decisive_ratio": verdict["decisive_ratio"],
            "evidence_ratio": verdict["evidence_ratio_true"],
            "total_votes": verdict["total_votes"],
            "verdict_calculated_at": now_iso,
            "published_at": now_iso,
            "phase4_locked": True,
            "updated_at": now_iso,
        })
        .eq("id", claim_id)
        .in_("status", ["OPEN", "ACTIVE", "EARLY_VERDICT", "LOCKED", "VOTING_CLOSED"])
        .execute()
    )

    print("[verdict v1] claim:", claim_id, "->", final_status, verdict["reason"], flush=True)

    return {
        "ok": True,
        "claim_id": claim_id,
        "finalized": bool(update_result.data),
        "status": final_status,
        **verdict,
    }


# PHASE 4 STEP 8
@app.get("/ai/library")
def ai_library():
    library = load_verifact_ai_library()
    sections = [
        "verifact_rules",
        "claim_type_rules",
        "source_quality_rules",
        "red_flag_rules",
        "confidence_rules",
        "source_credibility",
    ]

    return {
        "ok": True,
        "rules_loaded": all(section in library for section in sections),
        "sections": sections,
    }


def analyze_claim(payload: AiPrecheckRequest) -> dict:
    # PHASE 4 STEP 9
    source_metadata = score_source_url(payload.source_url)
    # PHASE 4 STEP 21
    source_page = fetch_source_page(payload.source_url)
    return analyze_claim_with_openai(
        title=payload.title,
        description=payload.description,
        source_url=payload.source_url,
        category=payload.category,
        source_metadata=source_metadata,
        source_page=source_page,
    )


# PHASE 4 STEP 9
def log_source_score(endpoint_label: str, source_metadata: dict) -> None:
    print("[source] domain:", source_metadata.get("domain"), flush=True)
    print("[source] quality:", source_metadata.get("source_quality"), flush=True)
    print("[source] score:", source_metadata.get("source_score"), flush=True)


# PHASE 4 STEP 21
def log_source_page(endpoint_label: str, source_page: dict) -> None:
    print(f"[{endpoint_label}] source_read_status:", source_page.get("status"), flush=True)
    print(f"[{endpoint_label}] source_page_title:", source_page.get("title"), flush=True)
    if source_page.get("error"):
        print(f"[{endpoint_label}] source_page_error:", source_page.get("error"), flush=True)
    print(f"[{endpoint_label}] source_excerpt_chars:", len(str(source_page.get("excerpt") or "")), flush=True)


# PHASE 4 STEP 10
def fetch_evidence_rows(claim_id: str) -> list[dict]:
    print("[ai evidence] claim_id:", claim_id, flush=True)
    supabase = get_supabase_client()
    response = (
        supabase.table("evidence")
        # PHASE 6 STEP 2 — also pull citation fields so offline sources reach the AI.
        .select("id, url, note, evidence_type, source_quality_label, source_quality_score, created_at, reference_type, citation, citation_verified")
        .eq("claim_id", claim_id)
        .order("created_at", desc=True)
        .limit(10)
        .execute()
    )
    evidence_rows = response.data or []

    if isinstance(evidence_rows, dict):
        evidence_rows = [evidence_rows]

    print("[ai evidence] evidence_count:", len(evidence_rows), flush=True)
    return evidence_rows


def get_supabase_create_client():
    repo_root = str(Path(__file__).resolve().parents[1])
    removed_paths: list[str] = []

    for candidate_path in ("", repo_root):
        if candidate_path in sys.path:
            sys.path.remove(candidate_path)
            removed_paths.append(candidate_path)

    try:
        from supabase import create_client
    except ImportError as error:
        raise RuntimeError("Supabase Python client is not installed on backend.") from error
    finally:
        for candidate_path in reversed(removed_paths):
            sys.path.insert(0, candidate_path)

    return create_client


def get_supabase_client() -> Any:
    supabase_url = os.environ.get("SUPABASE_URL", "").rstrip("/")
    service_role_key = os.environ.get("SUPABASE_SERVICE_ROLE_KEY", "")

    if not supabase_url or not service_role_key:
        raise RuntimeError("Missing SUPABASE_URL or SUPABASE_SERVICE_ROLE_KEY on backend.")

    return get_supabase_create_client()(supabase_url, service_role_key)


# PHASE 4 STEP 5B
def get_supabase_project_ref() -> str:
    supabase_url = os.environ.get("SUPABASE_URL", "").rstrip("/")
    hostname = urlparse(supabase_url).hostname or ""

    if hostname.endswith(".supabase.co"):
        return hostname.split(".")[0]

    return hostname or "unknown"


# PHASE 4 STEP 27
def get_authenticated_user_id(request: Request) -> str:
    authorization = request.headers.get("authorization", "")

    if not authorization.lower().startswith("bearer "):
        raise HTTPException(status_code=401, detail="Authentication required.")

    access_token = authorization.split(" ", 1)[1].strip()

    if not access_token:
        raise HTTPException(status_code=401, detail="Authentication required.")

    try:
        auth_response = get_supabase_client().auth.get_user(access_token)
        auth_user = getattr(auth_response, "user", None)
        user_id = getattr(auth_user, "id", None)
    except Exception:
        raise HTTPException(status_code=401, detail="Authentication required.")

    if not user_id:
        raise HTTPException(status_code=401, detail="Authentication required.")

    return str(user_id)


def normalize_claim_write_url(value: str | None) -> str:
    """Match the mobile normalizeUrl() behavior for claim write payloads."""
    normalized = str(value or "").strip()

    if not normalized:
        return ""

    if normalized.startswith("//"):
        normalized = f"https:{normalized}"
    elif not re.match(r"^https?://", normalized, flags=re.IGNORECASE):
        normalized = f"https://{normalized}"

    parsed = urlparse(normalized)
    hostname = str(parsed.hostname or "").lower()
    valid_hostname = hostname == "localhost" or bool(
        re.fullmatch(r"(?:[a-z0-9-]+\.)+[a-z]{2,}", hostname, flags=re.IGNORECASE)
    )

    if parsed.scheme.lower() not in {"http", "https"} or not valid_hostname:
        return ""

    return normalized


def normalize_optional_claim_write_url(value: str | None, field_label: str) -> str | None:
    if value is None or not str(value).strip():
        return None

    normalized = normalize_claim_write_url(value)
    if not normalized:
        raise HTTPException(status_code=422, detail=f"Enter a valid {field_label} URL.")

    return normalized


def validate_claim_create_payload(payload: ClaimCreateRequest) -> dict:
    title = payload.title.strip()
    description = payload.description.strip()
    category = payload.category.strip()
    source_url = normalize_claim_write_url(payload.source_url)

    if not title:
        raise HTTPException(status_code=422, detail="Title is required.")
    if len(title) > CLAIM_TITLE_MAX_LENGTH:
        raise HTTPException(status_code=422, detail="Title must be 160 characters or fewer.")
    if not description:
        raise HTTPException(status_code=422, detail="Description is required.")
    if len(description) > CLAIM_DESCRIPTION_MAX_LENGTH:
        raise HTTPException(
            status_code=422,
            detail=f"Description must be {CLAIM_DESCRIPTION_MAX_LENGTH} characters or fewer.",
        )
    if not category:
        raise HTTPException(status_code=422, detail="Category is required.")
    if not str(payload.source_url or "").strip():
        raise HTTPException(status_code=422, detail="Source URL is required.")
    if not source_url:
        raise HTTPException(status_code=422, detail="Enter a valid source URL.")

    return {
        "title": title,
        "description": description,
        "category": category,
        "source_url": source_url,
        "video_url": normalize_optional_claim_write_url(payload.video_url, "video"),
        "image_url": normalize_optional_claim_write_url(payload.image_url, "image"),
        "image_path": str(payload.image_path or "").strip() or None,
        "thumbnail_url": normalize_optional_claim_write_url(payload.thumbnail_url, "thumbnail"),
        "sub_category": str(payload.sub_category or "").strip() or None,
        "politician_tag": str(payload.politician_tag or "").strip() or None,
    }


def get_active_write_profile(supabase: Any, user_id: str, action: str) -> dict:
    result = (
        supabase.table("profiles")
        .select("id,is_suspended,is_deleted,suspension_reason")
        .eq("id", user_id)
        .limit(1)
        .execute()
    )
    profile = get_first_row(result)

    if not profile or profile.get("is_deleted"):
        raise HTTPException(status_code=403, detail="An active profile is required.")

    if profile.get("is_suspended"):
        fallback = f"This account is suspended from {action}."
        raise HTTPException(status_code=403, detail=str(profile.get("suspension_reason") or fallback))

    return profile


def enforce_claims_per_day_limit(supabase: Any, user_id: str) -> None:
    since = (datetime.now(timezone.utc) - timedelta(hours=24)).isoformat()

    try:
        response = (
            supabase.table("claims")
            .select("id", count="exact")
            .eq("author_id", user_id)
            .gte("created_at", since)
            .execute()
        )
    except Exception as error:
        print("[claims/create] daily limit query failed:", str(error), flush=True)
        raise HTTPException(status_code=503, detail="Could not verify the posting limit. Please retry.")

    count = getattr(response, "count", None)
    if count is None:
        rows = getattr(response, "data", None) or []
        count = len(rows) if isinstance(rows, list) else 0

    if int(count or 0) >= CLAIMS_PER_DAY_LIMIT:
        raise HTTPException(status_code=429, detail="You reached today's claim limit. Please try again later.")


def generate_claim_write_slug(title: str) -> str:
    slug = re.sub(r"[^a-z0-9]+", "-", title.strip().lower()).strip("-")[:72].rstrip("-")
    return slug or "claim"


def build_claim_insert_payload(author_id: str, fields: dict, embedding: Any = None, topic_cluster_id: str | None = None) -> dict:
    created_at = datetime.now(timezone.utc)
    vote_window_end = created_at + timedelta(hours=20)
    score_lock_at = created_at + timedelta(hours=24)
    created_at_iso = created_at.isoformat()
    score_lock_at_iso = score_lock_at.isoformat()

    insert_payload = {
        "author_id": author_id,
        "created_at": created_at_iso,
        "title": fields["title"],
        "description": fields["description"],
        "source_url": fields["source_url"],
        "video_url": fields["video_url"],
        "category": fields["category"],
        "sub_category": fields["sub_category"] if fields["category"] == "Politics" else None,
        "politician_tag": (
            fields["politician_tag"]
            if fields["category"] == "Politics" and fields["sub_category"] == "Politician"
            else None
        ),
        "slug": generate_claim_write_slug(fields["title"]),
        "votes_true": 0,
        "votes_fake": 0,
        "votes_unsure": 0,
        "total_votes": 0,
        "verdict_reason": None,
        "verdict_calculated_at": None,
        "status": "ACTIVE",
        "ai_status": "PENDING",
        "claim_type": "UNCLEAR",
        "ai_confidence": None,
        "ai_reason": None,
        "report_count": 0,
        "evidence_count": 0,
        "evidence_used_count": 0,
        "is_flagged": False,
        "mode": "production",
        "current_phase": 1,
        "vote_window_minutes": 1200,
        "vote_window_end": vote_window_end.isoformat(),
        "vote_accept_until": vote_window_end.isoformat(),
        "score_lock_at": score_lock_at_iso,
        "published_at": None,
        "phase4_locked": False,
        "early_verdict_fired": False,
        "suspicious_activity": False,
        "weighted_community_score": 0,
        "final_score": 0,
        "min_votes_required": 15,
        "expected_participation": 30,
        "source_count": 0,
        "source_quality": "unknown",
        "source_domain": None,
        "source_score": None,
        "source_reason": None,
        "red_flags": [],
        "ai_summary": None,
        "expires_at": score_lock_at_iso,
        "safety_status": "APPROVED",
        "safety_category": None,
        "safety_checked_at": created_at_iso,
    }

    for key in ("image_url", "image_path", "thumbnail_url"):
        if fields.get(key):
            insert_payload[key] = fields[key]

    if embedding is not None:
        insert_payload["embedding"] = embedding
    if topic_cluster_id:
        insert_payload["topic_cluster_id"] = topic_cluster_id

    return insert_payload


# PHASE 5 STEP 1B
def require_admin_user(request: Request) -> str:
    authenticated_user_id = get_authenticated_user_id(request)
    supabase = get_supabase_client()
    result = (
        supabase.table("profiles")
        .select("id,is_admin,is_suspended,is_deleted")
        .eq("id", authenticated_user_id)
        .limit(1)
        .execute()
    )
    profile = (result.data or [None])[0]

    if not profile or not profile.get("is_admin"):
        raise HTTPException(status_code=403, detail="Admin access required.")

    if profile.get("is_suspended") or profile.get("is_deleted"):
        raise HTTPException(status_code=403, detail="Admin account is not active.")

    return authenticated_user_id


# PHASE 4 STEP 5B
# PHASE 4 STEP 5C
def normalize_red_flags(red_flags: object) -> list[str]:
    if red_flags is None:
        return []

    if isinstance(red_flags, str):
        return [red_flags]

    if isinstance(red_flags, list):
        return [str(flag) for flag in red_flags]

    return [str(red_flags)]


# PHASE 4 STEP 5B
# PHASE 4 STEP 5C
# PHASE 4 STEP 5D
# PHASE 4 STEP 16
# PHASE 4 STEP 17
# PHASE 4 STEP 20
# PHASE 4 STEP 20B
def build_claim_ai_update_payload(analysis: dict) -> dict:
    claim_type = str(analysis.get("claim_type") or "UNCLEAR").upper()
    if claim_type not in {"FACTUAL", "OPINION", "SATIRE", "QUESTION", "PROMOTION", "UNCLEAR"}:
        claim_type = "UNCLEAR"

    source_quality = normalize_source_quality(analysis.get("source_quality"))
    ai_status = analysis["ai_status"]
    ai_confidence = analysis["ai_confidence"]
    source_count = analysis["source_count"]

    if claim_type in {"OPINION", "QUESTION", "SATIRE", "PROMOTION"}:
        ai_status = "NOT_FACT_CHECKABLE"
        ai_confidence = 0.5
        source_quality = normalize_source_quality(source_quality)
        source_count = 0

    return {
        # PHASE 4 STEP 7
        "claim_type": claim_type,
        "ai_status": ai_status,
        "ai_confidence": ai_confidence,
        "source_quality": source_quality,
        # PHASE 4 STEP 9
        "source_domain": analysis.get("source_domain"),
        "source_score": analysis.get("source_score"),
        "source_reason": analysis.get("source_reason"),
        # PHASE 4 STEP 21
        "source_read_status": analysis.get("source_read_status") or "not_read",
        "source_page_title": analysis.get("source_page_title") or None,
        "source_excerpt": analysis.get("source_excerpt") or None,
        "source_supports_claim": analysis.get("source_supports_claim"),
        "source_support_summary": analysis.get("source_support_summary") or None,
        # PHASE 4 STEP 10
        "evidence_used_count": analysis.get("evidence_used_count", 0),
        "source_count": source_count,
        "red_flags": normalize_red_flags(analysis.get("red_flags")),
        "ai_summary": analysis["ai_summary"],
        # PHASE 6 STEP 3 — natural-truth classification, stored alongside
        # ai_status on claims (nullable; fallback/error paths yield None).
        # Requires migration 036 before deploy.
        "naturally_true_category": analysis.get("naturally_true_category"),
        "verdict_signal": analysis.get("verdict_signal"),
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }


# PHASE 4 STEP 5B
def format_supabase_error(error: Exception) -> dict:
    error_dict = getattr(error, "dict", None)

    if callable(error_dict):
        try:
            data = error_dict()
            return {
                "error": str(data.get("message") or error),
                "details": None if data.get("details") is None else str(data.get("details")),
                "hint": None if data.get("hint") is None else str(data.get("hint")),
            }
        except Exception:
            pass

    return {
        "error": str(getattr(error, "message", None) or error),
        "details": None if getattr(error, "details", None) is None else str(getattr(error, "details")),
        "hint": None if getattr(error, "hint", None) is None else str(getattr(error, "hint")),
    }


# PHASE 4 STEP 5C
def format_supabase_response_error(error: object) -> dict:
    if isinstance(error, dict):
        return {
            "error": str(error.get("message") or error),
            "details": None if error.get("details") is None else str(error.get("details")),
            "hint": None if error.get("hint") is None else str(error.get("hint")),
        }

    return {
        "error": str(getattr(error, "message", None) or error),
        "details": None if getattr(error, "details", None) is None else str(getattr(error, "details")),
        "hint": None if getattr(error, "hint", None) is None else str(getattr(error, "hint")),
    }


# PHASE 4 STEP 5B
# PHASE 4 STEP 5C
# PHASE 4 STEP 5D
# PHASE 4 STEP 16
# PHASE 4 STEP 20
# PHASE 4 STEP 20B
# PHASE 4 STEP 20C
def update_claim_ai_fields(claim_id: str, ai_result: dict, endpoint_label: str) -> dict:
    update_payload = build_claim_ai_update_payload(ai_result)
    # PHASE 4 STEP 17
    # PHASE 4 STEP 20
    # PHASE 4 STEP 20B
    # PHASE 4 STEP 20C
    update_payload["source_quality"] = normalize_source_quality(update_payload.get("source_quality"))
    if update_payload["source_quality"] not in ALLOWED_SOURCE_QUALITIES:
        print("[ai] invalid source_quality blocked:", update_payload["source_quality"], flush=True)
        update_payload["source_quality"] = "unknown"

    print(f"[{endpoint_label}] Supabase project_ref:", get_supabase_project_ref(), flush=True)
    print(f"[{endpoint_label}] claim_id:", claim_id, flush=True)
    print("[AI UPDATE FINAL SOURCE QUALITY]", update_payload["source_quality"], flush=True)
    print("[ai] normalized source_quality:", update_payload["source_quality"], flush=True)
    print("[ai] claim_type:", update_payload.get("claim_type"), flush=True)
    # PHASE 4 STEP 27
    print(
        f"[{endpoint_label}] AI result summary:",
        {
            "ai_status": ai_result.get("ai_status"),
            "ai_confidence": ai_result.get("ai_confidence"),
            "source_read_status": ai_result.get("source_read_status"),
            "source_supports_claim": ai_result.get("source_supports_claim"),
        },
        flush=True,
    )

    try:
        supabase = get_supabase_client()
        update_result = (
            supabase.table("claims")
            .update(update_payload)
            .eq("id", claim_id)
            .execute()
        )
        print(f"[{endpoint_label}] update_result rows:", len(update_result.data or []), flush=True)

        update_error = getattr(update_result, "error", None)
        if update_error:
            formatted_error = format_supabase_response_error(update_error)
            print(f"[{endpoint_label}] update_error:", formatted_error, flush=True)
            return {
                "ok": False,
                "update_payload": update_payload,
                **formatted_error,
            }

        fetch_result = (
            supabase.table("claims")
            .select("id, claim_type, ai_status, ai_confidence, source_quality, source_domain, source_score, source_reason, source_read_status, source_page_title, source_supports_claim, source_support_summary, evidence_used_count, red_flags, ai_summary, source_count, naturally_true_category, verdict_signal, updated_at")
            .eq("id", claim_id)
            .execute()
        )
    except Exception as error:
        formatted_error = format_supabase_error(error)
        print(f"[{endpoint_label}] update_error:", formatted_error, flush=True)
        return {
            "ok": False,
            "update_payload": update_payload,
            **formatted_error,
        }

    fetch_error = getattr(fetch_result, "error", None)
    if fetch_error:
        formatted_error = format_supabase_response_error(fetch_error)
        print(f"[{endpoint_label}] fetch_error:", formatted_error, flush=True)
        return {
            "ok": False,
            "update_payload": update_payload,
            **formatted_error,
        }

    updated_claim = fetch_result.data
    if isinstance(updated_claim, list):
        updated_claim = updated_claim[0] if updated_claim else None

    print(
        f"[{endpoint_label}] fetched updated claim summary:",
        {
            "id": updated_claim.get("id") if updated_claim else None,
            "ai_status": updated_claim.get("ai_status") if updated_claim else None,
            "source_read_status": updated_claim.get("source_read_status") if updated_claim else None,
        },
        flush=True,
    )

    if not updated_claim:
        return {
            "ok": False,
            "claim_id": claim_id,
            "error": "Supabase update ran but updated claim could not be fetched",
            "update_payload": update_payload,
        }

    if updated_claim.get("ai_summary") != update_payload["ai_summary"]:
        return {
            "ok": False,
            "claim_id": claim_id,
            "error": "Supabase update did not persist new AI fields",
            "ai_result": ai_result,
            "update_payload": update_payload,
            "updated_claim": updated_claim,
        }

    return {
        "ok": True,
        "ai_result": ai_result,
        "update_payload": update_payload,
        "updated_claim": updated_claim,
    }


# PHASE 4 STEP 3
# PHASE 4 STEP 17
def build_ai_error_analysis() -> dict:
    return {
        # PHASE 4 STEP 7
        "claim_type": "UNCLEAR",
        "ai_confidence": 0.5,
        "source_count": 0,
        "source_quality": "unknown",
        # PHASE 4 STEP 21
        "source_read_status": "not_read",
        "source_page_title": None,
        "source_excerpt": None,
        "source_supports_claim": None,
        "source_support_summary": "Source page was not read because AI pre-check failed.",
        # PHASE 4 STEP 10
        "evidence_used_count": 0,
        "red_flags": ["AI pre-check failed"],
        "ai_summary": "AI pre-check failed. Community voting and evidence are still available.",
        "ai_status": "ERROR",
    }


# PHASE 4 STEP 3
def mark_claim_ai_error(claim_id: str) -> str | None:
    error_analysis = build_ai_error_analysis()
    result = update_claim_ai_fields(claim_id, error_analysis, "ai/precheck/error")

    if not result.get("ok"):
        return str(result.get("error") or "Supabase error status update failed.")

    return None


def build_safe_ai_precheck_failure(claim_id: str) -> dict:
    return {
        "ok": False,
        "claim_id": claim_id,
        "error": "AI pre-check is unavailable right now.",
        "details": None,
        "hint": None,
    }


# PHASE 4 STEP 5B
# PHASE 4 STEP 5C
# PHASE 4 STEP 5D
# PHASE 4 STEP 16
def build_ai_precheck_response(claim_id: str, update_result: dict) -> dict:
    updated_claim = update_result["updated_claim"]
    red_flags = normalize_red_flags(updated_claim.get("red_flags"))
    # PHASE 4 STEP 27
    safe_updated_claim = {
        "id": updated_claim.get("id"),
        "claim_type": updated_claim.get("claim_type"),
        "ai_status": updated_claim.get("ai_status"),
        "ai_confidence": updated_claim.get("ai_confidence"),
        "source_quality": updated_claim.get("source_quality"),
        "source_domain": updated_claim.get("source_domain"),
        "source_score": updated_claim.get("source_score"),
        "source_reason": updated_claim.get("source_reason"),
        "source_read_status": updated_claim.get("source_read_status"),
        "source_page_title": updated_claim.get("source_page_title"),
        "source_supports_claim": updated_claim.get("source_supports_claim"),
        "source_support_summary": updated_claim.get("source_support_summary"),
        "evidence_used_count": updated_claim.get("evidence_used_count"),
        "red_flags": red_flags,
        "ai_summary": updated_claim.get("ai_summary"),
        "source_count": updated_claim.get("source_count"),
        # PHASE 6 STEP 3
        "naturally_true_category": updated_claim.get("naturally_true_category"),
        "verdict_signal": updated_claim.get("verdict_signal"),
        "updated_at": updated_claim.get("updated_at"),
    }

    return {
        "ok": True,
        "claim_id": claim_id,
        # PHASE 4 STEP 7
        "claim_type": updated_claim.get("claim_type"),
        "ai_confidence": updated_claim.get("ai_confidence"),
        "source_count": updated_claim.get("source_count"),
        "source_quality": updated_claim.get("source_quality"),
        # PHASE 4 STEP 9
        "source_domain": updated_claim.get("source_domain"),
        "source_score": updated_claim.get("source_score"),
        "source_reason": updated_claim.get("source_reason"),
        # PHASE 4 STEP 21
        "source_read_status": updated_claim.get("source_read_status"),
        "source_page_title": updated_claim.get("source_page_title"),
        "source_supports_claim": updated_claim.get("source_supports_claim"),
        "source_support_summary": updated_claim.get("source_support_summary"),
        # PHASE 4 STEP 10
        "evidence_used_count": updated_claim.get("evidence_used_count"),
        "red_flags": red_flags,
        "ai_summary": updated_claim.get("ai_summary"),
        "ai_status": updated_claim.get("ai_status"),
        # PHASE 6 STEP 3
        "naturally_true_category": updated_claim.get("naturally_true_category"),
        "verdict_signal": updated_claim.get("verdict_signal"),
        "error": None,
        "details": None,
        "hint": None,
        "supabase_updated": True,
        "updated_claim": safe_updated_claim,
    }


# PHASE 4 STEP 3
def fetch_claim_row(claim_id: str) -> dict | None:
    supabase = get_supabase_client()
    response = supabase.table("claims").select("*").eq("id", claim_id).limit(1).execute()
    rows = response.data or []

    if isinstance(rows, dict):
        return rows

    if not rows:
        return None

    return rows[0]


# PHASE 4 STEP 3
def build_precheck_payload_from_claim(claim: dict) -> AiPrecheckRequest:
    return AiPrecheckRequest(
        claim_id=str(claim.get("id") or ""),
        title=str(claim.get("title") or ""),
        description=str(claim.get("description") or ""),
        source_url=str(claim.get("source_url") or ""),
        category=str(claim.get("category") or ""),
    )


# PHASE 6 STEP 1 — Duplicate claim detection.
def store_claim_embedding(claim_id: str) -> bool:
    """Generate and persist the embedding for an already-saved claim.

    WHY separate from claim creation: claims are inserted client-side (Supabase),
    so the embedding — which needs the server-only OpenAI key and service-role
    write access — is filled in by a follow-up backend call. Returns True on a
    stored embedding, False on any soft failure (missing claim / embedding error),
    which the caller treats as non-fatal.
    """
    claim = fetch_claim_row(claim_id)

    if not claim:
        print("[claims/embed] claim not found:", claim_id, flush=True)
        return False

    if claim.get("embedding"):
        print("[claims/embed] existing embedding reused for claim:", claim_id, flush=True)
        return True

    embedding = generate_claim_embedding(
        title=str(claim.get("title") or ""),
        description=str(claim.get("description") or ""),
    )

    if embedding is None:
        # Embedding failure must never be treated as an error worth surfacing;
        # the backfill script will pick this claim up later.
        print("[claims/embed] embedding unavailable for claim:", claim_id, flush=True)
        return False

    supabase = get_supabase_client()
    supabase.table("claims").update({"embedding": embedding}).eq("id", claim_id).execute()
    print("[claims/embed] stored embedding for claim:", claim_id, flush=True)
    return True


@app.post("/api/claims/embed")
def api_claims_embed(payload: ClaimEmbedRequest, request: Request):
    """Fire-and-forget endpoint the client calls after creating a claim.

    Always returns 200 with {"ok": bool}: the frontend does not await the result,
    and an embedding failure must never look like a claim-creation failure.
    """
    claim_id = payload.claim_id.strip()

    if not claim_id:
        raise HTTPException(status_code=400, detail="claim_id is required")

    enforce_rate_limit(
        request,
        "claims_embed",
        DUPLICATE_CHECK_RATE_LIMIT_MAX_REQUESTS,
        AI_RATE_LIMIT_WINDOW_SECONDS,
    )
    get_authenticated_user_id(request)

    try:
        stored = store_claim_embedding(claim_id)
    except Exception as error:
        # Fail soft — never bubble a 500 into the post flow.
        print(f"[claims/embed] failure: {error}", flush=True)
        return {"ok": False}

    # TOPIC CLUSTER — additive, never blocks claim creation (PHASE 6 STEP 4).
    # INSPECTION NOTE: there is no backend claim-creation endpoint — claims are
    # inserted client-side (services/claimService.ts createClaim), and this
    # fire-and-forget endpoint is the client's follow-up call right after the
    # INSERT succeeds. That makes it the claim-creation hook. It also just
    # stored the claim's embedding, which find_or_create_topic_cluster reuses
    # instead of generating a second one. Response value is unchanged.
    try:
        cluster_claim = fetch_claim_row(claim_id)
        if cluster_claim and not cluster_claim.get("topic_cluster_id"):
            find_or_create_topic_cluster(
                claim_id=claim_id,
                title=str(cluster_claim.get("title") or ""),
                description=str(cluster_claim.get("description") or ""),
                category=str(cluster_claim.get("category") or ""),
            )
    except Exception as cluster_error:
        print(f"[topic_cluster] non-blocking error: {cluster_error}", flush=True)

    return {"ok": stored}


def find_duplicate_claims(title: str, description: str, embedding: Any = None) -> dict:
    """Suggest existing claims that semantically match the one being submitted.

    Runs BEFORE the claim is saved so the frontend can offer "vote on this
    instead". Fails OPEN: any failure returns an empty duplicate list with 200 so
    it can never block a user from posting.

    Two-stage matching keeps precision high without paying for an LLM call on
    every candidate:
      * 0.85 < sim < 0.95 -> ask gpt-4.1-mini if it's the SAME assertion, and drop
        NO / OPPOSITE (this is what stops "voted for" matching "voted against").
      * sim >= 0.95        -> trust the embedding, skip the stance check.
    """
    embedding = embedding or generate_claim_embedding(title=title, description=description)

    if embedding is None:
        # Fail open: no embedding -> no suggestions, but posting is never blocked.
        # PHASE 6 STEP 4: topic_cluster is an additive key (always present, may
        # be null) so clients can rely on its shape.
        return {"duplicates": [], "topic_cluster": None, "embedding": None}

    try:
        supabase = get_supabase_client()
        response = supabase.rpc(
            "match_claims",
            {
                "query_embedding": embedding,
                "match_threshold": DUPLICATE_SIMILARITY_THRESHOLD,
                "match_count": DUPLICATE_MATCH_LIMIT,
                "allowed_statuses": DUPLICATE_VOTABLE_STATUSES,
            },
        ).execute()
        candidates = response.data or []
    except Exception as error:
        print(f"[check-duplicate] vector search failed: {error}", flush=True)
        return {"duplicates": [], "topic_cluster": None, "embedding": embedding}

    duplicates = []

    for candidate in candidates:
        similarity = float(candidate.get("similarity") or 0.0)
        candidate_title = str(candidate.get("title") or "")

        # Ambiguous band: confirm the assertion actually matches before suggesting.
        if similarity < DUPLICATE_STANCE_SKIP_SIMILARITY:
            stance = classify_claim_stance(title, candidate_title)
            if stance in {"NO", "OPPOSITE"}:
                print(
                    f"[check-duplicate] excluded by stance ({stance}): "
                    f"{candidate.get('id')} sim={similarity:.3f}",
                    flush=True,
                )
                continue

        duplicates.append(
            {
                "claim_id": str(candidate.get("id") or ""),
                "title": candidate_title,
                "similarity": similarity,
                "vote_count": int(candidate.get("vote_count") or 0),
                "verdict_status": str(candidate.get("verdict_status") or ""),
            }
        )

    # PHASE 6 STEP 4 — TOPIC CLUSTER info (additive field, null if no cluster).
    # INSPECTION NOTE: the spec references /api/claims/check-title, which does
    # not exist in this codebase; this endpoint is its equivalent (the pre-save
    # title/description check with an embedding already in hand). Reuses that
    # embedding — no second OpenAI call.
    topic_cluster = None
    try:
        cluster_response = get_supabase_client().rpc(
            "match_claim_topics",
            {
                "query_embedding": embedding,
                "match_threshold": CLUSTER_MATCH_THRESHOLD,
                "match_count": 1,
            },
        ).execute()
        cluster_rows = cluster_response.data or []

        if cluster_rows:
            cluster_row = cluster_rows[0]
            topic_cluster = {
                "topic_cluster_id": str(cluster_row.get("id") or ""),
                "topic_label": str(cluster_row.get("topic_label") or ""),
                "claim_count": int(cluster_row.get("claim_count") or 0),
                "cluster_verdict": str(cluster_row.get("cluster_verdict") or "INSUFFICIENT_DATA"),
                "total_vote_count": int(cluster_row.get("total_vote_count") or 0),
            }
    except Exception as cluster_error:
        print(f"[check-duplicate] topic cluster lookup failed: {cluster_error}", flush=True)

    return {"duplicates": duplicates, "topic_cluster": topic_cluster, "embedding": embedding}


@app.post("/api/claims/check-duplicate")
def api_claims_check_duplicate(payload: DuplicateCheckRequest, request: Request):
    enforce_rate_limit(
        request,
        "check_duplicate",
        DUPLICATE_CHECK_RATE_LIMIT_MAX_REQUESTS,
        AI_RATE_LIMIT_WINDOW_SECONDS,
    )
    get_authenticated_user_id(request)
    result = find_duplicate_claims(payload.title.strip(), payload.description.strip())
    return {
        "duplicates": result["duplicates"],
        "topic_cluster": result["topic_cluster"],
    }


# CONTENT SAFETY (NEW, additive) — objectionable-content gate at submission.
# Separate from the truth/source AI analysis. Runs BEFORE the claim is saved
# (claims are inserted client-side via Supabase, so the frontend calls this
# first and only inserts when safe). FAIL-OPEN: check_content_safety never
# raises on API failure, so a down safety API can never block posting.
BLOCKED_CONTENT_MESSAGE = "This claim may contain violent or threatening language. Please rewrite it before posting."


def log_content_safety_block(user_id: str | None, title: str, category: str, reason: str) -> None:
    """Record a blocked attempt for moderation visibility. Never fails the request."""
    try:
        supabase = get_supabase_client()
        supabase.table("content_safety_blocks").insert({
            "user_id": user_id,
            "title_snippet": str(title or "")[:120],
            "category": str(category or "")[:40],
            "reason": str(reason or "")[:200],
        }).execute()
    except Exception as error:
        print(f"[content-safety] block log failed: {error}", flush=True)


# /api/content/check-safety is the documented endpoint name; /api/claims/safety-check
# is the original path the app already calls. Both map to the same handler (alias).
@app.get("/admin/content-safety/openai-status")
def admin_content_safety_openai_status(request: Request):
    require_admin_role(request, {"SUPER_ADMIN", "ADMIN", "MODERATOR"})
    return {
        "ok": True,
        "content_safety_openai": get_content_safety_openai_status(),
    }


@app.post("/api/content/check-safety")
@app.post("/api/claims/safety-check")
def api_claims_safety_check(payload: ContentSafetyRequest, request: Request):
    """Classify a claim's SAFETY before it is saved.

    Returns 200 {ok, safe:true} when clear; 400 {ok:false, blocked:true, ...}
    when objectionable. Politically neutral — partisan/controversial factual
    claims pass. Fails open: any classifier failure returns safe.
    """
    enforce_rate_limit(
        request,
        "content_safety",
        DUPLICATE_CHECK_RATE_LIMIT_MAX_REQUESTS,
        AI_RATE_LIMIT_WINDOW_SECONDS,
    )
    user_id = get_authenticated_user_id(request)

    verdict = check_content_safety(payload.title, payload.description)
    print(
        f"[content-safety] ENDPOINT RAN user={user_id} "
        f"title_length={len(str(payload.title or ''))} "
        f"description_length={len(str(payload.description or ''))} verdict={verdict}",
        flush=True,
    )

    if verdict.get("safe", True):
        return {
            "ok": True,
            "safe": True,
            "category": "",
            "severity": "",
            "reason": "",
            "matched_layer": "",
        }

    category = str(verdict.get("category") or "objectionable")
    severity = str(verdict.get("severity") or "")
    matched_layer = str(verdict.get("matched_layer") or "")
    log_content_safety_block(user_id, payload.title, category, str(verdict.get("reason") or ""))

    return JSONResponse(
        status_code=400,
        content={
            "ok": False,
            "code": "UNSAFE_CLAIM_TEXT",
            "safe": False,
            "blocked": True,
            "message": BLOCKED_CONTENT_MESSAGE,
            "reason": BLOCKED_CONTENT_MESSAGE,
            "category": category,
            "severity": severity,
            "matched_layer": matched_layer,
        },
    )


@app.post("/api/claims", status_code=201)
def api_create_claim(payload: ClaimCreateRequest, request: Request, background_tasks: BackgroundTasks):
    """Authoritative claim write path; legacy RLS inserts remain during rollout."""
    author_id = get_authenticated_user_id(request)
    enforce_rate_limit(
        request,
        "claims_create",
        CLAIMS_CREATE_RATE_LIMIT_MAX_REQUESTS,
        AI_RATE_LIMIT_WINDOW_SECONDS,
    )
    fields = validate_claim_create_payload(payload)
    supabase = get_supabase_client()
    get_active_write_profile(supabase, author_id, "posting")
    enforce_claims_per_day_limit(supabase, author_id)

    verdict = check_content_safety(fields["title"], fields["description"])
    if not verdict.get("safe", True):
        category = str(verdict.get("category") or "objectionable")
        reason = str(verdict.get("reason") or BLOCKED_CONTENT_MESSAGE)
        log_content_safety_block(author_id, fields["title"], category, reason)
        return JSONResponse(
            status_code=400,
            content={
                "ok": False,
                "code": "UNSAFE_CLAIM_TEXT",
                "blocked": True,
                "reason": BLOCKED_CONTENT_MESSAGE,
                "category": category,
            },
        )

    duplicate_result = find_duplicate_claims(fields["title"], fields["description"])
    embedding = duplicate_result.get("embedding")
    topic_cluster = duplicate_result.get("topic_cluster") or {}
    topic_cluster_id = str(topic_cluster.get("topic_cluster_id") or "") or None
    insert_payload = build_claim_insert_payload(
        author_id,
        fields,
        embedding=embedding,
        topic_cluster_id=topic_cluster_id,
    )

    try:
        insert_result = supabase.table("claims").insert(insert_payload).execute()
    except Exception as error:
        print("[claims/create] insert failed:", str(error), flush=True)
        raise HTTPException(status_code=500, detail="Could not create claim right now.")

    created_claim = get_first_row(insert_result)
    claim_id = str((created_claim or {}).get("id") or "")
    if not created_claim or not claim_id:
        print("[claims/create] insert returned no row", flush=True)
        raise HTTPException(status_code=500, detail="Could not create claim right now.")

    share_url = f"{PUBLIC_SITE_URL}/claim/{claim_id}"
    try:
        update_result = (
            supabase.table("claims")
            .update({"share_url": share_url})
            .eq("id", claim_id)
            .execute()
        )
        created_claim = get_first_row(update_result) or {**created_claim, "share_url": share_url}
    except Exception as error:
        print("[claims/create] share URL update failed:", str(error), flush=True)
        created_claim = {**created_claim, "share_url": share_url}

    background_tasks.add_task(run_claim_post_insert_tasks, claim_id, author_id)
    print(
        "[claims/create] created",
        {
            "claim_id": claim_id,
            "author_id": author_id,
            "duplicate_candidates": len(duplicate_result.get("duplicates") or []),
            "topic_cluster_id": topic_cluster_id,
        },
        flush=True,
    )
    return created_claim


def parse_claim_timestamp(value: Any) -> datetime | None:
    normalized = str(value or "").strip()
    if not normalized:
        return None

    try:
        parsed = datetime.fromisoformat(normalized.replace("Z", "+00:00"))
    except ValueError:
        return None

    return parsed if parsed.tzinfo else parsed.replace(tzinfo=timezone.utc)


@app.post("/api/claims/{claim_id}/vote", status_code=201)
def api_vote_on_claim(claim_id: str, payload: ClaimVoteRequest, request: Request):
    user_id = get_authenticated_user_id(request)
    enforce_rate_limit(
        request,
        "claims_vote",
        DUPLICATE_CHECK_RATE_LIMIT_MAX_REQUESTS,
        AI_RATE_LIMIT_WINDOW_SECONDS,
    )
    normalized_claim_id = claim_id.strip()
    if not is_uuid(normalized_claim_id):
        raise HTTPException(status_code=404, detail="Claim not found.")

    vote_type = payload.vote_type.strip().upper().replace("NOT_SURE", "UNSURE")
    if vote_type not in {"TRUE", "FAKE", "UNSURE"}:
        raise HTTPException(status_code=422, detail="vote_type must be TRUE, FAKE, or UNSURE.")

    supabase = get_supabase_client()
    get_active_write_profile(supabase, user_id, "voting")
    claim_result = (
        supabase.table("claims")
        .select("*")
        .eq("id", normalized_claim_id)
        .limit(1)
        .execute()
    )
    claim = get_first_row(claim_result)
    if not claim or claim.get("is_deleted"):
        raise HTTPException(status_code=404, detail="Claim not found.")

    read_only_statuses = {
        "FINALIZED_TRUE",
        "FINALIZED_FAKE",
        "INSUFFICIENT_DATA",
        "COMMUNITY_TRUE",
        "COMMUNITY_FAKE",
        "NEEDS_MORE_EVIDENCE",
        "VOTING_CLOSED",
        "LOCKED",
    }
    vote_deadline = parse_claim_timestamp(claim.get("vote_accept_until") or claim.get("vote_window_end"))
    if (
        claim.get("published_at")
        or claim.get("phase4_locked")
        or str(claim.get("status") or "").upper() in read_only_statuses
        or (vote_deadline is not None and vote_deadline <= datetime.now(timezone.utc))
    ):
        raise HTTPException(status_code=409, detail="Voting is closed.")

    existing_result = (
        supabase.table("votes")
        .select("*")
        .eq("claim_id", normalized_claim_id)
        .eq("user_id", user_id)
        .limit(1)
        .execute()
    )
    existing_vote = get_first_row(existing_result)
    if existing_vote:
        return JSONResponse(
            status_code=409,
            content={
                "ok": False,
                "already_voted": True,
                "message": "You already voted on this claim.",
                "vote": existing_vote,
            },
        )

    vote_value = 1.0 if vote_type == "TRUE" else 0.0 if vote_type == "FAKE" else 0.5
    vote_payload = {
        "claim_id": normalized_claim_id,
        "user_id": user_id,
        "vote_type": vote_type,
        "vote_value": vote_value,
        "accepted": True,
        "suspicious": False,
        "rejected_reason": None,
    }

    try:
        insert_result = supabase.table("votes").insert(vote_payload).execute()
    except Exception as error:
        message = str(error).lower()
        if "23505" in message or "duplicate" in message or "unique" in message:
            return JSONResponse(
                status_code=409,
                content={
                    "ok": False,
                    "already_voted": True,
                    "message": "You already voted on this claim.",
                },
            )
        print("[claims/vote] insert failed:", str(error), flush=True)
        raise HTTPException(status_code=500, detail="Could not save vote right now.")

    inserted_vote = get_first_row(insert_result) or vote_payload
    try:
        supabase.rpc("recalculate_claim_vote_scores", {"target_claim_id": normalized_claim_id}).execute()
    except Exception as error:
        print("[claims/vote] score refresh warning:", str(error), flush=True)

    topic_cluster_id = str(claim.get("topic_cluster_id") or "")
    if topic_cluster_id:
        update_cluster_stats(topic_cluster_id)

    refreshed_claim = fetch_claim_row(normalized_claim_id) or claim
    return {"ok": True, "vote": inserted_vote, "claim": refreshed_claim}


# SINGLE WRITE PATH (step 3) — server-side report + evidence writes. These were
# the last client-only rulebooks; every rule the mobile client enforced now runs
# here (JWT auth, enum/length validation, suspension, dedup, per-day limits) so
# mobile and the future web share one implementation. Reuses existing services
# (get_active_write_profile, score_source_url) — no duplicated logic.
ALLOWED_REPORT_REASONS = {
    "SPAM", "FAKE_SOURCE", "DUPLICATE_CLAIM", "HARMFUL_CONTENT", "MISLEADING_TITLE",
    "HARASSMENT_OR_ABUSE", "MISINFORMATION_ABUSE", "EXPLICIT_CONTENT",
    "MALICIOUS_EVIDENCE", "OTHER",
}
REPORT_NOTE_MAX_LENGTH = 1000
REPORTS_PER_DAY_LIMIT = 20

ALLOWED_EVIDENCE_TYPES = {"SUPPORTS_TRUE", "SUPPORTS_FAKE", "ADDS_CONTEXT", "UNCLEAR"}
EVIDENCE_NOTE_MIN_LENGTH = 10
EVIDENCE_NOTE_MAX_LENGTH = 500
_EVIDENCE_DOMAIN_PATTERN = re.compile(r"^(?:[a-z0-9-]+\.)+[a-z]{2,}$", re.IGNORECASE)


def _normalize_evidence_type(raw: str) -> str:
    normalized = (raw or "").strip().upper().replace(" ", "_")
    if normalized in ("SUPPORTS_TRUE", "TRUE"):
        return "SUPPORTS_TRUE"
    if normalized in ("SUPPORTS_FAKE", "FAKE"):
        return "SUPPORTS_FAKE"
    if normalized in ("ADDS_CONTEXT", "CONTEXT", "ADDS"):
        return "ADDS_CONTEXT"
    if normalized == "UNCLEAR":
        return "UNCLEAR"
    return "ADDS_CONTEXT"


def _normalize_evidence_url(raw: str) -> str:
    trimmed = (raw or "").strip()
    if not trimmed:
        return ""
    if not re.match(r"^https?://", trimmed, re.IGNORECASE):
        return f"https://{trimmed}"
    return trimmed


def _is_valid_evidence_url(raw: str) -> bool:
    # Mirrors the client's isValidEvidenceUrl (evidenceService.ts / claim/[id].tsx):
    # http/https only, real hostname, no credentials, dotted-domain pattern.
    normalized = _normalize_evidence_url(raw)
    if not re.match(r"^https?://", normalized, re.IGNORECASE):
        return False
    try:
        parsed = urlparse(normalized)
    except Exception:
        return False
    if parsed.scheme not in ("http", "https") or parsed.username or parsed.password:
        return False
    hostname = (parsed.hostname or "").lower()
    if hostname.startswith("www."):
        hostname = hostname[4:]
    if not hostname:
        return False
    return bool(_EVIDENCE_DOMAIN_PATTERN.match(hostname))


def _sanitize_evidence_note(raw: str) -> str:
    return re.sub(r"\s+", " ", (raw or "").replace("<", "").replace(">", "")).strip()


@app.post("/api/claims/{claim_id}/report", status_code=201)
def api_report_claim(claim_id: str, payload: ClaimReportRequest, request: Request):
    user_id = get_authenticated_user_id(request)
    enforce_rate_limit(
        request, "claims_report", DUPLICATE_CHECK_RATE_LIMIT_MAX_REQUESTS, AI_RATE_LIMIT_WINDOW_SECONDS
    )
    normalized_claim_id = claim_id.strip()
    if not is_uuid(normalized_claim_id):
        raise HTTPException(status_code=404, detail="Claim not found.")

    reason = (payload.reason or "").strip().upper()
    if reason not in ALLOWED_REPORT_REASONS:
        raise HTTPException(status_code=422, detail="Invalid report reason.")

    note = (payload.note or "").replace("<", "").replace(">", "").strip()
    if len(note) > REPORT_NOTE_MAX_LENGTH:
        raise HTTPException(status_code=422, detail=f"Note must be {REPORT_NOTE_MAX_LENGTH} characters or fewer.")

    supabase = get_supabase_client()
    get_active_write_profile(supabase, user_id, "reporting")

    claim = fetch_claim_row(normalized_claim_id)
    if not claim or claim.get("is_deleted"):
        raise HTTPException(status_code=404, detail="Claim not found.")

    since = (datetime.now(timezone.utc) - timedelta(hours=24)).isoformat()
    try:
        recent = (
            supabase.table("reports").select("id", count="exact")
            .eq("user_id", user_id).gte("created_at", since).execute()
        )
        recent_count = getattr(recent, "count", None)
        if recent_count is None:
            recent_count = len(getattr(recent, "data", None) or [])
    except Exception as error:
        print("[claims/report] rate-limit query failed:", str(error), flush=True)
        recent_count = 0
    if int(recent_count or 0) >= REPORTS_PER_DAY_LIMIT:
        raise HTTPException(status_code=429, detail="Too many reports today. Please try again later.")

    existing = (
        supabase.table("reports").select("id")
        .eq("target_type", "CLAIM").eq("claim_id", normalized_claim_id).eq("user_id", user_id)
        .limit(1).execute()
    )
    if get_first_row(existing):
        return JSONResponse(
            status_code=409,
            content={"ok": False, "already_reported": True, "message": "You already reported this claim."},
        )

    insert_payload = {
        "target_type": "CLAIM",
        "claim_id": normalized_claim_id,
        "evidence_id": None,
        "profile_id": None,
        "user_id": user_id,
        "reason": reason,
        "note": note or None,
        "status": "OPEN",
    }
    try:
        insert_result = supabase.table("reports").insert(insert_payload).execute()
    except Exception as error:
        message = str(error).lower()
        if "23505" in message or "duplicate" in message or "unique" in message:
            return JSONResponse(
                status_code=409,
                content={"ok": False, "already_reported": True, "message": "You already reported this claim."},
            )
        print("[claims/report] insert failed:", str(error), flush=True)
        raise HTTPException(status_code=500, detail="Could not submit report right now.")

    created = get_first_row(insert_result) or insert_payload
    try:
        supabase.rpc("recalculate_claim_report_count", {"target_claim_id": normalized_claim_id}).execute()
    except Exception as error:
        print("[claims/report] count refresh warning:", str(error), flush=True)

    refreshed_claim = fetch_claim_row(normalized_claim_id) or claim
    return {"ok": True, "report": created, "claim": refreshed_claim}


@app.post("/api/claims/{claim_id}/evidence", status_code=201)
def api_add_evidence(claim_id: str, payload: ClaimEvidenceRequest, request: Request):
    user_id = get_authenticated_user_id(request)
    enforce_rate_limit(
        request, "claims_evidence", DUPLICATE_CHECK_RATE_LIMIT_MAX_REQUESTS, AI_RATE_LIMIT_WINDOW_SECONDS
    )
    normalized_claim_id = claim_id.strip()
    if not is_uuid(normalized_claim_id):
        raise HTTPException(status_code=404, detail="Claim not found.")

    if not (payload.url or "").strip():
        raise HTTPException(status_code=422, detail="Evidence URL is required.")
    if not _is_valid_evidence_url(payload.url):
        raise HTTPException(status_code=422, detail="Please check the source URL.")

    note = _sanitize_evidence_note(payload.note)
    if not note:
        raise HTTPException(status_code=422, detail="Evidence note is required.")
    if len(note) < EVIDENCE_NOTE_MIN_LENGTH:
        raise HTTPException(status_code=422, detail="Short note must be at least 10 characters.")
    if len(note) > EVIDENCE_NOTE_MAX_LENGTH:
        raise HTTPException(status_code=422, detail=f"Evidence note must be {EVIDENCE_NOTE_MAX_LENGTH} characters or fewer.")

    evidence_type = _normalize_evidence_type(payload.evidence_type)
    normalized_url = _normalize_evidence_url(payload.url)

    supabase = get_supabase_client()
    get_active_write_profile(supabase, user_id, "adding evidence")

    claim = fetch_claim_row(normalized_claim_id)
    if not claim or claim.get("is_deleted"):
        raise HTTPException(status_code=404, detail="Claim not found.")

    score = score_source_url(normalized_url) or {}
    insert_payload = {
        "claim_id": normalized_claim_id,
        "user_id": user_id,
        "evidence_type": evidence_type,
        "url": normalized_url,
        "note": note,
        "source_quality_label": score.get("source_trust_label") or score.get("source_quality"),
        "source_quality_score": score.get("source_score"),
        "source_quality_reason": (score.get("source_message") or {}).get("text") or score.get("source_domain"),
    }
    try:
        insert_result = supabase.table("evidence").insert(insert_payload).execute()
    except Exception as error:
        print("[claims/evidence] insert failed:", str(error), flush=True)
        raise HTTPException(status_code=500, detail="Could not save evidence right now.")

    created = get_first_row(insert_result)
    if not created:
        raise HTTPException(status_code=500, detail="Could not save evidence right now.")

    try:
        supabase.rpc("recalculate_claim_evidence_count", {"target_claim_id": normalized_claim_id}).execute()
    except Exception as error:
        print("[claims/evidence] count refresh warning:", str(error), flush=True)

    return {"ok": True, "evidence": created}


# MODERATION CHECK (NEW) — dedicated endpoint with a single, stable response
# contract for the Create Claim screen's moderation state machine. It wraps the
# same check_content_safety() classifier used by /api/claims/safety-check, but
# returns the {ok, allowed, flagged, category, message} shape the client expects.
#
# FAIL-OPEN (product decision): the deterministic layer inside check_content_safety
# ALWAYS runs and blocks targeted threats regardless of OpenAI availability; only
# the OpenAI moderation/semantic layers fail open when the API is unreachable, so
# a classifier outage never blocks legitimate posting (Render free-tier cold
# starts + active Apple review). Public endpoint; authentication is optional and
# used only to associate moderation logs with a user when a bearer token exists.
MODERATION_TITLE_MAX = 300
MODERATION_DESCRIPTION_MAX = 2000
MODERATION_VIOLENCE_MESSAGE = "This claim may contain violent or threatening language. Please rewrite it before posting."
MODERATION_GENERIC_MESSAGE = "This content violates our community guidelines and cannot be posted."


def _moderation_category_label(verdict: dict) -> str:
    """Map an internal safety verdict to a stable public category label."""
    raw = f"{verdict.get('category', '')} {verdict.get('matched_layer', '')} {verdict.get('reason', '')}".lower()
    if "violence" in raw or "threat" in raw or "kill" in raw or "indirect_violence" in raw:
        return "THREATENING_VIOLENCE"
    if "hate" in raw:
        return "HATE"
    if "harassment" in raw:
        return "HARASSMENT"
    if "sexual" in raw:
        return "SEXUAL"
    if "self_harm" in raw or "self-harm" in raw:
        return "SELF_HARM"
    if "spam" in raw:
        return "SPAM"
    return "OBJECTIONABLE"


def get_optional_authenticated_user_id(request: Request) -> str | None:
    authorization = request.headers.get("authorization", "")

    if not authorization.lower().startswith("bearer "):
        return None

    try:
        return get_authenticated_user_id(request)
    except HTTPException:
        return None


@app.post("/moderation/check")
def moderation_check(payload: ContentSafetyRequest, request: Request):
    """Classify a claim's SAFETY (not truth/politics) before it is saved.

    Request:  {"title": "...", "description": "optional"}
    Allowed:  200 {"ok": true,  "allowed": true,  "flagged": false, "category": null, "message": null}
    Blocked:  200 {"ok": true,  "allowed": false, "flagged": true,  "category": "THREATENING_VIOLENCE", "message": "..."}
    Bad input: 400 {"ok": false, "code": "INVALID_INPUT", "message": "..."}
    """
    enforce_rate_limit(
        request,
        "moderation_check",
        DUPLICATE_CHECK_RATE_LIMIT_MAX_REQUESTS,
        AI_RATE_LIMIT_WINDOW_SECONDS,
    )
    user_id = get_optional_authenticated_user_id(request)

    title = str(payload.title or "").strip()
    description = str(payload.description or "").strip()

    if not title:
        return JSONResponse(
            status_code=400,
            content={"ok": False, "code": "INVALID_INPUT", "message": "A claim title is required."},
        )
    if len(title) > MODERATION_TITLE_MAX:
        return JSONResponse(
            status_code=400,
            content={"ok": False, "code": "INVALID_INPUT", "message": "Claim title is too long."},
        )
    if len(description) > MODERATION_DESCRIPTION_MAX:
        return JSONResponse(
            status_code=400,
            content={"ok": False, "code": "INVALID_INPUT", "message": "Claim description is too long."},
        )

    verdict = check_content_safety(title, description)
    allowed = bool(verdict.get("safe", True))
    print(
        f"[moderation-check] user={user_id} title_length={len(title)} "
        f"description_length={len(description)} allowed={allowed} "
        f"layer={verdict.get('matched_layer', '')} category={verdict.get('category', '')}",
        flush=True,
    )

    if allowed:
        return {"ok": True, "allowed": True, "flagged": False, "category": None, "message": None}

    category = _moderation_category_label(verdict)
    message = MODERATION_VIOLENCE_MESSAGE if category == "THREATENING_VIOLENCE" else MODERATION_GENERIC_MESSAGE
    log_content_safety_block(user_id, title, category, message)

    return {
        "ok": True,
        "allowed": False,
        "flagged": True,
        "category": category,
        "message": message,
    }


# ═══════════════════════════════════════════════════════════════════════════
# SERVER-SIDE SAFETY GATE — Layers 1+2 via database webhook (Part 1c).
# ═══════════════════════════════════════════════════════════════════════════
# Claims are inserted client-side (supabase-js). A Supabase Database Webhook
# fires INSERT on public.claims -> POST /internal/safety-check with the
# X-Safety-Secret header. This runs OpenAI moderation + gpt-4.1-mini semantic
# (classify_for_gate — the SAME layer helpers as /api/content/check-safety, but
# fail-CLOSED) and writes safety_status APPROVED/BLOCKED. Layer 0 (the Postgres
# trigger in migration 048) has already stamped hard blocklist/regex hits BLOCKED
# before we ever see the row, so this only decides rows that arrive PENDING.
#
# The RLS policy "Only approved claims visible" (048) is the real enforcement:
# a PENDING or BLOCKED claim is invisible to everyone except its author/admins
# regardless of what any client does. This endpoint just resolves PENDING -> a
# final state; it can never make unsafe content visible.
#
# /internal/safety-sweep re-checks PENDING claims older than 2 minutes so a
# missed or failed webhook self-heals (driven by a Render cron).
SAFETY_BLOCKED_NOTIFICATION_TITLE = "Claim removed"
SAFETY_SWEEP_MIN_AGE_SECONDS = 120
SAFETY_SWEEP_BATCH_SIZE = 50


def _require_safety_secret(request: Request) -> None:
    """Guard the internal safety endpoints with the shared webhook secret."""
    expected = os.environ.get("SAFETY_WEBHOOK_SECRET", "")
    provided = request.headers.get("x-safety-secret", "")

    if not expected or not hmac.compare_digest(provided, expected):
        raise HTTPException(status_code=401, detail="Unauthorized safety request.")


def _notify_claim_blocked(supabase: Any, author_id: str | None, claim_id: str | None, category: str) -> None:
    """Tell the author their claim was removed. Never fails the caller."""
    if not author_id:
        return

    label = (category or "objectionable content").strip() or "objectionable content"
    try:
        supabase.table("notifications").insert({
            "user_id": author_id,
            "type": "claim_blocked",
            "title": SAFETY_BLOCKED_NOTIFICATION_TITLE,
            "body": f"Your claim was removed for violating community guidelines: {label}.",
            "claim_id": claim_id,
        }).execute()
    except Exception as error:
        print(f"[safety-gate] blocked notification failed claim={claim_id}: {error}", flush=True)


def _process_claim_safety(supabase: Any, record: dict) -> dict:
    """Resolve one claim row's safety_status. Idempotent, never raises.

    - APPROVED already: nothing to do.
    - BLOCKED already (Layer 0 trigger): ensure the author is notified once.
    - PENDING: run the AI layers; write APPROVED/BLOCKED, or leave PENDING on
      an OpenAI error (fail-CLOSED — the sweep retries; we never approve blind).
    """
    claim_id = str(record.get("id") or "").strip()
    author_id = record.get("author_id")
    current = str(record.get("safety_status") or "PENDING").upper()

    if not claim_id:
        return {"ok": False, "reason": "missing_claim_id"}

    if current == "APPROVED":
        return {"ok": True, "claim_id": claim_id, "status": "APPROVED", "action": "skipped"}

    if current == "BLOCKED":
        # Layer 0 trigger already blocked it — still notify the author once
        # (the webhook fires exactly once per INSERT, so no duplicate).
        _notify_claim_blocked(supabase, author_id, claim_id, str(record.get("safety_category") or "violence"))
        return {"ok": True, "claim_id": claim_id, "status": "BLOCKED", "action": "notified"}

    verdict = classify_for_gate(
        str(record.get("title") or ""),
        str(record.get("description") or ""),
        log_id=claim_id,
    )
    decision = verdict.get("decision", "PENDING")

    if decision == "PENDING":
        # Could not verify (OpenAI down/errored). Leave PENDING for the sweep.
        print(f"[safety-gate] left PENDING claim={claim_id} reason={verdict.get('reason', '')}", flush=True)
        return {"ok": True, "claim_id": claim_id, "status": "PENDING", "action": "deferred"}

    category = str(verdict.get("category") or "")
    try:
        supabase.table("claims").update({
            "safety_status": decision,
            "safety_category": category or None,
            "safety_checked_at": datetime.now(timezone.utc).isoformat(),
        }).eq("id", claim_id).eq("safety_status", "PENDING").execute()
    except Exception as error:
        print(f"[safety-gate] update failed claim={claim_id}: {error}", flush=True)
        return {"ok": False, "claim_id": claim_id, "reason": "update_failed"}

    if decision == "BLOCKED":
        _notify_claim_blocked(supabase, author_id, claim_id, category)

    print(f"[safety-gate] claim={claim_id} -> {decision} layer={verdict.get('matched_layer', '')}", flush=True)
    return {"ok": True, "claim_id": claim_id, "status": decision, "action": "updated"}


@app.post("/internal/safety-check")
async def internal_safety_check(request: Request):
    """Supabase Database Webhook target — INSERT on public.claims.

    Body (Supabase supafunc webhook shape):
      {"type":"INSERT","table":"claims","record":{...},"old_record":null,...}
    We also accept a bare claim row for manual retries.
    """
    _require_safety_secret(request)

    try:
        body = await request.json()
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid JSON body.")

    record = body.get("record") if isinstance(body, dict) else None
    if not isinstance(record, dict):
        record = body if isinstance(body, dict) else {}

    supabase = get_supabase_client()
    result = _process_claim_safety(supabase, record)
    return {"ok": bool(result.get("ok")), **result}


@app.post("/internal/safety-sweep")
def internal_safety_sweep(request: Request):
    """Re-check PENDING claims older than 2 minutes (missed/failed webhooks).

    Driven by a Render cron. Self-heals any claim a webhook never resolved.
    """
    _require_safety_secret(request)

    supabase = get_supabase_client()
    cutoff = (datetime.now(timezone.utc) - timedelta(seconds=SAFETY_SWEEP_MIN_AGE_SECONDS)).isoformat()

    try:
        query = (
            supabase.table("claims")
            .select("id,title,description,author_id,safety_status,safety_category,created_at")
            .eq("safety_status", "PENDING")
            .lt("created_at", cutoff)
            .order("created_at", desc=False)
            .limit(SAFETY_SWEEP_BATCH_SIZE)
        )
        rows = query.execute().data or []
    except Exception as error:
        print(f"[safety-sweep] query failed: {error}", flush=True)
        raise HTTPException(status_code=500, detail="Sweep query failed.")

    results = {"APPROVED": 0, "BLOCKED": 0, "PENDING": 0}
    for record in rows:
        outcome = _process_claim_safety(supabase, record)
        status = str(outcome.get("status") or "PENDING")
        results[status] = results.get(status, 0) + 1

    print(f"[safety-sweep] checked={len(rows)} results={results}", flush=True)
    return {"ok": True, "checked": len(rows), "results": results}


# PHASE 6 STEP 2 — Offline reference citations.
# Evidence types accepted, mirroring the existing evidence table CHECK constraint.
CITATION_EVIDENCE_TYPES = {"SUPPORTS_TRUE", "SUPPORTS_FAKE", "ADDS_CONTEXT", "UNCLEAR"}
CITATION_DISPUTE_MIN_REASON_CHARS = 20


def fetch_evidence_row(evidence_id: str) -> "dict | None":
    """Fetch a single evidence row by id, or None. New helper (URL path untouched)."""
    supabase = get_supabase_client()
    response = supabase.table("evidence").select("*").eq("id", evidence_id).limit(1).execute()
    rows = response.data or []

    if isinstance(rows, dict):
        return rows

    return rows[0] if rows else None


@app.post("/api/evidence/citation")
def api_evidence_citation(payload: CitationEvidenceRequest, request: Request):
    """Submit an OFFLINE citation as evidence (book/newspaper/journal/document).

    WHY a new endpoint rather than the client-side Supabase insert used for URL
    evidence: offline citations need server-only work — schema validation,
    external existence checks, and consistent trust weighting — that cannot run
    on the client. URL evidence is completely unchanged; it still goes through the
    existing client path. Verification is fail-open: it never blocks submission.
    """
    enforce_rate_limit(request, "evidence_citation", AI_RATE_LIMIT_MAX_REQUESTS, AI_RATE_LIMIT_WINDOW_SECONDS)
    user_id = get_authenticated_user_id(request)

    claim_id = payload.claim_id.strip()
    reference_type = payload.reference_type.strip().lower()
    evidence_type = payload.evidence_type.strip().upper()

    if not claim_id:
        raise HTTPException(status_code=400, detail="claim_id is required")

    if evidence_type not in CITATION_EVIDENCE_TYPES:
        raise HTTPException(status_code=400, detail="Invalid evidence_type.")

    # This endpoint is offline-only; 'url' evidence must keep using the existing path.
    if not reference_type or reference_type == "url":
        raise HTTPException(
            status_code=400,
            detail="reference_type must be one of: book, newspaper, journal, document.",
        )

    # 1. Shape validation — reject missing required fields with a clear message.
    is_valid, validation_error = validate_citation(reference_type, payload.citation)
    if not is_valid:
        raise HTTPException(status_code=400, detail=validation_error)

    # 2. Existence check (fail-open: None on any timeout/failure).
    citation_verified = verify_citation_exists(reference_type, payload.citation)

    # 3. Trust weighting for the offline source (never touches URL scoring).
    scored = score_citation_source(reference_type, payload.citation, citation_verified)

    # NOTE: evidence.url is NOT NULL (pre-existing constraint we do not alter), so
    # offline citations store an empty string for url. Flagged in the migration.
    insert_payload = {
        "claim_id": claim_id,
        "user_id": user_id,
        "evidence_type": evidence_type,
        "url": "",
        "note": payload.note.strip(),
        "reference_type": reference_type,
        "citation": payload.citation,
        "citation_verified": citation_verified,
        "source_quality_label": scored["source_quality_label"],
        "source_quality_score": scored["source_quality_score"],
        "source_quality_reason": scored["source_quality_reason"],
    }

    supabase = get_supabase_client()
    result = supabase.table("evidence").insert(insert_payload).execute()
    inserted = get_first_row(result)

    # Keep claims.evidence_count in sync, exactly like the URL client path does.
    try:
        supabase.rpc("recalculate_claim_evidence_count", {"target_claim_id": claim_id}).execute()
    except Exception as error:
        print(f"[evidence/citation] evidence count recalc failed: {error}", flush=True)

    return {
        "ok": True,
        "evidence": inserted,
        "citation_verified": citation_verified,
        "source_quality": scored["source_quality"],
        "source_quality_score": scored["source_quality_score"],
        "source_quality_reason": scored["source_quality_reason"],
    }


@app.post("/api/evidence/{evidence_id}/dispute")
def api_evidence_dispute(evidence_id: str, payload: CitationDisputeRequest, request: Request):
    """Open a community dispute against an offline citation.

    WHY the guard rails: disputes are a trust mechanism, so they must be
    meaningful (>= 20 chars), targeted at disputable content (offline citations,
    not URLs which are self-verifiable), and not self-serving (you can't dispute
    your own citation). One-per-user is enforced by the DB unique constraint.
    """
    enforce_rate_limit(request, "evidence_dispute", AI_RATE_LIMIT_MAX_REQUESTS, AI_RATE_LIMIT_WINDOW_SECONDS)
    user_id = get_authenticated_user_id(request)

    reason = payload.reason.strip()
    if len(reason) < CITATION_DISPUTE_MIN_REASON_CHARS:
        raise HTTPException(
            status_code=400,
            detail=f"Dispute reason must be at least {CITATION_DISPUTE_MIN_REASON_CHARS} characters.",
        )

    evidence = fetch_evidence_row(evidence_id.strip())
    if not evidence:
        raise HTTPException(status_code=404, detail="Evidence not found.")

    if str(evidence.get("reference_type") or "url").strip().lower() == "url":
        raise HTTPException(status_code=400, detail="URL evidence cannot be disputed as a citation.")

    if str(evidence.get("user_id") or "") == user_id:
        raise HTTPException(status_code=403, detail="You cannot dispute your own citation.")

    supabase = get_supabase_client()

    try:
        result = (
            supabase.table("citation_disputes")
            .insert(
                {
                    "evidence_id": evidence_id,
                    "disputer_id": user_id,
                    "reason": reason,
                }
            )
            .execute()
        )
    except Exception as error:
        # The unique (evidence_id, disputer_id) constraint surfaces here as a
        # duplicate-key error; translate it into a clean 409.
        message = str(error).lower()
        if "duplicate" in message or "unique" in message or "23505" in message:
            raise HTTPException(status_code=409, detail="You have already disputed this citation.")
        print(f"[evidence/dispute] insert failed: {error}", flush=True)
        raise HTTPException(status_code=500, detail="Could not record dispute.")

    inserted = get_first_row(result)
    return {"ok": True, "dispute": inserted}


@app.get("/api/evidence/{evidence_id}/disputes")
def api_evidence_disputes(evidence_id: str, request: Request):
    """List disputes (with status) for a given evidence row.

    Read-only view backing the community dispute UI and manual admin resolution.
    """
    enforce_rate_limit(request, "evidence_disputes_list", SOURCE_SCORE_RATE_LIMIT_MAX_REQUESTS, AI_RATE_LIMIT_WINDOW_SECONDS)
    get_authenticated_user_id(request)

    supabase = get_supabase_client()
    response = (
        supabase.table("citation_disputes")
        .select("id, evidence_id, disputer_id, reason, status, created_at")
        .eq("evidence_id", evidence_id.strip())
        .order("created_at", desc=True)
        .execute()
    )
    disputes = response.data or []
    if isinstance(disputes, dict):
        disputes = [disputes]

    return {"disputes": disputes}


# PHASE 4 STEP 9
@app.get("/ai/source-score")
def ai_source_score(request: Request, url: str = ""):
    # PHASE 4 STEP 27
    enforce_rate_limit(request, "ai_source_score", SOURCE_SCORE_RATE_LIMIT_MAX_REQUESTS, AI_RATE_LIMIT_WINDOW_SECONDS)
    source_metadata = score_source_url(url)
    log_source_score("ai/source-score", source_metadata)
    return source_metadata


def run_claim_ai_precheck_pipeline(claim_id: str, authenticated_user_id: str) -> dict:
    claim = fetch_claim_row(claim_id)

    if not claim:
        raise HTTPException(status_code=404, detail="Claim not found")

    payload = build_precheck_payload_from_claim(claim)
    print("[ai/precheck] authenticated user:", authenticated_user_id, flush=True)
    # PHASE 4 STEP 9
    source_metadata = score_source_url(payload.source_url)
    log_source_score("ai/precheck", source_metadata)
    # PHASE 4 STEP 21
    source_page = fetch_source_page(payload.source_url)
    log_source_page("ai/precheck", source_page)
    ai_result = analyze_claim_with_openai(
        title=payload.title,
        description=payload.description,
        source_url=payload.source_url,
        category=payload.category,
        source_metadata=source_metadata,
        source_page=source_page,
    )
    print("[ai/precheck] called", flush=True)
    print(f"[ai/precheck] OPENAI_MODEL={get_openai_model()}", flush=True)
    print(f"[ai/precheck] claim_id={payload.claim_id}", flush=True)
    print(f"[ai/precheck] source_url={payload.source_url}", flush=True)
    # PHASE 4 STEP 7
    print("[ai] claim_type:", ai_result.get("claim_type"), flush=True)
    print(f"[ai/precheck] ai_status={ai_result['ai_status']}", flush=True)
    print(f"[ai/precheck] ai_confidence={ai_result['ai_confidence']}", flush=True)
    print(f"[ai/precheck] source_quality={ai_result['source_quality']}", flush=True)
    print("[ai/precheck] OpenAI analysis completed", flush=True)
    update_result = update_claim_ai_fields(payload.claim_id, ai_result, "ai/precheck")

    if not update_result.get("ok"):
        print(f"[ai/precheck] Supabase update failure: {update_result.get('error')}", flush=True)
        return build_safe_ai_precheck_failure(payload.claim_id)

    print("[ai/precheck] Supabase update success", flush=True)

    # PHASE 6 STEP 3 — SEO metadata trigger (NEW, additive). Claims are
    # created client-side (services/claimService.ts), and /ai/precheck is the
    # backend call that fires right after creation, so the 'creation' SEO
    # version is generated here. Synchronous because the codebase has no
    # fire-and-forget pattern; generate_claim_seo never raises, so the
    # existing precheck response is never blocked or altered by SEO failures.
    try:
        generate_claim_seo(
            claim_id=payload.claim_id,
            title=payload.title,
            description=payload.description,
            category=payload.category,
            verdict=None,
            version="creation",
        )
    except Exception as seo_error:  # Defensive belt-and-suspenders.
        print("[ai/precheck] SEO generation failed (non-fatal):", str(seo_error), flush=True)

    return build_ai_precheck_response(payload.claim_id, update_result)


@app.post("/ai/precheck", response_model=AiPrecheckResponse, response_model_exclude_none=True)
def ai_precheck(payload: AiPrecheckRequest, request: Request):
    claim_id = payload.claim_id.strip()

    if not claim_id:
        raise HTTPException(status_code=400, detail="claim_id is required")

    enforce_rate_limit(request, "ai_precheck", AI_RATE_LIMIT_MAX_REQUESTS, AI_RATE_LIMIT_WINDOW_SECONDS)
    authenticated_user_id = get_authenticated_user_id(request)
    enforce_claim_ai_cooldown(claim_id)
    return run_claim_ai_precheck_pipeline(claim_id, authenticated_user_id)


def run_claim_post_insert_tasks(claim_id: str, author_id: str) -> None:
    """Non-blocking enrichment shared by every endpoint-created claim."""
    try:
        store_claim_embedding(claim_id)
        claim = fetch_claim_row(claim_id)
        if claim:
            cluster_id = str(claim.get("topic_cluster_id") or "")
            if cluster_id:
                update_cluster_stats(cluster_id)
            else:
                find_or_create_topic_cluster(
                    claim_id=claim_id,
                    title=str(claim.get("title") or ""),
                    description=str(claim.get("description") or ""),
                    category=str(claim.get("category") or ""),
                )
    except Exception as error:
        print(f"[claims/create] embedding/topic background warning: {error}", flush=True)

    try:
        run_claim_ai_precheck_pipeline(claim_id, author_id)
    except Exception as error:
        print(f"[claims/create] AI/SEO background warning: {error}", flush=True)
        try:
            claim = fetch_claim_row(claim_id)
            if claim:
                generate_claim_seo(
                    claim_id=claim_id,
                    title=str(claim.get("title") or ""),
                    description=str(claim.get("description") or ""),
                    category=str(claim.get("category") or ""),
                    verdict=None,
                    version="creation",
                )
        except Exception as seo_error:
            print(f"[claims/create] SEO background warning: {seo_error}", flush=True)


# PHASE 4 STEP 3
@app.post("/ai/precheck/retry", response_model=AiPrecheckResponse, response_model_exclude_none=True)
def retry_ai_precheck(payload: AiPrecheckRetryRequest, request: Request):
    claim_id = payload.claim_id.strip()

    if not claim_id:
        raise HTTPException(status_code=400, detail="claim_id is required")

    # PHASE 4 STEP 27
    enforce_rate_limit(request, "ai_precheck_retry", AI_RATE_LIMIT_MAX_REQUESTS, AI_RATE_LIMIT_WINDOW_SECONDS)
    authenticated_user_id = get_authenticated_user_id(request)
    enforce_claim_ai_cooldown(claim_id)
    print("[ai/precheck/retry] called", flush=True)
    print(f"[ai/precheck/retry] OPENAI_MODEL={get_openai_model()}", flush=True)
    print(f"[ai/precheck/retry] claim_id={claim_id}", flush=True)
    print("[ai/precheck/retry] authenticated user:", authenticated_user_id, flush=True)

    try:
        claim = fetch_claim_row(claim_id)

        if not claim:
            raise HTTPException(status_code=404, detail="Claim not found")

        previous_status = claim.get("ai_status") or "PENDING"
        print(f"[ai/precheck/retry] previous_ai_status={previous_status}", flush=True)

        retry_payload = build_precheck_payload_from_claim(claim)
        # PHASE 4 STEP 9
        source_metadata = score_source_url(retry_payload.source_url)
        log_source_score("ai/precheck/retry", source_metadata)
        # PHASE 4 STEP 21
        source_page = fetch_source_page(retry_payload.source_url)
        log_source_page("ai/precheck/retry", source_page)
        # PHASE 4 STEP 10
        evidence_rows = fetch_evidence_rows(claim_id)
        ai_result = analyze_claim_with_openai(
            title=retry_payload.title,
            description=retry_payload.description,
            source_url=retry_payload.source_url,
            category=retry_payload.category,
            source_metadata=source_metadata,
            source_page=source_page,
            evidence_rows=evidence_rows,
        )
        # PHASE 4 STEP 7
        print("[ai] claim_type:", ai_result.get("claim_type"), flush=True)
        print(f"[ai/precheck/retry] new_ai_status={ai_result['ai_status']}", flush=True)
        print(f"[ai/precheck/retry] ai_confidence={ai_result['ai_confidence']}", flush=True)
        print(f"[ai/precheck/retry] source_quality={ai_result['source_quality']}", flush=True)
        print("[ai/precheck/retry] OpenAI analysis completed", flush=True)
        update_result = update_claim_ai_fields(claim_id, ai_result, "ai/precheck/retry")

        if not update_result.get("ok"):
            print(f"[ai/precheck/retry] Supabase update failure: {update_result.get('error')}", flush=True)
            return build_safe_ai_precheck_failure(claim_id)

        print("[ai/precheck/retry] Supabase update success", flush=True)
        return build_ai_precheck_response(claim_id, update_result)
    except HTTPException:
        raise
    except Exception as error:
        print(f"[ai/precheck/retry] failure: {error}", flush=True)
        error_update = mark_claim_ai_error(claim_id)

        if error_update:
            print(f"[ai/precheck/retry] Supabase update failure: {error_update}", flush=True)
        else:
            print("[ai/precheck/retry] Supabase ERROR status update success", flush=True)

        error_analysis = build_ai_error_analysis()
        return {
            "ok": False,
            "claim_id": claim_id,
            **error_analysis,
            "error": "AI pre-check failed. Please retry.",
        }


# ============================================================
# PHASE 6 STEP 3 — NEW endpoints (additive; nothing above changed)
# ============================================================


def parse_supabase_timestamp(value: Any) -> datetime | None:
    """Parse a Supabase timestamptz string into an aware UTC datetime."""
    text = str(value or "").strip()

    if not text:
        return None

    try:
        parsed = datetime.fromisoformat(text.replace("Z", "+00:00"))
    except ValueError:
        return None

    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=timezone.utc)

    return parsed


# PHASE 6 STEP 3 — FEATURE 1: claim deletion, 3-hour window.
@app.delete("/api/claims/{claim_id}")
def delete_own_claim(claim_id: str, request: Request):
    """Author-only claim deletion within 3 hours of posting.

    After 3 hours OR after finalization (whichever comes first) the claim is
    permanent — no one can delete it through this endpoint, admins included
    (the separate POST /admin/claims/delete moderation path is unchanged).

    Inspection notes:
    - "Finalized" in this codebase is claims.verdict_calculated_at (set by
      finalize_expired_claim / claimService.ts), so it plays the spec's
      finalized_at role; no new column was added.
    - Existing deletions are HARD deletes (admin endpoint and the client's
      claimService.deleteClaim); matched exactly. Child rows (votes,
      evidence, reports, mention_tags, claim_seo) cascade in the DB.

    Frontend should show a countdown timer on the user's own claims for the
    first 3 hours. After 3 hours or finalization, hide the delete option
    entirely.
    """
    normalized_claim_id = claim_id.strip()

    if not is_uuid(normalized_claim_id):
        raise HTTPException(status_code=404, detail="Claim not found")

    authenticated_user_id = get_authenticated_user_id(request)
    claim = fetch_claim_row(normalized_claim_id)

    if not claim:
        raise HTTPException(status_code=404, detail="Claim not found")

    # Ownership check — same pattern as the RLS policy (auth.uid() = author_id).
    if str(claim.get("author_id") or "") != authenticated_user_id:
        raise HTTPException(status_code=403, detail="You can only remove your own claims.")

    if claim.get("verdict_calculated_at") is not None:
        raise HTTPException(status_code=403, detail="Claim is finalized and permanent.")

    created_at = parse_supabase_timestamp(claim.get("created_at"))

    # An unreadable created_at means the window cannot be proven still open,
    # so the claim is treated as permanent (fail closed).
    if created_at is None:
        raise HTTPException(status_code=403, detail="Claim can only be removed within 3 hours of posting.")

    age = datetime.now(timezone.utc) - created_at

    if age > timedelta(hours=3):
        raise HTTPException(status_code=403, detail="Claim can only be removed within 3 hours of posting.")

    # Hard delete — matches the existing admin deletion behavior exactly;
    # the 3-hour window above is the guard.
    try:
        supabase = get_supabase_client()
        supabase.table("claims").delete().eq("id", normalized_claim_id).execute()
    except Exception as error:
        print("[claims delete] failed:", normalized_claim_id, str(error), flush=True)
        raise HTTPException(status_code=500, detail="Could not remove the claim right now.")

    print("[claims delete] removed claim:", normalized_claim_id, "by:", authenticated_user_id, flush=True)
    return {"ok": True, "message": "Claim removed."}


# PHASE 6 STEP 3 — FEATURE 2: auth-level PII scrub + session invalidation,
# called from DELETE /account. Follows the existing GoTrue admin REST
# pattern (see fetch_auth_users_for_username_check): there is no stored
# session table in this codebase — auth is stateless Supabase JWTs — so
# "invalidate all sessions" means revoking every refresh token via the
# admin logout endpoint; outstanding access tokens then lapse at expiry.
def scrub_auth_user_and_sign_out(user_id: str) -> None:
    supabase_url = os.environ.get("SUPABASE_URL", "").rstrip("/")
    service_role_key = os.environ.get("SUPABASE_SERVICE_ROLE_KEY", "")

    if not supabase_url or not service_role_key:
        raise RuntimeError("Missing SUPABASE_URL or SUPABASE_SERVICE_ROLE_KEY on backend.")

    headers = {
        "apikey": service_role_key,
        "authorization": f"Bearer {service_role_key}",
        "content-type": "application/json",
    }
    encoded_user_id = quote(str(user_id), safe="")

    # a. Replace the auth email with a non-deliverable placeholder and clear
    #    metadata (username/display name copies live there). profiles has no
    #    email column, so this is the only place email exists.
    update_body = json.dumps({
        "email": f"deleted+{str(user_id).replace('-', '')[:16]}@deleted.invalid",
        "user_metadata": {},
    }).encode("utf-8")
    update_request = UrlRequest(
        f"{supabase_url}/auth/v1/admin/users/{encoded_user_id}",
        data=update_body,
        headers=headers,
        method="PUT",
    )

    with urlopen(update_request, timeout=10):
        pass

    # b. Revoke all refresh tokens for this user (global sign-out).
    logout_request = UrlRequest(
        f"{supabase_url}/auth/v1/admin/users/{encoded_user_id}/logout",
        data=b"",
        headers=headers,
        method="POST",
    )

    with urlopen(logout_request, timeout=10):
        pass

    print("[account delete] auth PII scrubbed and sessions revoked:", user_id, flush=True)


# PHASE 6 STEP 3 — FEATURE 3: SEO metadata endpoint.
def get_claim_final_verdict_label(claim: dict) -> str | None:
    """Human-readable verdict once a claim is finalized, else None."""
    if not claim.get("verdict_calculated_at"):
        return None

    status = str(claim.get("status") or "").strip().upper()
    labels = {
        "COMMUNITY_TRUE": "True",
        "COMMUNITY_FAKE": "Fake",
        "NEEDS_MORE_EVIDENCE": "Unverified - needs more evidence",
    }
    return labels.get(status, status or None)


@app.get("/api/claims/{claim_id}/seo")
def api_claim_seo(claim_id: str):
    """Public SEO metadata for a claim (no auth — crawlers must reach it).

    Returns the latest claim_seo row: the 'finalization' version when the
    claim has finalized, otherwise the 'creation' version.

    Finalization trigger note: claims finalize client-side
    (services/claimService.ts finalizeExpiredClaim / the DB RPC), so there is
    no backend code path that sets verdict_calculated_at to hook into.
    Instead the 'finalization' SEO version is generated lazily on first fetch
    after finalization — same net effect (verdict-aware SEO exists once the
    verdict is public), no change to the existing finalize flow. The
    'creation' version is likewise backfilled here if the /ai/precheck
    trigger never ran for an old claim.
    """
    normalized_claim_id = claim_id.strip()

    if not is_uuid(normalized_claim_id):
        raise HTTPException(status_code=404, detail="Claim not found")

    claim = fetch_claim_row(normalized_claim_id)

    if not claim or claim.get("hidden"):
        raise HTTPException(status_code=404, detail="Claim not found")

    seo_row = fetch_latest_claim_seo(normalized_claim_id)
    final_verdict = get_claim_final_verdict_label(claim)
    needs_finalization_version = final_verdict is not None and (
        seo_row is None or seo_row.get("version") != "finalization"
    )

    if needs_finalization_version or seo_row is None:
        generate_claim_seo(
            claim_id=normalized_claim_id,
            title=str(claim.get("title") or ""),
            description=str(claim.get("description") or ""),
            category=str(claim.get("category") or ""),
            verdict=final_verdict,
            version="finalization" if needs_finalization_version else "creation",
        )
        seo_row = fetch_latest_claim_seo(normalized_claim_id)

    if seo_row is None:
        raise HTTPException(status_code=404, detail="SEO metadata is not available for this claim yet.")

    return {"ok": True, "claim_id": normalized_claim_id, "seo": seo_row}


# ============================================================
# PHASE 6 STEP 4 — Topic clustering endpoints (NEW, additive)
# ============================================================
#
# VOTE-HOOK NOTE (Feature: cluster stats on vote): the spec asks to wire
# update_cluster_stats into "the vote recording endpoint in main.py". No such
# endpoint exists — votes are inserted client-side straight into Supabase
# (services/voteService.ts -> supabase.from("votes").insert), with DB triggers
# maintaining the per-claim totals. Cluster stats are therefore refreshed
# LAZILY on read below (search results and the topic detail endpoint), which
# keeps them eventually consistent without touching the existing vote flow.

TOPIC_SEARCH_SIMILARITY_THRESHOLD = 0.65
TOPIC_SEARCH_CLUSTER_LIMIT = 10
TOPIC_SEARCH_PREVIEW_CLAIMS = 3
TOPIC_SEARCH_INDIVIDUAL_LIMIT = 20
TOPIC_CLAIMS_DEFAULT_LIMIT = 20
TOPIC_CLAIMS_MAX_LIMIT = 50

# Public author-card fields, mirroring what the client's own profile join
# selects (services/claimService.ts CLAIM_PROFILE_SAFE_SELECT-style subset) so
# claim rows returned here render authors through the existing mapper.
TOPIC_AUTHOR_PROFILE_SELECT = (
    "id,username,display_name,avatar_url,verified,reputation_score,trust_score,"
    "rank_title,highest_rank_achieved,badge_list,is_deleted,deleted_at,created_at"
)


def attach_author_profiles_to_claim_rows(rows: list) -> list:
    """Embed each claim row's author profile under row['profiles'].

    The frontend claim mapper (claimService.ts mapAuthor/getEmbeddedProfile)
    reads the author from row.profiles; the client normally merges this itself,
    so backend-served claim lists must do the same for authors to render.
    Fails soft: rows come back unmodified if the profile fetch errors."""
    author_ids = sorted({str(row.get("author_id") or "") for row in rows if row.get("author_id")})

    if not author_ids:
        return rows

    try:
        profiles_response = (
            get_supabase_client()
            .table("profiles")
            .select(TOPIC_AUTHOR_PROFILE_SELECT)
            .in_("id", author_ids)
            .execute()
        )
        profiles_by_id = {str(profile.get("id")): profile for profile in (profiles_response.data or [])}
    except Exception as error:
        print(f"[topics] author profile fetch failed: {error}", flush=True)
        return rows

    for row in rows:
        row["profiles"] = profiles_by_id.get(str(row.get("author_id") or ""))

    return rows


def strip_claim_row_for_api(row: dict) -> dict:
    """Drop the raw embedding vector — large and useless to clients."""
    safe_row = dict(row)
    safe_row.pop("embedding", None)
    return safe_row


@app.get("/api/topics/search")
def api_topics_search(request: Request, q: str = ""):
    """Two-layer public search: topic clusters first, individual claims after.

    Public (no auth) — Google must be able to index topic cluster results.
    Layer 1: clusters by embedding similarity to the query (> 0.65, top 10),
             each with up to 3 preview claims ordered by vote count.
    Layer 2: text-search fallback over claims with NO cluster.
             INSPECTION NOTE: the existing search (claimService.ts
             searchClaimsPage) is a client-side substring filter over
             title/description/source_url/category — there is no backend text
             search to reuse, so this is its server-side equivalent (ilike over
             title/description), returning raw claim rows in the same shape the
             existing claim list endpoints/queries produce.
    """
    enforce_rate_limit(request, "topics_search", SOURCE_SCORE_RATE_LIMIT_MAX_REQUESTS, AI_RATE_LIMIT_WINDOW_SECONDS)
    query = (q or "").strip()

    if len(query) < 2:
        return {"topics": [], "individual_claims": []}

    topics = []
    embedding = generate_claim_embedding(title=query, description="")

    if embedding is not None:
        try:
            supabase = get_supabase_client()
            match_response = supabase.rpc(
                "match_claim_topics",
                {
                    "query_embedding": embedding,
                    "match_threshold": TOPIC_SEARCH_SIMILARITY_THRESHOLD,
                    "match_count": TOPIC_SEARCH_CLUSTER_LIMIT,
                },
            ).execute()
            matched_clusters = match_response.data or []

            for cluster in matched_clusters:
                cluster_id = str(cluster.get("id") or "")
                # Lazy stats refresh (see VOTE-HOOK NOTE above), then re-read.
                update_cluster_stats(cluster_id)
                topic_row = fetch_topic_row(cluster_id)

                if not topic_row:
                    continue

                preview_rows = []
                try:
                    preview_response = (
                        supabase.table("claims")
                        .select("id,title,author_id,votes_true,votes_fake,total_votes,status")
                        .eq("topic_cluster_id", cluster_id)
                        .order("total_votes", desc=True)
                        .limit(TOPIC_SEARCH_PREVIEW_CLAIMS)
                        .execute()
                    )
                    preview_rows = attach_author_profiles_to_claim_rows(preview_response.data or [])
                except Exception as preview_error:
                    print(f"[topics/search] preview fetch failed: {preview_error}", flush=True)

                preview_claims = []
                for preview in preview_rows:
                    profile = preview.get("profiles") or {}
                    is_deleted_author = bool(profile.get("is_deleted"))
                    preview_claims.append(
                        {
                            "claim_id": str(preview.get("id") or ""),
                            "title": str(preview.get("title") or ""),
                            "author_display_name": (
                                "Deleted User"
                                if is_deleted_author
                                else str(profile.get("display_name") or profile.get("username") or "Verifact contributor")
                            ),
                            "true_votes": int(preview.get("votes_true") or 0),
                            "fake_votes": int(preview.get("votes_fake") or 0),
                            "verdict_status": str(preview.get("status") or ""),
                        }
                    )

                topics.append(
                    {
                        "topic_cluster_id": cluster_id,
                        "topic_label": str(topic_row.get("topic_label") or ""),
                        "slug": str(topic_row.get("slug") or ""),
                        "cluster_verdict": str(topic_row.get("cluster_verdict") or "INSUFFICIENT_DATA"),
                        "total_true_votes": int(topic_row.get("total_true_votes") or 0),
                        "total_fake_votes": int(topic_row.get("total_fake_votes") or 0),
                        "total_vote_count": int(topic_row.get("total_vote_count") or 0),
                        "claim_count": int(topic_row.get("claim_count") or 0),
                        "preview_claims": preview_claims,
                    }
                )
        except Exception as error:
            print(f"[topics/search] cluster search failed: {error}", flush=True)
            topics = []

    individual_claims = []
    try:
        # PostgREST or= syntax breaks on commas/parens inside the pattern, so
        # strip them from the user query before building the ilike filter.
        safe_pattern = re.sub(r"[,()]+", " ", query).strip()

        if safe_pattern:
            individual_response = (
                get_supabase_client()
                .table("claims")
                .select("*")
                .is_("topic_cluster_id", "null")
                .or_(f"title.ilike.%{safe_pattern}%,description.ilike.%{safe_pattern}%")
                .order("created_at", desc=True)
                .limit(TOPIC_SEARCH_INDIVIDUAL_LIMIT)
                .execute()
            )
            individual_rows = attach_author_profiles_to_claim_rows(individual_response.data or [])
            individual_claims = [strip_claim_row_for_api(row) for row in individual_rows]
    except Exception as error:
        print(f"[topics/search] individual claim search failed: {error}", flush=True)
        individual_claims = []

    return {"topics": topics, "individual_claims": individual_claims}


@app.get("/api/topics/{cluster_id}/claims")
def api_topic_claims(cluster_id: str, request: Request, limit: int = TOPIC_CLAIMS_DEFAULT_LIMIT, offset: int = 0):
    """All claims in a topic cluster, paginated. Public (no auth) — Google
    must be able to index topic cluster pages."""
    enforce_rate_limit(request, "topic_claims", SOURCE_SCORE_RATE_LIMIT_MAX_REQUESTS, AI_RATE_LIMIT_WINDOW_SECONDS)
    normalized_cluster_id = cluster_id.strip()

    if not is_uuid(normalized_cluster_id):
        raise HTTPException(status_code=404, detail="Topic not found")

    safe_limit = max(1, min(TOPIC_CLAIMS_MAX_LIMIT, int(limit or TOPIC_CLAIMS_DEFAULT_LIMIT)))
    safe_offset = max(0, int(offset or 0))

    # Lazy stats refresh BEFORE reading the topic row (see VOTE-HOOK NOTE):
    # votes land client-side, so this read is where cluster totals catch up.
    update_cluster_stats(normalized_cluster_id)
    topic = fetch_topic_row(normalized_cluster_id)

    if not topic:
        raise HTTPException(status_code=404, detail="Topic not found")

    try:
        claims_response = (
            get_supabase_client()
            .table("claims")
            .select("*", count="exact")
            .eq("topic_cluster_id", normalized_cluster_id)
            .order("created_at", desc=True)
            .range(safe_offset, safe_offset + safe_limit - 1)
            .execute()
        )
        claim_rows = attach_author_profiles_to_claim_rows(claims_response.data or [])
        total = int(claims_response.count or len(claim_rows))
    except Exception as error:
        print(f"[topics/claims] fetch failed: {error}", flush=True)
        raise HTTPException(status_code=500, detail="Could not load topic claims right now.")

    return {
        "topic": topic,
        "claims": [strip_claim_row_for_api(row) for row in claim_rows],
        "total": total,
        "offset": safe_offset,
        "limit": safe_limit,
    }
