from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import load_workbook


DEFAULT_INPUT = Path("app_moderation_blocklist.xlsx")
DEFAULT_OUTPUT = Path("backend/data/moderation_blocklist.json")

REQUIRED_COLUMNS = {
    "category": "category",
    "subcategory": "subcategory",
    "flagged keyword / phrase": "phrase",
    "severity level": "severity",
    "recommended system action": "action",
}


def normalize_header(value: object) -> str:
    return " ".join(str(value or "").strip().lower().split())


def cell_text(value: object) -> str:
    return str(value or "").strip()


def read_entries(input_path: Path) -> list[dict[str, str]]:
    workbook = load_workbook(input_path, data_only=True, read_only=True)
    worksheet = workbook.active

    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError(f"{input_path} is empty")

    header_lookup = {normalize_header(value): index for index, value in enumerate(rows[0])}
    missing = [label for label in REQUIRED_COLUMNS if label not in header_lookup]
    if missing:
        raise ValueError(f"{input_path} is missing columns: {', '.join(missing)}")

    entries: list[dict[str, str]] = []
    for row in rows[1:]:
        phrase = cell_text(row[header_lookup["flagged keyword / phrase"]])
        if not phrase:
            continue

        entries.append(
            {
                "phrase": phrase,
                "category": cell_text(row[header_lookup["category"]]),
                "subcategory": cell_text(row[header_lookup["subcategory"]]),
                "severity": cell_text(row[header_lookup["severity level"]]),
                "action": cell_text(row[header_lookup["recommended system action"]]),
            }
        )

    return entries


def main() -> int:
    input_path = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_INPUT
    output_path = Path(sys.argv[2]) if len(sys.argv) > 2 else DEFAULT_OUTPUT

    if not input_path.exists():
        raise FileNotFoundError(
            f"Blocklist workbook not found: {input_path}. "
            "Pass the xlsx path as the first argument."
        )

    entries = read_entries(input_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(
        json.dumps(entries, indent=2, ensure_ascii=False) + "\n",
        encoding="utf-8",
    )
    print(f"Wrote {len(entries)} entries to {output_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
